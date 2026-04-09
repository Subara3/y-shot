"""
y-shot: Web Screenshot Automation Tool  v2.3 (Flet)
  - v1.5: highlight fix, abort, tel capture, input check fix
  - v1.6: popup menu, reorder fix, pattern count sync, modal dialogs
  - v1.7: dropdown page selector, 1-column test list, start_number per page,
           pattern numbering in filenames/UI/Excel, manual number override,
           fullshot (CDP full-page capture)
  - v2.2: highlight self-correction, element browser width/scroll,
           F5 refresh step, page duplicate/reorder,
           step/pattern copy-paste, taskbar icon fix
  - v2.3: alert OK/cancel (confirm dialog handling),
           step delete confirmation setting, snackbar fix
  - v2.6: POST値キャプチャモード(スクショmode=post),
           hidden radio/checkbox のlabel経由クリック,
           テストケース追加/削除時のUI更新エラー修正
"""

import csv, os, sys, json, threading, time, logging, traceback, copy, shutil
# Set AppUserModelID so Windows taskbar shows the exe icon, not the Flet client icon
try:
    import ctypes
    ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID("yshot.app")
except Exception:
    pass
from datetime import datetime
from urllib.parse import urlparse, urlunparse
import flet as ft

APP_NAME = "y-shot"
APP_VERSION = "3.1"
APP_AUTHOR = "Yuri Norimatsu"

# ── Constants ──
LOG_MAX_LINES = 400
SAVE_DELAY_SEC = 2.0
BANK_MAX_URLS = 50
WIN_CREATE_NO_WINDOW = 0x08000000  # subprocess: hide console window on Windows

# ── File logger (log/YYYYMMDD.log, append) ──
def _setup_file_logger():
    _log_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)) if not getattr(sys, 'frozen', False)
                            else os.path.dirname(sys.executable), "log")
    os.makedirs(_log_dir, exist_ok=True)
    _log_file = os.path.join(_log_dir, datetime.now().strftime("%Y%m%d") + ".log")
    logger = logging.getLogger("y-shot")
    logger.setLevel(logging.DEBUG)
    if not logger.handlers:
        fh = logging.FileHandler(_log_file, encoding="utf-8")
        fh.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(message)s", datefmt="%H:%M:%S"))
        logger.addHandler(fh)
    return logger
_flog = _setup_file_logger()

import re
def _safe_filename(s, max_len=30):
    """Remove characters unsafe for filenames and truncate."""
    s = re.sub(r'[\\/:*?"<>|\x00-\x1f]', '_', s).strip().strip('.')
    return s[:max_len] if s else '_'

def _has_non_bmp(s):
    """Check if string contains characters outside the Basic Multilingual Plane (e.g. emoji, 𠮷)."""
    return any(ord(c) > 0xFFFF for c in s)

# ===================================================================
# Backend
# ===================================================================

def _sel_by(selector):
    """Return (By.XPATH, selector) if selector starts with '//' or '(/', else (By.CSS_SELECTOR, selector)."""
    from selenium.webdriver.common.by import By
    if selector.startswith("//") or selector.startswith("(/"):
        return (By.XPATH, selector)
    return (By.CSS_SELECTOR, selector)

def _safe_float(val, default=1.0):
    """float変換 — 不正値はdefaultにフォールバック"""
    try: return float(val)
    except (ValueError, TypeError): return default

# ── JS-based bulk element collection (replaces per-element round-trips) ──
_JS_COLLECT_ELEMENTS = """
(function(includeHidden) {
    function escAttr(v) { return v.replace(/\\\\/g,'\\\\\\\\').replace(/"/g,'\\\\"'); }
    function isSafeClass(c) {
        if (!c || /^\\d/.test(c)) return false;
        return /^[a-zA-Z0-9_-]+$/.test(c);
    }
    function cnt(sel) { try { return document.querySelectorAll(sel).length; } catch(e) { return 999; } }
    function getVis(el) {
        try {
            var s = window.getComputedStyle(el);
            if (s.display==='none') return [false,'display:none'];
            if (s.visibility==='hidden') return [false,'visibility:hidden'];
            if (parseFloat(s.opacity)===0) return [false,'opacity:0'];
            var r = el.getBoundingClientRect();
            if (r.width===0 && r.height===0) return [false,'size:0'];
            if (r.bottom<0 || r.right<0) return [false,'off-viewport'];
            return [true,''];
        } catch(e) { return [true,'']; }
    }
    function buildSel(el, tag, eid, ename) {
        if (eid) {
            if (/^\\d/.test(eid) || !/^[a-zA-Z0-9_-]+$/.test(eid))
                return '[id="'+escAttr(eid)+'"]';
            return '#'+eid;
        }
        var ta=['data-testid','data-cy','data-test'];
        for (var i=0;i<ta.length;i++) {
            var tv=el.getAttribute(ta[i]);
            if (tv) { var s='['+ta[i]+'="'+escAttr(tv)+'"]'; if (cnt(s)===1) return s; }
        }
        var aria=el.getAttribute('aria-label');
        if (aria) { var s=tag+'[aria-label="'+escAttr(aria)+'"]'; if (cnt(s)===1) return s; }
        if (ename) { var s=tag+'[name="'+escAttr(ename)+'"]'; if (cnt(s)===1) return s; }
        var etype=el.getAttribute('type')||'';
        if (etype && ename) {
            var s=tag+'[type="'+etype+'"][name="'+escAttr(ename)+'"]';
            if (cnt(s)===1) return s;
            if (etype==='checkbox'||etype==='radio') {
                var val=el.getAttribute('value')||'';
                if (val) { var vs=tag+'[type="'+etype+'"][name="'+escAttr(ename)+'"][value="'+escAttr(val)+'"]'; if (cnt(vs)>=1) return vs; }
            }
        }
        var cls=(el.getAttribute('class')||'').trim();
        if (cls) {
            var parts=cls.split(/\\s+/).slice(0,3).filter(isSafeClass);
            if (parts.length) {
                var cs=tag+'.'+parts.slice(0,2).join('.');
                var cn=cnt(cs);
                if (cn===1) return cs;
                if (cn<=3 && parts.length>=2) return cs;
            }
        }
        // nth-of-type with parent ID
        var p=el.parentElement;
        if (p) {
            var sibs=p.children, idx=0;
            for (var j=0;j<sibs.length;j++) { if (sibs[j].tagName===el.tagName) idx++; if (sibs[j]===el) break; }
            if (idx>0) {
                var pid=p.id||'';
                if (pid) {
                    if (/^\\d/.test(pid)||!/^[a-zA-Z0-9_-]+$/.test(pid))
                        return '[id="'+escAttr(pid)+'"] > '+tag+':nth-of-type('+idx+')';
                    return '#'+pid+' > '+tag+':nth-of-type('+idx+')';
                }
            }
        }
        // ancestor ID path: build child-combinator path from nearest ancestor with ID
        var anc=p;
        while (anc && anc!==document.body && anc!==document.documentElement) {
            if (anc.id) {
                var ancSel=/^\\d/.test(anc.id)||!/^[a-zA-Z0-9_-]+$/.test(anc.id)
                    ? '[id="'+escAttr(anc.id)+'"]' : '#'+anc.id;
                // Simple descendant (if unique)
                var descendant=ancSel+' '+tag;
                if (cnt(descendant)===1) return descendant;
                // Build child-combinator path from ancestor to element
                var pathParts=[];
                var cur=el;
                var tooDeep=false;
                while (cur && cur!==anc) {
                    var pp=cur.parentElement;
                    if (!pp) { tooDeep=true; break; }
                    var ci=1;
                    for (var si=0;si<pp.children.length;si++) {
                        if (pp.children[si]===cur) break;
                        ci++;
                    }
                    var ctag=cur.tagName.toLowerCase();
                    pathParts.unshift(ctag+':nth-child('+ci+')');
                    cur=pp;
                    if (pathParts.length>6) { tooDeep=true; break; }
                }
                if (!tooDeep && pathParts.length>0) {
                    var fullSel=ancSel+' > '+pathParts.join(' > ');
                    try { if (cnt(fullSel)===1) return fullSel; } catch(e) {}
                }
                break;
            }
            anc=anc.parentElement;
        }
        // XPath position-based (e.g. (//img)[2] — useful for verification of img resources)
        var allSameTag=document.querySelectorAll(tag);
        if (allSameTag.length>1) {
            for (var ni=0;ni<allSameTag.length;ni++) {
                if (allSameTag[ni]===el) {
                    var xp='(//'+tag+')['+(ni+1)+']';
                    try { var xr=document.evaluate(xp,document,null,9,null); if (xr.singleNodeValue===el) return xp; } catch(e){}
                    break;
                }
            }
        }
        // img: use alt or src as last resort
        if (tag==='img') {
            var alt=el.getAttribute('alt');
            if (alt) { var s='img[alt="'+escAttr(alt)+'"]'; if (cnt(s)===1) return s; }
            var src=el.getAttribute('src');
            if (src) {
                var s='img[src="'+escAttr(src)+'"]'; if (cnt(s)===1) return s;
                var fname=src.split('/').pop().split('\\\\').pop().split('?')[0];
                if (fname) { var s2='img[src*="'+escAttr(fname)+'"]'; if (cnt(s2)===1) return s2; }
            }
        }
        // XPath fallback: use text content to build a unique XPath
        var txt=(el.textContent||'').trim();
        if (txt && txt.length<=60 && txt.indexOf("'")<0) {
            if (txt.indexOf("\\n")>=0 || txt.length>30) {
                var shortTxt=txt.replace(/\\s+/g,' ').substring(0,20);
                var xp='//'+tag+"[contains(normalize-space(),'"+shortTxt+"')]";
                try { var xr=document.evaluate(xp,document,null,7,null); if (xr.snapshotLength===1) return xp; } catch(e){}
            } else {
                var xp='//'+tag+"[normalize-space()='"+txt+"']";
                try { var xr=document.evaluate(xp,document,null,7,null); if (xr.snapshotLength===1) return xp; } catch(e){}
            }
        }
        // XPath by @value (for input[type=submit/button] etc.)
        var val=el.getAttribute('value');
        if (val && val.indexOf("'")<0 && (tag==='input'||tag==='button')) {
            var xp='//'+tag+"[@value='"+val+"']";
            try { var xr=document.evaluate(xp,document,null,7,null); if (xr.snapshotLength===1) return xp; } catch(e){}
        }
        // aria-label XPath (may have been non-unique as CSS, try XPath)
        var ariaL=el.getAttribute('aria-label');
        if (ariaL && ariaL.indexOf("'")<0) {
            var xp='//'+tag+"[@aria-label='"+ariaL+"']";
            try { var xr=document.evaluate(xp,document,null,7,null); if (xr.snapshotLength===1) return xp; } catch(e){}
        }
        return tag;
    }
    function getHint(el, tag, etype) {
        var h=el.getAttribute('placeholder')
            ||el.getAttribute('alt')
            ||el.getAttribute('title')
            ||el.getAttribute('aria-label')
            ||(el.textContent||'').trim().substring(0,50)
            ||(el.getAttribute('value')||'').substring(0,30)
            ||'';
        if (!h && tag==='a') {
            var href=el.getAttribute('href')||'';
            if (href) { var q=href.indexOf('?'); h=(q>=0?href.substring(0,q):href).slice(-50); }
        }
        if (!h && tag==='img') {
            var src=el.getAttribute('src')||'';
            if (src) { var q=src.indexOf('?'); h=(q>=0?src.substring(0,q):src).slice(-50); }
        }
        if ((!h||h===(el.getAttribute('value')||'')) && tag==='input' && (etype==='checkbox'||etype==='radio')) {
            var lbl=el.closest('label');
            if (lbl) { var lt=lbl.textContent.trim(); if (lt) h=lt.substring(0,50); }
            if (!h && el.id) {
                var fl=document.querySelector('label[for="'+el.id+'"]');
                if (fl) h=fl.textContent.trim().substring(0,50);
            }
        }
        return h;
    }
    var CSS="input,textarea,select,button,a,label,[role='button'],[type='submit'],"
           +"[type='image'],img,[onclick],li[id],span[id],div[onclick],"
           +"[class*='btn'],[class*='Btn'],[class*='button'],[class*='Button']";
    var els; try { els=document.querySelectorAll(CSS); } catch(e) { return []; }
    var results=[], seen={};
    for (var i=0;i<els.length;i++) {
        try {
            var el=els[i];
            var etype=el.getAttribute('type')||'';
            var vis=getVis(el);
            if (!vis[0]) {
                if (etype!=='radio' && etype!=='checkbox') { if (!includeHidden) continue; }
            }
            var tag=el.tagName.toLowerCase();
            if (etype==='hidden') continue;
            var eid=el.getAttribute('id')||'';
            var ename=el.getAttribute('name')||'';
            var sel=buildSel(el,tag,eid,ename);
            if (seen[sel]) continue; seen[sel]=true;
            // Collect extra metadata for test case creation
            var meta={};
            if (tag==='input'||tag==='textarea'||tag==='select') {
                if (el.hasAttribute('required')) meta.required=true;
                var ml=el.getAttribute('maxlength'); if (ml) meta.maxlength=parseInt(ml);
                var pt=el.getAttribute('pattern'); if (pt) meta.pattern=pt;
                if (el.disabled) meta.disabled=true;
                if (el.readOnly) meta.readonly=true;
                if (tag==='select') meta.option_count=el.options?el.options.length:0;
                // Associated label
                var lbl='';
                if (eid) { var le=document.querySelector('label[for="'+eid+'"]'); if (le) lbl=le.textContent.trim().substring(0,40); }
                if (!lbl) { var cl=el.closest('label'); if (cl) lbl=cl.textContent.trim().substring(0,40); }
                if (!lbl && el.closest('div,li,td')) {
                    var pp=el.closest('div,li,td');
                    for (var ci=0;ci<pp.childNodes.length;ci++) {
                        if (pp.childNodes[ci].nodeType===3) { var t=pp.childNodes[ci].textContent.trim(); if(t){lbl=t.substring(0,40);break;} }
                    }
                }
                if (lbl) meta.label=lbl;
            }
            if (tag==='a') { var hr=el.getAttribute('href'); if (hr) meta.href=hr.substring(0,80); }
            results.push({selector:sel,tag:tag,type:etype,name:ename,id:eid,
                          hint:getHint(el,tag,etype),visible:vis[0],hidden_reason:vis[1],meta:meta});
        } catch(e) { continue; }
    }
    return results;
})(arguments[0]);
"""

def collect_elements_js(driver, include_hidden=False):
    """Collect interactive elements using a single JS call (fast)."""
    try:
        results = driver.execute_script("return " + _JS_COLLECT_ELEMENTS.strip(), include_hidden)
        return results if results else []
    except Exception as e:
        _flog.warning(f"JS collect failed, falling back to Python: {e}")
        return collect_elements_python(driver, include_hidden)

def collect_elements_python(driver, include_hidden=False):
    from selenium.webdriver.common.by import By
    results, seen = [], set()
    css = ("input, textarea, select, button, a, label, [role='button'], [type='submit'], "
           "[type='image'], img, [onclick], li[id], span[id], div[onclick], "
           "[class*='btn'], [class*='Btn'], [class*='button'], [class*='Button']")
    try: elements = driver.find_elements(By.CSS_SELECTOR, css)
    except Exception: return results
    # Pass 1: collect basic info
    hidden_els = []  # (index_in_results, selenium_element)
    for el in elements:
        try:
            visible = el.is_displayed()
            etype = el.get_attribute("type") or ""
            if not visible:
                if (etype or "") not in ("radio", "checkbox"):
                    if not include_hidden:
                        continue
            tag = el.tag_name.lower()
            if etype == "hidden": continue
            eid = el.get_attribute("id") or ""; ename = el.get_attribute("name") or ""
            sel = _build_selector(driver, el, tag, eid, ename)
            if sel in seen: continue; seen.add(sel)
            hint = (el.get_attribute("placeholder") or el.get_attribute("alt") or
                    el.get_attribute("title") or
                    (el.text or "").strip()[:50] or
                    (el.get_attribute("textContent") or "").strip()[:50] or
                    (el.get_attribute("value") or "")[:30])
            if not hint and tag == "a":
                href = el.get_attribute("href") or ""
                if href: hint = href.split("?")[0][-50:]
            if not hint and tag == "img":
                src = el.get_attribute("src") or ""
                if src: hint = src.split("?")[0][-50:]
            if (not hint or hint == (el.get_attribute("value") or "")) and tag == "input" and etype in ("checkbox", "radio"):
                try:
                    label_text = driver.execute_script(
                        "var e=arguments[0];"
                        "var p=e.closest('label');"
                        "if(p){var c=p.textContent.trim();if(c)return c;}"
                        "var id=e.id;"
                        "if(id){var l=document.querySelector('label[for=\"'+id+'\"]');"
                        "if(l)return l.textContent.trim();}"
                        "return '';", el) or ""
                    if label_text: hint = label_text[:50]
                except Exception: pass
            entry = {"selector": sel, "tag": tag, "type": etype, "name": ename, "id": eid, "hint": hint, "visible": visible, "hidden_reason": ""}
            if not visible:
                hidden_els.append((len(results), el))
            results.append(entry)
        except Exception: continue
    # Pass 2: batch JS for hidden reasons (single call instead of per-element)
    if hidden_els:
        batch_els = [el for _, el in hidden_els]
        try:
            reasons = driver.execute_script(
                "var results=[];"
                "for(var i=0;i<arguments.length;i++){"
                "  var e=arguments[i];"
                "  try{var s=window.getComputedStyle(e);"
                "  if(s.display==='none'){results.push('display:none');continue;}"
                "  if(s.visibility==='hidden'){results.push('visibility:hidden');continue;}"
                "  if(parseFloat(s.opacity)===0){results.push('opacity:0');continue;}"
                "  var r=e.getBoundingClientRect();"
                "  if(r.width===0&&r.height===0){results.push('size:0');continue;}"
                "  if(r.bottom<0||r.right<0){results.push('off-viewport');continue;}"
                "  results.push('other');}catch(x){results.push('unknown');}"
                "}return results;", *batch_els) or []
            for j, (idx, _) in enumerate(hidden_els):
                if j < len(reasons):
                    results[idx]["hidden_reason"] = reasons[j]
        except Exception:
            for idx, _ in hidden_els:
                results[idx]["hidden_reason"] = "unknown"
    return results

def _css_escape_attr(val):
    return val.replace('\\', '\\\\').replace('"', '\\"')

def _is_safe_class(cls):
    if not cls: return False
    if cls[0].isdigit(): return False
    return all(c.isalnum() or c in '-_' for c in cls)

def _build_selector(driver, el, tag, eid, ename):
    from selenium.webdriver.common.by import By
    if eid:
        if eid[0].isdigit() or not all(c.isalnum() or c in '-_' for c in eid):
            return f'[id="{_css_escape_attr(eid)}"]'
        return f"#{eid}"
    # Prefer stable test attributes over name/class
    for attr in ("data-testid", "data-cy", "data-test"):
        val = el.get_attribute(attr) or ""
        if val:
            s = f'[{attr}="{_css_escape_attr(val)}"]'
            try:
                if len(driver.find_elements(By.CSS_SELECTOR, s)) == 1: return s
            except Exception: pass
    aria = el.get_attribute("aria-label") or ""
    if aria:
        s = f'{tag}[aria-label="{_css_escape_attr(aria)}"]'
        try:
            if len(driver.find_elements(By.CSS_SELECTOR, s)) == 1: return s
        except Exception: pass
    if ename:
        safe_name = _css_escape_attr(ename)
        s = f'{tag}[name="{safe_name}"]'
        try:
            if len(driver.find_elements(By.CSS_SELECTOR, s)) == 1: return s
        except Exception: pass
    etype = el.get_attribute("type") or ""
    if etype and ename:
        safe_name = _css_escape_attr(ename)
        s = f'{tag}[type="{etype}"][name="{safe_name}"]'
        try:
            if len(driver.find_elements(By.CSS_SELECTOR, s)) == 1: return s
        except Exception: pass
        # checkbox/radio: disambiguate by value (type+name+value is semantically unique)
        if etype in ("checkbox", "radio"):
            val = el.get_attribute("value") or ""
            if val:
                s = f'{tag}[type="{etype}"][name="{safe_name}"][value="{_css_escape_attr(val)}"]'
                try:
                    if driver.find_elements(By.CSS_SELECTOR, s): return s
                except Exception: pass
    classes = (el.get_attribute("class") or "").strip()
    if classes:
        safe_classes = [c for c in classes.split()[:3] if _is_safe_class(c)]
        if safe_classes:
            cs = tag + "".join(f".{c}" for c in safe_classes[:2])
            try:
                n = len(driver.find_elements(By.CSS_SELECTOR, cs))
                if n == 1: return cs
                # Allow small duplicates for multi-class selectors (e.g. PC/SP dual layout)
                if n <= 3 and len(safe_classes) >= 2: return cs
            except Exception: pass
    try:
        idx = driver.execute_script(
            "var e=arguments[0],p=e.parentElement;if(!p)return 0;"
            "var s=[];for(var i=0;i<p.children.length;i++)if(p.children[i].tagName===e.tagName)s.push(p.children[i]);"
            "for(var j=0;j<s.length;j++)if(s[j]===e)return j+1;return 0;", el)
        if idx and idx > 0:
            pid = driver.execute_script("var p=arguments[0].parentElement;return p?(p.id||''):'';", el)
            if pid:
                if pid[0].isdigit() or not all(c.isalnum() or c in '-_' for c in pid):
                    return f'[id="{_css_escape_attr(pid)}"] > {tag}:nth-of-type({idx})'
                return f"#{pid} > {tag}:nth-of-type({idx})"
    except Exception: pass
    # XPath position-based (e.g. (//img)[2])
    try:
        all_same = driver.find_elements(By.CSS_SELECTOR, tag)
        if len(all_same) > 1:
            for ni, _el in enumerate(all_same):
                if _el == el:
                    xp = f"(//{tag})[{ni+1}]"
                    found = driver.find_elements(By.XPATH, xp)
                    if len(found) == 1: return xp
                    break
    except Exception: pass
    # img: use alt or src for unique selector (last resort)
    if tag == "img":
        try:
            alt = el.get_attribute("alt") or ""
            if alt:
                s = f'img[alt="{_css_escape_attr(alt)}"]'
                if len(driver.find_elements(By.CSS_SELECTOR, s)) == 1: return s
            src = el.get_attribute("src") or ""
            if src:
                s = f'img[src="{_css_escape_attr(src)}"]'
                if len(driver.find_elements(By.CSS_SELECTOR, s)) == 1: return s
                fname = src.split("/")[-1].split("\\")[-1].split("?")[0]
                if fname:
                    s2 = f'img[src*="{_css_escape_attr(fname)}"]'
                    if len(driver.find_elements(By.CSS_SELECTOR, s2)) == 1: return s2
        except Exception: pass
    # XPath fallback: use text content to build a unique XPath
    try:
        txt = (el.text or "").strip()
        if txt and len(txt) <= 60 and "'" not in txt:
            if "\n" in txt or len(txt) > 30:
                short = " ".join(txt.split())[:20]
                xp = f"//{tag}[contains(normalize-space(),'{short}')]"
            else:
                xp = f"//{tag}[normalize-space()='{txt}']"
            from selenium.webdriver.common.by import By
            if len(driver.find_elements(By.XPATH, xp)) == 1:
                return xp
    except Exception: pass
    # XPath by @value (for input[type=submit/button] etc.)
    try:
        val = el.get_attribute("value") or ""
        if val and "'" not in val and tag in ("input", "button"):
            xp = f"//{tag}[@value='{val}']"
            from selenium.webdriver.common.by import By
            if len(driver.find_elements(By.XPATH, xp)) == 1:
                return xp
    except Exception: pass
    return tag

_JS_CAPTURE_FORM_VALUES = """
(function(){
    // Reuse buildSel from _JS_COLLECT_ELEMENTS (inlined for independence)
    function escAttr(v){return v.replace(/\\\\/g,'\\\\\\\\').replace(/"/g,'\\\\"');}
    function isSafeClass(c){if(!c||/^\\d/.test(c))return false;return /^[a-zA-Z0-9_-]+$/.test(c);}
    function cnt(sel){try{return document.querySelectorAll(sel).length;}catch(e){return 999;}}
    function buildSel(el,tag,eid,ename){
        if(eid){if(/^\\d/.test(eid)||!/^[a-zA-Z0-9_-]+$/.test(eid))return'[id="'+escAttr(eid)+'"]';return'#'+eid;}
        var ta=['data-testid','data-cy','data-test'];
        for(var i=0;i<ta.length;i++){var tv=el.getAttribute(ta[i]);if(tv){var s='['+ta[i]+'="'+escAttr(tv)+'"]';if(cnt(s)===1)return s;}}
        if(ename){var s=tag+'[name="'+escAttr(ename)+'"]';if(cnt(s)===1)return s;}
        var etype=el.getAttribute('type')||'';
        if(etype&&ename){var s=tag+'[type="'+etype+'"][name="'+escAttr(ename)+'"]';if(cnt(s)===1)return s;
            if(etype==='checkbox'||etype==='radio'){var val=el.getAttribute('value')||'';if(val){var vs=tag+'[type="'+etype+'"][name="'+escAttr(ename)+'"][value="'+escAttr(val)+'"]';if(cnt(vs)>=1)return vs;}}}
        var cls=(el.getAttribute('class')||'').trim();
        if(cls){var parts=cls.split(/\\s+/).slice(0,3).filter(isSafeClass);if(parts.length){var cs=tag+'.'+parts.slice(0,2).join('.');if(cnt(cs)<=3)return cs;}}
        var p=el.parentElement;
        if(p){var sibs=p.children,idx=0;for(var j=0;j<sibs.length;j++){if(sibs[j].tagName===el.tagName)idx++;if(sibs[j]===el)break;}
            if(idx>0&&p.id){var pid=p.id;if(/^\\d/.test(pid)||!/^[a-zA-Z0-9_-]+$/.test(pid))return'[id="'+escAttr(pid)+'"] > '+tag+':nth-of-type('+idx+')';return'#'+pid+' > '+tag+':nth-of-type('+idx+')';}}
        return tag;
    }
    var steps=[],seen={};
    var textCss="input[type='text'],input[type='tel'],input[type='email'],input[type='number'],input[type='url'],input[type='search'],input[type='password'],input[type='date'],input[type='time'],input[type='datetime-local'],input[type='month'],input[type='week'],input[type='color'],input[type='range'],input:not([type]),textarea";
    var els=document.querySelectorAll(textCss);
    for(var i=0;i<els.length;i++){try{var el=els[i];var t=(el.getAttribute('type')||'text').toLowerCase();if(t==='hidden')continue;
        var val=el.value||'';if(!val.trim())continue;var tag=el.tagName.toLowerCase();var sel=buildSel(el,tag,el.getAttribute('id')||'',el.getAttribute('name')||'');
        if(seen[sel])continue;seen[sel]=true;steps.push({type:'入力',selector:sel,value:val});}catch(e){continue;}}
    els=document.querySelectorAll('select');
    for(var i=0;i<els.length;i++){try{var el=els[i];var sel=buildSel(el,'select',el.getAttribute('id')||'',el.getAttribute('name')||'');
        if(seen[sel])continue;seen[sel]=true;var val=el.value||'';if(val)steps.push({type:'選択',selector:sel,value:val});}catch(e){continue;}}
    var checks=document.querySelectorAll("input[type='radio']:checked,input[type='checkbox']:checked");
    for(var i=0;i<checks.length;i++){try{var el=checks[i];var sel=buildSel(el,'input',el.getAttribute('id')||'',el.getAttribute('name')||'');
        if(sel==='input'||seen[sel])continue;seen[sel]=true;steps.push({type:'クリック',selector:sel});}catch(e){continue;}}
    return steps;
})();
"""

def capture_form_values(driver):
    """Capture current form values as steps (JS-based, fast)."""
    try:
        results = driver.execute_script("return " + _JS_CAPTURE_FORM_VALUES.strip())
        return results if results else []
    except Exception as e:
        _flog.warning(f"JS capture_form_values failed: {e}")
        return _capture_form_values_python(driver)

def _capture_form_values_python(driver):
    """Fallback: Python-based form value capture."""
    from selenium.webdriver.common.by import By
    steps, seen = [], set()
    text_css = ("input[type='text'],input[type='tel'],input[type='email'],"
                "input[type='number'],input[type='url'],input[type='search'],"
                "input[type='password'],input[type='date'],input[type='time'],"
                "input[type='datetime-local'],input[type='month'],input[type='week'],"
                "input[type='color'],input[type='range'],"
                "input:not([type]),textarea")
    for el in driver.find_elements(By.CSS_SELECTOR, text_css):
        try:
            if (el.get_attribute("type") or "text") == "hidden": continue
            val = el.get_attribute("value") or ""
            if not val.strip(): continue
            tag = el.tag_name.lower()
            sel = _build_selector(driver, el, tag, el.get_attribute("id") or "", el.get_attribute("name") or "")
            if sel in seen: continue; seen.add(sel)
            steps.append({"type": "入力", "selector": sel, "value": val})
        except Exception: continue
    for el in driver.find_elements(By.CSS_SELECTOR, "select"):
        try:
            sel = _build_selector(driver, el, "select", el.get_attribute("id") or "", el.get_attribute("name") or "")
            if sel in seen: continue; seen.add(sel)
            val = el.get_attribute("value") or ""
            if val: steps.append({"type": "選択", "selector": sel, "value": val})
        except Exception: continue
    for css_q in ["input[type='radio']:checked", "input[type='checkbox']:checked"]:
        for el in driver.find_elements(By.CSS_SELECTOR, css_q):
            try:
                eid = el.get_attribute("id") or ""
                ename = el.get_attribute("name") or ""
                sel = _build_selector(driver, el, "input", eid, ename)
                if sel in ("input",) or sel in seen: continue
                seen.add(sel)
                steps.append({"type": "クリック", "selector": sel})
            except Exception: continue
    return steps

def collect_element_options(driver, el_info):
    from selenium.webdriver.common.by import By
    tag = el_info.get("tag", "")
    etype = el_info.get("type", "").lower()
    sel = el_info.get("selector", "")
    if tag == "select":
        try:
            el = driver.find_element(*_sel_by(sel))
            options = el.find_elements(By.TAG_NAME, "option")
            result = []
            for opt in options:
                val = opt.get_attribute("value") or ""
                text = opt.text.strip() or val
                if val or text:
                    result.append({"label": text, "value": val})
            return "選択", result
        except Exception:
            return None, []
    elif etype == "radio":
        try:
            name = el_info.get("name", "") or driver.find_element(*_sel_by(sel)).get_attribute("name") or ""
            if not name: return None, []
            safe_name = _css_escape_attr(name)
            radios = driver.find_elements(By.CSS_SELECTOR, f'input[type="radio"][name="{safe_name}"]')
            result = []
            for r in radios:
                val = r.get_attribute("value") or ""
                rid = r.get_attribute("id") or ""
                label_text = ""
                if rid:
                    try:
                        lbl = driver.find_element(By.CSS_SELECTOR, f'label[for="{_css_escape_attr(rid)}"]')
                        label_text = lbl.text.strip()
                    except Exception: pass
                if not label_text:
                    label_text = val or f"radio_{len(result)+1}"
                radio_sel = f'input[type="radio"][name="{safe_name}"][value="{_css_escape_attr(val)}"]'
                result.append({"label": label_text, "value": radio_sel})
            return "クリック", result
        except Exception:
            return None, []
    return None, []

# Generate XPath for an element (executed via driver.execute_script)
XPATH_JS = """
var el = arguments[0];
if (!el) return '';
if (el.id) return '//*[@id=' + JSON.stringify(el.id) + ']';
var segs = [];
while (el && el.nodeType === 1) {
    var i = 1;
    var sib = el.previousSibling;
    while (sib) {
        if (sib.nodeType === 1 && sib.tagName === el.tagName) i++;
        sib = sib.previousSibling;
    }
    segs.unshift(el.tagName.toLowerCase() + '[' + i + ']');
    el = el.parentNode;
}
return '/' + segs.join('/');
"""

HIGHLIGHT_JS = ("(function(s){try{"
    "var p=document.getElementById('__yshot_hl');if(p)p.remove();"
    "if(window.__yshot_scroll_rm){window.removeEventListener('scroll',window.__yshot_scroll_rm,true);}"
    "var all,cnt;"
    "if(s.substring(0,2)==='//'){"
    "var xr=document.evaluate(s,document,null,7,null);cnt=xr.snapshotLength;"
    "all=[];for(var xi=0;xi<cnt;xi++)all.push(xr.snapshotItem(xi));"
    "}else{all=document.querySelectorAll(s);cnt=all.length;}"
    "if(!cnt)return JSON.stringify({found:0});"
    "var e=all[0];"
    # Override scroll-behavior: smooth to ensure instant scroll
    "var html=document.documentElement;"
    "var origSB=html.style.scrollBehavior;"
    "html.style.setProperty('scroll-behavior','auto','important');"
    "e.scrollIntoView({block:'center',behavior:'instant'});"
    "html.style.scrollBehavior=origSB;"
    # Force reflow so getBoundingClientRect returns post-scroll coords
    "void html.offsetHeight;"
    "var r=e.getBoundingClientRect();"
    "var h=document.createElement('div');h.id='__yshot_hl';"
    "var color=cnt===1?'#FF4444':'#FF8800';"
    "h.style.cssText='position:fixed;border:3px solid '+color+';background:rgba(255,68,68,0.15);"
    "z-index:2147483647;pointer-events:none;border-radius:3px;transition:none;';"
    "h.style.top=r.top-3+'px';h.style.left=r.left-3+'px';"
    "h.style.width=r.width+6+'px';h.style.height=r.height+6+'px';"
    # Append to documentElement to avoid body margin/transform issues
    "html.appendChild(h);"
    # Self-correct: if ancestor transforms shifted fixed positioning, compensate
    "var hr=h.getBoundingClientRect();"
    "var dx=hr.left-(r.left-3),dy=hr.top-(r.top-3);"
    "if(Math.abs(dx)>1||Math.abs(dy)>1){"
    "h.style.left=(r.left-3-dx)+'px';h.style.top=(r.top-3-dy)+'px';}"
    "setTimeout(function(){"
    "window.__yshot_scroll_rm=function(){"
    "var x=document.getElementById('__yshot_hl');if(x)x.remove();"
    "window.removeEventListener('scroll',window.__yshot_scroll_rm,true);};"
    "window.addEventListener('scroll',window.__yshot_scroll_rm,true);"
    "},600);"
    "return JSON.stringify({found:cnt,tag:e.tagName,id:e.id||'',name:e.getAttribute('name')||''});"
    "}catch(x){return JSON.stringify({found:0,error:x.message});}})(arguments[0]);")

JS_SET_VALUE = (
    "(function(el,val){"
    "var proto=el.tagName==='TEXTAREA'"
    "?window.HTMLTextAreaElement.prototype"
    ":window.HTMLInputElement.prototype;"
    "var setter=Object.getOwnPropertyDescriptor(proto,'value');"
    "if(setter&&setter.set){setter.set.call(el,val);}else{el.value=val;}"
    "el.dispatchEvent(new Event('input',{bubbles:true}));"
    "el.dispatchEvent(new Event('change',{bubbles:true}));"
    "})(arguments[0],arguments[1]);"
)

# ── Overlay removal: hide cookie banners, tracking overlays, etc. before screenshot ──
JS_REMOVE_OVERLAYS = """
(function(){
    var removed=0;
    var all=document.querySelectorAll('div,section,aside,dialog,[role="dialog"],[role="alertdialog"]');
    for(var i=0;i<all.length;i++){
        var el=all[i]; var s=window.getComputedStyle(el);
        if(s.position==='fixed'&&s.zIndex&&parseInt(s.zIndex)>999){
            var r=el.getBoundingClientRect();
            if(r.width>window.innerWidth*0.3||r.height>window.innerHeight*0.2){
                var t=(el.textContent||'').toLowerCase();
                if(t.indexOf('cookie')>=0||t.indexOf('privacy')>=0||t.indexOf('consent')>=0
                  ||t.indexOf('accept')>=0||t.indexOf('同意')>=0||t.indexOf('プライバシー')>=0
                  ||el.id.toLowerCase().indexOf('overlay')>=0
                  ||el.id.toLowerCase().indexOf('consent')>=0
                  ||el.id.toLowerCase().indexOf('cookie')>=0
                  ||el.id.toLowerCase().indexOf('privacy')>=0
                  ||(el.className&&el.className.toLowerCase&&(
                    el.className.toLowerCase().indexOf('overlay')>=0
                    ||el.className.toLowerCase().indexOf('consent')>=0
                    ||el.className.toLowerCase().indexOf('cookie')>=0
                    ||el.className.toLowerCase().indexOf('banner')>=0
                    ||el.className.toLowerCase().indexOf('popup')>=0
                    ||el.className.toLowerCase().indexOf('modal')>=0))){
                    el.style.display='none'; removed++;
                }
            }
        }
    }
    // Also hide backdrop overlays (semi-transparent full-screen)
    for(var i=0;i<all.length;i++){
        var el=all[i]; var s=window.getComputedStyle(el);
        if(s.position==='fixed'&&parseFloat(s.opacity)<1&&s.zIndex&&parseInt(s.zIndex)>999){
            var r=el.getBoundingClientRect();
            if(r.width>=window.innerWidth*0.95&&r.height>=window.innerHeight*0.95){
                var bg=s.backgroundColor||'';
                if(bg.indexOf('rgba')>=0||bg.indexOf('0,0,0')>=0||bg==='transparent'){
                    el.style.display='none'; removed++;
                }
            }
        }
    }
    return removed;
})();
"""

# ── Lazy image preload: force-load lazy images by rewriting attributes ──
JS_PRELOAD_LAZY_IMAGES = """
(function(){
    var loaded=0;
    // 1. Force-load images with data-src / data-lazy / loading="lazy"
    var imgs=document.querySelectorAll('img[data-src],img[data-lazy],img[data-original],img[loading="lazy"]');
    for(var i=0;i<imgs.length;i++){
        var img=imgs[i];
        var src=img.getAttribute('data-src')||img.getAttribute('data-lazy')||img.getAttribute('data-original');
        if(src&&!img.src){img.src=src;loaded++;}
        if(img.loading==='lazy'){img.loading='eager';loaded++;}
    }
    // 2. Trigger IntersectionObserver by scrolling with yields
    // (execute_script is sync, so we use a different approach: disconnect all observers)
    // This is a best-effort approach — some custom lazy-load won't be caught
    // 3. Force srcset images
    var srcsets=document.querySelectorAll('source[data-srcset],img[data-srcset]');
    for(var i=0;i<srcsets.length;i++){
        var el=srcsets[i];
        var ss=el.getAttribute('data-srcset');
        if(ss){el.srcset=ss;loaded++;}
    }
    // 4. Trigger scroll event to wake up scroll-based lazy loaders
    window.scrollTo(0,document.body.scrollHeight);
    window.dispatchEvent(new Event('scroll'));
    window.dispatchEvent(new Event('resize'));
    return loaded;
})();
"""

def kill_driver(drv, timeout=5):
    if drv is None:
        return
    pid = None
    try:
        svc = getattr(drv, 'service', None)
        proc = getattr(svc, 'process', None) if svc else None
        if proc: pid = proc.pid
    except Exception: pass
    try: drv.quit()
    except Exception: pass
    if pid:
        import subprocess as _sp
        try:
            if proc:
                proc.wait(timeout=2)
                return
        except Exception: pass
        try:
            if sys.platform == 'win32':
                _sp.run(['taskkill', '/F', '/T', '/PID', str(pid)], capture_output=True, timeout=5)
            else:
                import signal as _sig
                os.killpg(os.getpgid(pid), _sig.SIGKILL)
        except Exception:
            try:
                if proc: proc.kill()
            except Exception: pass

def build_auth_url(url, user, password):
    if not user: return url
    p = urlparse(url); nl = f"{user}:{password}@{p.hostname}"
    if p.port: nl += f":{p.port}"
    return urlunparse(p._replace(netloc=nl))

def setup_basic_auth(driver, config):
    """Configure CDP-based Basic Auth headers on a WebDriver instance.
    NOTE: setExtraHTTPHeaders applies to ALL requests including third-party CDNs.
    This is acceptable for dev/staging servers but credentials could leak to external domains."""
    import base64 as _b64
    ba = config.get("basic_auth_user", "").strip()
    if not ba: return
    token = _b64.b64encode(f"{ba}:{config.get('basic_auth_pass', '')}".encode()).decode()
    try:
        driver.execute_cdp_cmd("Network.enable", {})
        driver.execute_cdp_cmd("Network.setExtraHTTPHeaders", {"headers": {"Authorization": f"Basic {token}"}})
    except Exception:
        pass

STEP_TYPES = ["入力", "クリック", "ホバー", "選択", "待機", "要素待機", "スクロール", "スクショ", "検証", "戻る", "更新", "アラートOK", "アラートキャンセル", "ナビゲーション", "セッション削除", "見出し", "コメント"]
STEP_ICONS = {"入力": ft.Icons.EDIT, "クリック": ft.Icons.MOUSE,
              "ホバー": ft.Icons.NEAR_ME,
              "選択": ft.Icons.ARROW_DROP_DOWN_CIRCLE,
              "待機": ft.Icons.HOURGLASS_BOTTOM, "要素待機": ft.Icons.VISIBILITY,
              "スクロール": ft.Icons.SWAP_VERT,
              "スクショ": ft.Icons.CAMERA_ALT,
              "戻る": ft.Icons.ARROW_BACK, "更新": ft.Icons.REFRESH,
              "アラートOK": ft.Icons.CHECK_CIRCLE, "アラートキャンセル": ft.Icons.CANCEL,
              "ナビゲーション": ft.Icons.OPEN_IN_BROWSER,
              "セッション削除": ft.Icons.DELETE_SWEEP,
              "見出し": ft.Icons.TITLE, "コメント": ft.Icons.COMMENT}
SCROLL_MODES = [("element", "要素へスクロール"), ("pixel", "ピクセル指定"), ("top", "先頭に戻る")]
INPUT_MODES = [("overwrite", "上書き"), ("append", "追記"), ("clear", "クリアのみ")]

def step_short(step):
    t = step["type"]
    if t == "見出し": return step.get("text","")
    if t == "コメント": return step.get("text","")
    if t == "入力":
        v = step.get("value") or "{パターン}"
        if len(v) > 20: v = v[:17]+"..."
        sel = step.get("selector") or ""
        if len(sel) > 20: sel = sel[:17]+"..."
        mode_label = {"append": "[追記]", "clear": "[クリア]"}.get(step.get("input_mode",""), "")
        return f"{sel} \u2190 {v} {mode_label}".strip()
    if t == "クリック":
        sel = step.get("selector") or ""
        if sel == "{パターン}": return "{パターン} (全パターン)"
        return sel[:30]+"..." if len(sel) > 30 else sel
    if t == "ホバー":
        sel = step.get("selector") or ""
        return sel[:30]+"..." if len(sel) > 30 else sel
    if t == "選択":
        sel = step.get("selector") or ""
        if len(sel) > 20: sel = sel[:17]+"..."
        v = step.get("value") or ""
        if len(v) > 15: v = v[:12]+"..."
        return f"{sel} \u2190 [{v}]"
    if t == "戻る": return f"ブラウザバック +{step.get('seconds','1.0')}秒"
    if t == "更新": return f"F5更新 +{step.get('seconds','1.0')}秒"
    if t == "アラートOK": return "ダイアログOK"
    if t == "アラートキャンセル": return "ダイアログキャンセル"
    if t == "セッション削除": return "Cookie/セッション全削除"
    if t == "ナビゲーション":
        url = step.get("url") or ""
        return url[:40]+"..." if len(url) > 40 else url
    if t == "待機": return f"{step.get('seconds','1.0')}秒"
    if t == "スクロール":
        sm = step.get("scroll_mode", "element")
        if sm == "top": return "先頭に戻る"
        if sm == "pixel": return f"↓{step.get('scroll_px','0')}px"
        sel = step.get("selector","")
        if len(sel) > 25: sel = sel[:22]+"..."
        return f"→ {sel}"
    if t == "要素待機":
        sel = step.get("selector") or ""; timeout = step.get("seconds") or "10"
        if len(sel) > 25: sel = sel[:22]+"..."
        return f"{sel} (最大{timeout}秒)"
    if t == "スクショ":
        m = step.get("mode","fullpage")
        if m == "fullpage": return "表示範囲"
        if m == "fullshot": return "ページ全体(縦長)"
        if m == "margin": return f"要素+{step.get('margin_px','500')}px"
        if m in ("post","state","attrs"):
            return {"post":"POST値","state":"要素状態","attrs":"要素属性"}.get(m, m)
        return "要素のみ"
    if t == "検証":
        m = step.get("verify_type") or "attrs"
        sel = step.get("selector") or ""
        label = {"post":"POST値","state":"要素状態","attrs":"要素属性"}.get(m, m)
        if sel:
            if len(sel) > 25: sel = sel[:22]+"..."
            return f"{label} {sel}"
        return label
    return str(step)


# Pre-compiled regexes for _normalize_source (avoid re-compiling on every screenshot)
_NS_PATTERNS = None
def _get_ns_patterns():
    global _NS_PATTERNS
    if _NS_PATTERNS is None:
        _NS_PATTERNS = [
            (re.compile(r'(<input[^>]*name=["\'](?:_token|csrf_token|csrfmiddlewaretoken|__RequestVerificationToken|authenticity_token|nonce)["\'][^>]*value=["\'])[^"\']*(["\'])', re.IGNORECASE), r'\1__NORMALIZED__\2'),
            (re.compile(r'(<meta[^>]*name=["\'](?:csrf-token|_token)["\'][^>]*content=["\'])[^"\']*(["\'])', re.IGNORECASE), r'\1__NORMALIZED__\2'),
            (re.compile(r'(<input[^>]*name=["\'](?:PHPSESSID|session_id|_session|jsessionid)["\'][^>]*value=["\'])[^"\']*(["\'])', re.IGNORECASE), r'\1__NORMALIZED__\2'),
            (re.compile(r'\d{4}[-/]\d{2}[-/]\d{2}[\sT]\d{2}:\d{2}:\d{2}'), '__DATETIME__'),
            (re.compile(r'\d{4}[-/]\d{2}[-/]\d{2}[\sT]\d{2}:\d{2}'), '__DATETIME__'),
            (re.compile(r'(?<=["\'\s=])\d{10,13}(?=["\'\s&;])'), '__TIMESTAMP__'),
            (re.compile(r'(nonce=["\'])[A-Za-z0-9+/=]+(["\'])'), r'\1__NONCE__\2'),
            (re.compile(r'(\?(?:v|t|_|ver|version|cache|cb)=)[^"\'&\s]+'), r'\1__CACHE__'),
        ]
    return _NS_PATTERNS

def _normalize_source(html):
    """Normalize HTML source for diff comparison.
    Removes volatile values that change between runs (timestamps, CSRF tokens, etc.)."""
    for pattern, repl in _get_ns_patterns():
        html = pattern.sub(repl, html)
    return html

def _generate_report(outdir, log_cb, pages=None):
    """Generate an HTML report. Walks subdirectories for PNGs."""
    from html import escape as _esc
    try:
        all_pngs = []
        for root, dirs, files in os.walk(outdir):
            dirs[:] = [d for d in sorted(dirs) if not d.startswith("_")]
            for fn in sorted(files):
                if fn.lower().endswith(".png"):
                    rel = os.path.relpath(os.path.join(root, fn), outdir).replace("\\", "/")
                    all_pngs.append(rel)
        if not all_pngs: return
        # Collect page URLs
        url_lines = ""
        if pages:
            urls = [f'<li><strong>{_esc(pg.get("number",""))}.{_esc(pg.get("name",""))}</strong>: <a href="{_esc(pg.get("url",""))}">{_esc(pg.get("url",""))}</a></li>'
                    for pg in pages if pg.get("url","")]
            if urls: url_lines = '<h2>対象URL</h2><ul>' + ''.join(urls) + '</ul>'
        html = ['<!DOCTYPE html><html lang="ja"><head><meta charset="UTF-8">',
                '<title>y-shot レポート</title>',
                '<style>body{font-family:sans-serif;max-width:1200px;margin:0 auto;padding:20px;background:#f8f9fa}',
                'h1{color:#333}h2{color:#555;margin-top:32px;border-bottom:2px solid #ddd;padding-bottom:4px}',
                '.card{background:#fff;border-radius:8px;box-shadow:0 1px 4px rgba(0,0,0,.1);margin:16px 0;padding:16px}',
                '.card img{max-width:100%;border:1px solid #ddd;border-radius:4px}',
                '.card .name{font-weight:bold;color:#555;margin-bottom:8px;font-size:14px}',
                '.card .meta{font-size:12px;color:#888;margin-bottom:6px}</style></head><body>',
                f'<h1>y-shot レポート</h1><p>{os.path.basename(outdir)} — {len(all_pngs)} 枚</p>',
                url_lines]
        cur_dir = None
        for rel in all_pngs:
            d = os.path.dirname(rel)
            if d != cur_dir:
                cur_dir = d
                if d: html.append(f'<h2>{_esc(d)}</h2>')
            # A5: Parse filename for metadata
            fn_base = os.path.splitext(os.path.basename(rel))[0]
            parts = fn_base.split('_')
            meta = ""
            if len(parts) >= 4:
                # Format: 001_番号_テスト名_p01_パターン_ss1
                tc_num = parts[1] if len(parts) > 1 else ""
                tc_name = parts[2] if len(parts) > 2 else ""
                pat_label = parts[4] if len(parts) > 4 else ""
                meta = f'<div class="meta">{_esc(tc_num)} {_esc(tc_name)} — {_esc(pat_label)}</div>'
            html.append(f'<div class="card">{meta}<div class="name">{_esc(rel)}</div><img src="{_esc(rel)}" loading="lazy"></div>')
        html.append('</body></html>')
        rp = os.path.join(outdir, "report.html")
        with open(rp, "w", encoding="utf-8") as f: f.write("\n".join(html))
        log_cb(f"[レポート] {rp}")
    except Exception as x:
        log_cb(f"[WARN] レポート生成失敗: {x}")

def _generate_excel_report(outdir, log_cb, pages=None, test_cases=None, run_label=""):
    """Generate Excel report. Walks subdirectories; one sheet per subfolder."""
    try:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XlImage
        from PIL import Image as PILImage
    except ImportError:
        log_cb("[WARN] Excel出力には openpyxl と Pillow が必要です"); return
    try:
        # Collect all PNGs with full paths, grouped by subfolder
        all_pngs = []  # (full_path, display_name, subfolder_name)
        for root, dirs, files in os.walk(outdir):
            dirs[:] = [d for d in sorted(dirs) if not d.startswith("_")]
            for fn in sorted(files):
                if fn.lower().endswith(".png"):
                    fp = os.path.join(root, fn)
                    rel = os.path.relpath(fp, outdir).replace("\\", "/")
                    subdir = os.path.basename(root) if root != outdir else ""
                    all_pngs.append((fp, rel, subdir))
        if not all_pngs: return
        wb = Workbook()
        MAX_IMG_WIDTH = 800

        # Write URL summary on a cover sheet
        ws_cover = wb.active; ws_cover.title = "概要"
        ws_cover.column_dimensions["A"].width = 20; ws_cover.column_dimensions["B"].width = 80
        ws_cover.cell(row=1, column=1, value="y-shot レポート").font = ws_cover.cell(row=1, column=1).font.copy(bold=True, size=14)
        ws_cover.cell(row=2, column=1, value="実行日時"); ws_cover.cell(row=2, column=2, value=os.path.basename(outdir))
        ws_cover.cell(row=3, column=1, value="スクショ数"); ws_cover.cell(row=3, column=2, value=len(all_pngs))
        if run_label: ws_cover.cell(row=4, column=1, value="実行ラベル"); ws_cover.cell(row=4, column=2, value=run_label)
        row = 6
        ws_cover.cell(row=row, column=1, value="対象URL").font = ws_cover.cell(row=row, column=1).font.copy(bold=True, size=12)
        row += 1
        if pages:
            for pg in pages:
                url = pg.get("url", "")
                if url:
                    ws_cover.cell(row=row, column=1, value=f"{pg.get('number','')}.{pg.get('name','')}")
                    ws_cover.cell(row=row, column=2, value=url)
                    row += 1

        def _write_sheet(ws, png_list):
            ws.column_dimensions["A"].width = 60; ws.column_dimensions["B"].width = 5
            current_row = 1
            for fp, display, _ in png_list:
                cell = ws.cell(row=current_row, column=1, value=display)
                cell.font = cell.font.copy(bold=True, size=11)
                ws.row_dimensions[current_row].height = 20; current_row += 1
                pil_img = PILImage.open(fp)
                orig_w, orig_h = pil_img.size
                scale = min(1.0, MAX_IMG_WIDTH / orig_w) if orig_w > 0 else 1.0
                xl_img = XlImage(fp)
                xl_img.width = int(orig_w * scale); xl_img.height = int(orig_h * scale)
                ws.add_image(xl_img, f"A{current_row}")
                current_row += max(1, int(orig_h * scale) // 15 + 2)

        # Group by subfolder
        groups = {}
        for item in all_pngs:
            key = item[2] or "root"
            groups.setdefault(key, []).append(item)

        if len(groups) > 1 or (len(groups) == 1 and "root" not in groups):
            for folder_name, pngs in groups.items():
                sheet_name = folder_name[:31] if folder_name != "root" else "エビデンス"
                _base, _n = sheet_name, 2
                while sheet_name in [s.title for s in wb.worksheets]:
                    sheet_name = f"{_base[:28]}({_n})"; _n += 1
                ws = wb.create_sheet(title=sheet_name)
                _write_sheet(ws, pngs)
        else:
            ws = wb.create_sheet(title="エビデンス")
            _write_sheet(ws, all_pngs)

        xp = os.path.join(outdir, "evidence.xlsx")
        wb.save(xp); log_cb(f"[Excel] {xp}")
    except Exception as x:
        log_cb(f"[WARN] Excel生成失敗: {x}")

def run_all_tests(config, test_cases, pattern_sets, log_cb, done_cb, stop_event=None, progress_cb=None, driver_ref=None, pages=None, run_label="", project_name=""):
    try:
        from selenium import webdriver
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.webdriver.support.ui import Select as SeleniumSelect
        from selenium.webdriver.common.keys import Keys
    except ImportError:
        log_cb("[ERROR] selenium が見つかりません。"); done_cb(); return
    import base64 as _b64
    try:
        from PIL import Image as _PILImage
    except ImportError:
        _PILImage = None
    driver = None
    outdir = None
    try:
        opts = webdriver.ChromeOptions()
        opts.add_argument("--ignore-certificate-errors")
        opts.add_argument("--allow-insecure-localhost")
        opts.add_argument("--disable-features=HttpsUpgrades")
        opts.add_argument("--disable-blink-features=AutomationControlled")
        opts.add_experimental_option("excludeSwitches", ["enable-automation"])
        opts.add_argument("--disable-notifications")
        # Enable performance logging for POST data capture (mode="post")
        opts.set_capability("goog:loggingPrefs", {"performance": "ALL"})
        if config.get("headless") == "1":
            opts.add_argument("--headless=new")
            log_cb("[INFO] ヘッドレスモード")
        driver = webdriver.Chrome(options=opts); driver.set_window_size(1280, 900)
        # Stealth: remove navigator.webdriver flag
        try: driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {"source": "Object.defineProperty(navigator,'webdriver',{get:()=>false});"})
        except Exception: pass
        if driver_ref is not None: driver_ref.append(driver)
        ba = config.get("basic_auth_user","").strip()
        if ba:
            setup_basic_auth(driver, config)
            log_cb("[INFO] Basic認証を設定 (CDP)")
        # Build page URL lookup: page_id -> url
        _page_urls = {}
        for pg in (pages or []):
            pu = pg.get("url", "").strip()
            if pu: _page_urls[pg["_id"]] = pu
        _project_url = config.get("project_url", "").strip()
        def _resolve_url(tc):
            """Resolve start URL: project URL > test URL > page URL."""
            if _project_url: return _project_url
            tc_url = tc.get("url", "").strip()
            if tc_url: return tc_url
            return _page_urls.get(tc.get("page_id", ""), "")
        outdir_base = config.get("output_dir", os.path.join(get_app_dir(), "screenshots"))
        ts = datetime.now().strftime("%Y%m%d")
        if project_name:
            safe_proj = _safe_filename(project_name, 30)
            outdir = os.path.join(outdir_base, f"{ts}_{safe_proj}")
        else:
            outdir = os.path.join(outdir_base, ts)
        os.makedirs(outdir, exist_ok=True)
        gss = 0
        total_pats = 0
        for tc in test_cases:
            pn = tc.get("pattern")
            total_pats += len(pattern_sets.get(pn, [])) if pn else 1
        done_pats = 0

        # Build page lookup — folders are created lazily on first use
        page_dirs = {}  # page_id -> directory path (resolved, but not yet on disk)
        save_source = config.get("save_source", "1") == "1"
        source_dirs = {}  # page_id -> _source subdirectory path
        source_root = os.path.join(outdir, "_source")
        _needed_page_ids = set(tc.get("page_id", "") for tc in test_cases)
        _page_dir_paths = {}  # page_id -> planned path (not yet created)
        if pages:
            for pg in pages:
                if pg["_id"] not in _needed_page_ids: continue
                pg_num = pg.get("number", "0")
                pg_name = _safe_filename(pg.get("name", ""), 30)
                _page_dir_paths[pg["_id"]] = os.path.join(outdir, f"{pg_num}_{pg_name}")
        def _ensure_page_dir(pid):
            """Create page output directory on first use and return its path."""
            if pid in page_dirs:
                return page_dirs[pid]
            planned = _page_dir_paths.get(pid)
            if planned:
                # Clear existing page dir on re-run (same-day overwrite)
                if os.path.isdir(planned):
                    for old_f in os.listdir(planned):
                        old_fp = os.path.join(planned, old_f)
                        if os.path.isfile(old_fp):
                            os.remove(old_fp)
                os.makedirs(planned, exist_ok=True)
                page_dirs[pid] = planned
                if save_source:
                    os.makedirs(source_root, exist_ok=True)
                    src_dir = os.path.join(source_root, os.path.basename(planned))
                    # Clear existing source dir too
                    if os.path.isdir(src_dir):
                        for old_f in os.listdir(src_dir):
                            old_fp = os.path.join(src_dir, old_f)
                            if os.path.isfile(old_fp):
                                os.remove(old_fp)
                    os.makedirs(src_dir, exist_ok=True)
                    source_dirs[pid] = src_dir
                return planned
            return outdir

        for tc_idx, tc in enumerate(test_cases, 1):
            if stop_event and stop_event.is_set():
                log_cb("[中断] ユーザーにより中断されました"); break
            tc_name = tc.get("name", f"テスト{tc_idx}")
            tc_number = tc.get("number", "")
            steps = tc.get("steps", [])
            pat_name = tc.get("pattern")
            pats = pattern_sets.get(pat_name, []) if pat_name else []
            if not pats:
                pats = [{"label": "single", "value": ""}]
            tc_pid = tc.get("page_id")
            tc_outdir = None  # resolved lazily on first screenshot
            log_cb("")
            log_cb(f"{'='*50}")
            log_cb(f"テストケース: {tc_number} {tc_name} ({len(pats)} パターン)")
            log_cb(f"{'='*50}")
            log_cb("")

            for pi, pat in enumerate(pats, 1):
                if stop_event and stop_event.is_set():
                    log_cb("[中断] ユーザーにより中断されました"); break
                label, value = pat.get("label",f"p{pi:03d}"), pat.get("value","")
                log_cb(f"--- [{pi}/{len(pats)}] {label} ({len(value)}文字) ---")
                tc_start_url = _resolve_url(tc)
                if not tc_start_url:
                    log_cb(f"  [WARN] URL未設定 - スキップ"); continue
                tc_base = build_auth_url(tc_start_url, ba, config.get("basic_auth_pass","")) if ba else tc_start_url
                driver.get(tc_base)
                try: WebDriverWait(driver, 15).until(lambda d: d.execute_script("return document.readyState") == "complete")
                except Exception: pass
                # Wait for SPA/JS content to stabilize (DOM element count stops changing)
                try:
                    _prev_count = 0
                    for _wait_i in range(8):  # max 4 seconds (8 x 0.5s)
                        time.sleep(0.5)
                        _cur_count = driver.execute_script("return document.querySelectorAll('*').length;") or 0
                        if _cur_count > 0 and _cur_count == _prev_count:
                            break
                        _prev_count = _cur_count
                except Exception: pass
                sc = 0
                _step_failed = False
                for si, step in enumerate(steps, 1):
                    if stop_event and stop_event.is_set(): break
                    if _step_failed and step.get("type") not in ("スクショ", "検証"):
                        # Skip remaining steps after failure (except screenshots/verification for evidence)
                        continue
                    st = step["type"]
                    if st in ("見出し","コメント"):
                        if st == "見出し": log_cb(f"  ## {step.get('text','')}")
                        continue
                    # iframe switching
                    if "_frame" in step:
                        try:
                            driver.switch_to.default_content()
                            fi = step.get("_frame_index", 0)
                            iframes = driver.find_elements(By.CSS_SELECTOR, "iframe, frame")
                            if fi < len(iframes): driver.switch_to.frame(iframes[fi])
                        except Exception as fx: log_cb(f"  S{si} [WARN] iframe切替失敗: {fx}")
                    elif si > 1 and st not in ("アラートOK", "アラートキャンセル", "セッション削除"):
                        # Return to default content if previous step was in iframe
                        # (skip for alert steps — switch_to.default_content() auto-dismisses confirm dialogs)
                        try: driver.switch_to.default_content()
                        except Exception: pass
                    # 空セレクタの事前チェック（セレクタ必須のステップ）
                    _needs_sel = st in ("入力", "クリック", "ホバー", "選択", "要素待機")
                    if _needs_sel and not step.get("selector", "").strip():
                        log_cb(f"  S{si} [WARN] セレクタ未設定 — スキップ ({st})")
                        _step_failed = True; continue
                    if st == "入力":
                        sel = step.get("selector","")
                        iv = step.get("value","{パターン}").replace("{パターン}",value).replace("{pattern}",value)
                        input_mode = step.get("input_mode", "overwrite")
                        try:
                            e = WebDriverWait(driver,10).until(EC.presence_of_element_located(_sel_by(sel)))
                            driver.execute_script("arguments[0].scrollIntoView({block:'center',behavior:'instant'});", e)
                            etype = (e.get_attribute("type") or "").lower()
                            if input_mode == "clear":
                                # Clear only — no new value
                                if etype in ("date","time","datetime-local","month","week","color"):
                                    driver.execute_script("arguments[0].value='';arguments[0].dispatchEvent(new Event('change',{bubbles:true}));", e)
                                else:
                                    e.clear()
                                log_cb(f"  S{si} クリア: {sel}")
                            elif input_mode == "append":
                                # Append to existing value (click end, then type)
                                e.click()
                                e.send_keys(Keys.END)
                                e.send_keys(iv)
                                log_cb(f"  S{si} 追記: {sel}")
                            else:
                                # Overwrite (default)
                                if etype in ("date","time","datetime-local","month","week","color") or _has_non_bmp(iv):
                                    driver.execute_script("arguments[0].focus();arguments[0].value='';", e)
                                    driver.execute_script(JS_SET_VALUE, e, iv)
                                else:
                                    e.clear(); e.send_keys(iv)
                                log_cb(f"  S{si} 入力: {sel}")
                        except Exception as x:
                            log_cb(f"  S{si} [WARN] 入力失敗: {x}")
                            _step_failed = True
                    elif st == "クリック":
                        sel = step.get("selector","").replace("{パターン}",value).replace("{pattern}",value)
                        try:
                            _el = WebDriverWait(driver,10).until(EC.presence_of_element_located(_sel_by(sel)))
                            driver.execute_script("arguments[0].scrollIntoView({block:'center',behavior:'instant'});", _el)
                            try:
                                _el.click()
                            except Exception:
                                driver.execute_script("arguments[0].click();", _el)
                            log_cb(f"  S{si} クリック: {sel}")
                        except Exception as x:
                            log_cb(f"  S{si} [WARN] クリック失敗: {x}")
                            _step_failed = True
                    elif st == "ホバー":
                        sel = step.get("selector","")
                        try:
                            from selenium.webdriver.common.action_chains import ActionChains
                            _el = WebDriverWait(driver,10).until(EC.presence_of_element_located(_sel_by(sel)))
                            driver.execute_script("arguments[0].scrollIntoView({block:'center',behavior:'instant'});", _el)
                            ActionChains(driver).move_to_element(_el).perform()
                            time.sleep(0.3)
                            log_cb(f"  S{si} ホバー: {sel}")
                        except Exception as x:
                            log_cb(f"  S{si} [WARN] ホバー失敗: {x}")
                            _step_failed = True
                    elif st == "選択":
                        sel = step.get("selector","")
                        sv = step.get("value","").replace("{パターン}",value).replace("{pattern}",value)
                        try:
                            el = WebDriverWait(driver,10).until(EC.presence_of_element_located(_sel_by(sel)))
                            driver.execute_script("arguments[0].scrollIntoView({block:'center',behavior:'instant'});", el)
                            dd = SeleniumSelect(el)
                            try: dd.select_by_value(sv)
                            except Exception: dd.select_by_visible_text(sv)
                            log_cb(f"  S{si} 選択: {sel} -> {sv}")
                        except Exception as x:
                            log_cb(f"  S{si} [WARN] 選択失敗: {x}")
                            _step_failed = True
                    elif st == "戻る":
                        try:
                            driver.back()
                            s = _safe_float(step.get("seconds","1.0")); time.sleep(s)
                            log_cb(f"  S{si} 戻る (+{s}秒)")
                        except Exception as x: log_cb(f"  S{si} [WARN] 戻る失敗: {x}")
                    elif st == "更新":
                        try:
                            driver.refresh()
                            s = _safe_float(step.get("seconds","1.0")); time.sleep(s)
                            log_cb(f"  S{si} 更新 (+{s}秒)")
                        except Exception as x: log_cb(f"  S{si} [WARN] 更新失敗: {x}")
                    elif st == "セッション削除":
                        try:
                            driver.delete_all_cookies()
                            log_cb(f"  S{si} セッション削除: すべてのCookieを削除")
                        except Exception as x: log_cb(f"  S{si} [WARN] セッション削除失敗: {x}")
                    elif st in ("アラートOK", "アラートキャンセル"):
                        _accept = (st == "アラートOK")
                        try:
                            alert = None
                            _last_ae = None
                            for _retry in range(20):
                                try:
                                    alert = driver.switch_to.alert
                                    break
                                except Exception as _ae:
                                    _last_ae = _ae
                                    if _retry == 0:
                                        log_cb(f"  S{si} [DEBUG] アラート待機中... ({type(_ae).__name__})")
                                    time.sleep(0.5)
                            if alert is None:
                                raise Exception(f"アラートが見つかりません (20回リトライ, 最終エラー: {_last_ae})")
                            if _accept: alert.accept()
                            else: alert.dismiss()
                            log_cb(f"  S{si} {'確認OK' if _accept else '確認×'}")
                        except Exception as x:
                            log_cb(f"  S{si} [WARN] {st}失敗: {x}")
                    elif st == "ナビゲーション":
                        nav_url = step.get("url","").replace("{パターン}",value).replace("{pattern}",value)
                        try:
                            if ba: nav_url = build_auth_url(nav_url, ba, config.get("basic_auth_pass",""))
                            driver.get(nav_url)
                            try: WebDriverWait(driver, 15).until(lambda d: d.execute_script("return document.readyState") == "complete")
                            except Exception: pass
                            log_cb(f"  S{si} ナビ: {nav_url[:60]}")
                        except Exception as x: log_cb(f"  S{si} [WARN] ナビゲーション失敗: {x}")
                    elif st == "スクロール":
                        sm = step.get("scroll_mode", "element")
                        try:
                            if sm == "top":
                                driver.execute_script("window.scrollTo(0,0);")
                                log_cb(f"  S{si} スクロール: 先頭")
                            elif sm == "pixel":
                                px = int(_safe_float(step.get("scroll_px", "0"), 0))
                                driver.execute_script(f"window.scrollTo(0,{px});")
                                log_cb(f"  S{si} スクロール: {px}px")
                            else:
                                sel = step.get("selector","")
                                el = WebDriverWait(driver,10).until(EC.presence_of_element_located(_sel_by(sel)))
                                driver.execute_script("arguments[0].scrollIntoView({block:'center',behavior:'instant'});", el)
                                log_cb(f"  S{si} スクロール: {sel}")
                            time.sleep(0.3)
                        except Exception as x: log_cb(f"  S{si} [WARN] スクロール失敗: {x}")
                    elif st == "待機":
                        s = _safe_float(step.get("seconds","1.0")); time.sleep(s); log_cb(f"  S{si} 待機: {s}秒")
                    elif st == "要素待機":
                        sel = step.get("selector","")
                        timeout = _safe_float(step.get("seconds","10"), 10.0)
                        try:
                            WebDriverWait(driver, timeout).until(EC.visibility_of_element_located(_sel_by(sel)))
                            log_cb(f"  S{si} 要素待機OK: {sel}")
                        except Exception as x:
                            log_cb(f"  S{si} [WARN] 要素待機タイムアウト({timeout}秒): {sel}")
                    elif st == "スクショ":
                        # Remove overlays (cookie banners, tracking popups) before capture
                        try:
                            n_removed = driver.execute_script("return " + JS_REMOVE_OVERLAYS.strip()) or 0
                            if n_removed:
                                log_cb(f"  S{si} オーバーレイ {n_removed}件 非表示")
                        except Exception: pass
                        # Flash effect (brief white overlay before capture)
                        try:
                            driver.execute_script(
                                "var f=document.createElement('div');f.id='__yshot_flash';"
                                "f.style.cssText='position:fixed;top:0;left:0;width:100%;height:100%;"
                                "background:white;opacity:0.7;z-index:2147483647;pointer-events:none;';"
                                "document.body.appendChild(f);"
                                "setTimeout(function(){var x=document.getElementById('__yshot_flash');if(x)x.remove();},150);")
                            time.sleep(0.2)
                        except Exception: pass
                        sc += 1; gss += 1; mode = step.get("mode","fullpage"); sel = step.get("selector","")
                        if tc_outdir is None: tc_outdir = _ensure_page_dir(tc_pid)
                        safe_tc = _safe_filename(tc_name, 20)
                        safe_number = _safe_filename(tc_number, 10) if tc_number else ""
                        num_prefix = f"{safe_number}_" if safe_number else ""
                        if len(pats) > 1:
                            safe_label = _safe_filename(label, 30)
                            fn = f"{gss:03d}_{num_prefix}{safe_tc}_{safe_label}_ss{sc}.png"
                        else:
                            fn = f"{gss:03d}_{num_prefix}{safe_tc}_ss{sc}.png"
                        fp = os.path.join(tc_outdir, fn)
                        try:
                            if mode == "element" and sel:
                                driver.find_element(*_sel_by(sel)).screenshot(fp)
                            elif mode == "margin" and sel:
                                if _PILImage is None:
                                    log_cb("  [WARN] Pillow未インストール: marginモードはfullpageにフォールバック")
                                    driver.save_screenshot(fp)
                                else:
                                    mg = int(_safe_float(step.get("margin_px",500), 500))
                                    tgt = driver.find_element(*_sel_by(sel))
                                    driver.execute_script("arguments[0].scrollIntoView({block:'center',behavior:'instant'});",tgt)
                                    time.sleep(0.3)
                                    r = driver.execute_script("var r=arguments[0].getBoundingClientRect();return{x:r.x,y:r.y,w:r.width,h:r.height};",tgt)
                                    driver.save_screenshot(fp)
                                    img = _PILImage.open(fp)
                                    d = driver.execute_script("return window.devicePixelRatio||1;")
                                    x1,y1 = max(0,int(r["x"]*d)-mg), max(0,int(r["y"]*d)-mg)
                                    x2,y2 = min(img.width,int((r["x"]+r["w"])*d)+mg), min(img.height,int((r["y"]+r["h"])*d)+mg)
                                    if x2>x1 and y2>y1:
                                        img.crop((x1,y1,x2,y2)).save(fp)
                                    else:
                                        log_cb(f"  S{si} [WARN] マージンcrop無効 (rect={x1},{y1},{x2},{y2}) — 元画像を使用")
                            elif mode == "fullshot":
                                # Preload lazy images: force-load + scroll to trigger loaders
                                try:
                                    n_loaded = driver.execute_script("return " + JS_PRELOAD_LAZY_IMAGES.strip()) or 0
                                    # Scroll through page to trigger remaining lazy loaders
                                    total_h = driver.execute_script("return Math.max(document.body.scrollHeight,document.documentElement.scrollHeight);")
                                    view_h = driver.execute_script("return window.innerHeight;")
                                    if total_h and view_h:
                                        for pos in range(0, int(total_h), max(int(view_h * 0.8), 100)):
                                            driver.execute_script(f"window.scrollTo(0,{pos});")
                                            time.sleep(0.1)
                                    if n_loaded:
                                        log_cb(f"  S{si} 遅延画像 {n_loaded}件 先読み")
                                except Exception:
                                    pass
                                # CDP full-page screenshot (captures entire scrollable page)
                                try:
                                    metrics = driver.execute_cdp_cmd('Page.getLayoutMetrics', {})
                                    # Chrome 120+: cssContentSize, older: contentSize
                                    cs = metrics.get('cssContentSize') or metrics.get('contentSize', {})
                                    cw = cs.get('width', 1280)
                                    ch = cs.get('height', 900)
                                    if ch > 16384:
                                        log_cb(f"  S{si} [WARN] ページ高さ{ch}px > 上限16384px。画像が切れます")
                                        ch = 16384
                                    _flog.debug(f"fullshot: {cw}x{ch} (keys={list(metrics.keys())})")
                                    result = driver.execute_cdp_cmd('Page.captureScreenshot', {
                                        'format': 'png',
                                        'captureBeyondViewport': True,
                                        'clip': {'x': 0, 'y': 0, 'width': cw, 'height': ch, 'scale': 1}
                                    })
                                    with open(fp, 'wb') as _f:
                                        _f.write(_b64.b64decode(result['data']))
                                except Exception as cdp_err:
                                    _flog.error(f"CDP fullshot failed: {cdp_err}, falling back to resize method")
                                    # Fallback: resize window to full page height
                                    try:
                                        total_h = driver.execute_script("return Math.max(document.body.scrollHeight, document.documentElement.scrollHeight);")
                                        total_w = driver.execute_script("return Math.max(document.body.scrollWidth, document.documentElement.scrollWidth);")
                                        actual_h = min(total_h, 16384)
                                        if total_h > 16384:
                                            log_cb(f"  S{si} [WARN] ページ高さ{total_h}px > 上限16384px。画像が切れます")
                                        driver.set_window_size(max(total_w, 1280), actual_h)
                                        time.sleep(0.5)
                                        driver.save_screenshot(fp)
                                    finally:
                                        driver.set_window_size(1280, 900)
                            elif mode == "state" and sel:
                                # ── 要素状態キャプチャモード ──
                                # 指定要素の全属性・計算済みスタイルをHTMLテーブルとしてDOMに注入
                                try:
                                    tgt_el = driver.find_element(*_sel_by(sel))
                                    state_info = driver.execute_script("""
var el = arguments[0];
var result = {tagName: el.tagName.toLowerCase(), attrs: [], styles: []};
// 全HTML属性
for (var i = 0; i < el.attributes.length; i++) {
    var a = el.attributes[i];
    result.attrs.push([a.name, a.value]);
}
// 入力要素の現在値
if (typeof el.value !== 'undefined' && el.value !== '') {
    result.attrs.push(['[現在値]', el.value]);
}
// textContent (入力以外)
if (!('value' in el) || el.tagName === 'BUTTON') {
    var tc = (el.textContent || '').trim();
    if (tc && tc.length <= 200) result.attrs.push(['[textContent]', tc]);
    else if (tc) result.attrs.push(['[textContent]', tc.substring(0, 200) + '…']);
}
// 計算済みスタイル (リソースURL関連 + レイアウト)
var cs = window.getComputedStyle(el);
var props = ['display','visibility','opacity','width','height',
    'background-image','background-color','color','font-size',
    'border-width','border-style','border-color','position','overflow'];
var skip_none = {'background-image':1,'background-color':1,'border-style':1,'border-color':1};
for (var j = 0; j < props.length; j++) {
    var p = props[j], v = cs.getPropertyValue(p);
    if (!v) continue;
    if (v === 'normal' || v === 'auto' || v === '0px') continue;
    if (v === 'none' && skip_none[p]) continue;
    result.styles.push([p, v]);
}
return result;
""", tgt_el)
                                    def _esc(s):
                                        return str(s).replace("&","&amp;").replace("<","&lt;").replace(">","&gt;").replace('"',"&quot;")
                                    # 属性テーブル行
                                    attr_rows = "".join(
                                        f'<tr><td style="border:1px solid #666;padding:6px 10px;font-weight:bold;background:#f5f5f5;white-space:nowrap">{_esc(k)}</td>'
                                        f'<td style="border:1px solid #666;padding:6px 10px;word-break:break-all">{_esc(v) if v else "<span style=color:#999>(空)</span>"}</td></tr>'
                                        for k, v in state_info.get("attrs", [])
                                    )
                                    # 計算済みスタイル行
                                    style_rows = "".join(
                                        f'<tr><td style="border:1px solid #666;padding:6px 10px;font-weight:bold;background:#f5f5f5;white-space:nowrap">{_esc(k)}</td>'
                                        f'<td style="border:1px solid #666;padding:6px 10px;word-break:break-all">{_esc(v)}</td></tr>'
                                        for k, v in state_info.get("styles", [])
                                    )
                                    tag_name = _esc(state_info.get("tagName", "?"))
                                    safe_sel = _esc(sel)
                                    inject_html = (
                                        f'<div id="__yshot_state" style="margin:15px;padding:15px;border:3px solid #2E7D32;border-radius:8px;background:#E8F5E9;font-family:sans-serif">'
                                        f'<div style="font-size:16px;font-weight:bold;color:#2E7D32;margin-bottom:8px">🔍 要素の状態</div>'
                                        f'<div style="font-size:11px;color:#555;margin-bottom:10px;word-break:break-all">&lt;{tag_name}&gt; — {safe_sel}</div>'
                                        f'<table style="border-collapse:collapse;width:100%;font-size:13px;margin-bottom:10px">'
                                        f'<tr><th style="border:1px solid #666;padding:6px 10px;background:#2E7D32;color:#fff;text-align:left">属性名</th>'
                                        f'<th style="border:1px solid #666;padding:6px 10px;background:#2E7D32;color:#fff;text-align:left">値</th></tr>'
                                        f'{attr_rows}</table>'
                                    )
                                    if style_rows:
                                        inject_html += (
                                            f'<div style="font-size:13px;font-weight:bold;color:#2E7D32;margin:8px 0 4px">計算済みスタイル</div>'
                                            f'<table style="border-collapse:collapse;width:100%;font-size:13px">'
                                            f'<tr><th style="border:1px solid #666;padding:6px 10px;background:#558B2F;color:#fff;text-align:left">プロパティ</th>'
                                            f'<th style="border:1px solid #666;padding:6px 10px;background:#558B2F;color:#fff;text-align:left">値</th></tr>'
                                            f'{style_rows}</table>'
                                        )
                                    inject_html += '</div>'
                                    driver.execute_script(
                                        "var target=document.querySelector('main')||document.body;"
                                        "target.insertAdjacentHTML('afterbegin',arguments[0]);",
                                        inject_html
                                    )
                                    time.sleep(0.3)
                                    n_attrs = len(state_info.get("attrs", []))
                                    n_styles = len(state_info.get("styles", []))
                                    log_cb(f"  S{si} 要素状態 属性{n_attrs}件+スタイル{n_styles}件 挿入")
                                except Exception as state_err:
                                    log_cb(f"  S{si} [WARN] 要素状態取得失敗: {state_err}")
                                # fullshot と同じ方法でスクショ
                                try:
                                    metrics = driver.execute_cdp_cmd('Page.getLayoutMetrics', {})
                                    cs = metrics.get('cssContentSize') or metrics.get('contentSize', {})
                                    cw, ch = cs.get('width', 1280), cs.get('height', 900)
                                    if ch > 16384: ch = 16384
                                    result = driver.execute_cdp_cmd('Page.captureScreenshot', {
                                        'format': 'png', 'captureBeyondViewport': True,
                                        'clip': {'x': 0, 'y': 0, 'width': cw, 'height': ch, 'scale': 1}
                                    })
                                    with open(fp, 'wb') as _f:
                                        _f.write(_b64.b64decode(result['data']))
                                except Exception:
                                    driver.save_screenshot(fp)
                                # 挿入したDOMを除去
                                try:
                                    driver.execute_script("var e=document.getElementById('__yshot_state');if(e)e.remove();")
                                except Exception:
                                    pass
                            elif mode == "post":
                                # ── POST値キャプチャモード ──
                                # パフォーマンスログからPOSTリクエストを抽出
                                import json as _json_mod
                                post_entries = []
                                try:
                                    perf_logs = driver.get_log("performance")
                                    for entry in perf_logs:
                                        try:
                                            msg = _json_mod.loads(entry["message"])["message"]
                                            if msg.get("method") != "Network.requestWillBeSent":
                                                continue
                                            req = msg.get("params", {}).get("request", {})
                                            if req.get("method") != "POST":
                                                continue
                                            pd = req.get("postData", "")
                                            if not pd:
                                                continue
                                            post_entries.append({
                                                "url": req.get("url", ""),
                                                "postData": pd,
                                                "contentType": req.get("headers", {}).get("Content-Type", ""),
                                            })
                                        except Exception:
                                            continue
                                except Exception as plog_err:
                                    log_cb(f"  S{si} [WARN] パフォーマンスログ取得失敗: {plog_err}")
                                # POSTデータをパース→HTMLテーブル化→DOM挿入
                                if post_entries:
                                    from urllib.parse import unquote_plus
                                    last_post = post_entries[-1]  # 直近のPOSTリクエスト
                                    raw_pd = last_post["postData"]
                                    # URL-encoded のパース
                                    pairs = []
                                    for chunk in raw_pd.split("&"):
                                        if "=" in chunk:
                                            k, v = chunk.split("=", 1)
                                            pairs.append((unquote_plus(k), unquote_plus(v)))
                                        else:
                                            pairs.append((unquote_plus(chunk), ""))
                                    # HTMLテーブル生成（エスケープ付き）
                                    def _esc(s):
                                        return s.replace("&","&amp;").replace("<","&lt;").replace(">","&gt;").replace('"',"&quot;")
                                    tbl_rows = "".join(
                                        f'<tr><td style="border:1px solid #666;padding:6px 10px;font-weight:bold;background:#f5f5f5;white-space:nowrap">{_esc(k)}</td>'
                                        f'<td style="border:1px solid #666;padding:6px 10px;word-break:break-all">{_esc(v) if v else "<span style=color:#999>(空)</span>"}</td></tr>'
                                        for k, v in pairs
                                    )
                                    post_url = _esc(last_post["url"])
                                    inject_html = (
                                        f'<div id="__yshot_post" style="margin:15px;padding:15px;border:3px solid #1565C0;border-radius:8px;background:#E3F2FD;font-family:sans-serif">'
                                        f'<div style="font-size:16px;font-weight:bold;color:#1565C0;margin-bottom:8px">📡 POST データ (実際の送信値)</div>'
                                        f'<div style="font-size:11px;color:#555;margin-bottom:10px;word-break:break-all">送信先: {post_url}</div>'
                                        f'<table style="border-collapse:collapse;width:100%;font-size:13px">'
                                        f'<tr><th style="border:1px solid #666;padding:6px 10px;background:#1565C0;color:#fff;text-align:left">パラメータ名</th>'
                                        f'<th style="border:1px solid #666;padding:6px 10px;background:#1565C0;color:#fff;text-align:left">POST値</th></tr>'
                                        f'{tbl_rows}</table></div>'
                                    )
                                    try:
                                        driver.execute_script(
                                            "var target=document.querySelector('main')||document.body;"
                                            "target.insertAdjacentHTML('afterbegin',arguments[0]);",
                                            inject_html
                                        )
                                        time.sleep(0.3)
                                        log_cb(f"  S{si} POSTデータ {len(pairs)}件 挿入")
                                    except Exception as inj_err:
                                        log_cb(f"  S{si} [WARN] POST DOM挿入失敗: {inj_err}")
                                else:
                                    log_cb(f"  S{si} [INFO] POSTリクエストが見つかりません")
                                # fullshot と同じ方法でスクショ
                                try:
                                    metrics = driver.execute_cdp_cmd('Page.getLayoutMetrics', {})
                                    cs = metrics.get('cssContentSize') or metrics.get('contentSize', {})
                                    cw, ch = cs.get('width', 1280), cs.get('height', 900)
                                    if ch > 16384: ch = 16384
                                    result = driver.execute_cdp_cmd('Page.captureScreenshot', {
                                        'format': 'png', 'captureBeyondViewport': True,
                                        'clip': {'x': 0, 'y': 0, 'width': cw, 'height': ch, 'scale': 1}
                                    })
                                    with open(fp, 'wb') as _f:
                                        _f.write(_b64.b64decode(result['data']))
                                except Exception:
                                    driver.save_screenshot(fp)
                                # 挿入したDOMを除去（後続テストへの影響防止）
                                try:
                                    driver.execute_script("var e=document.getElementById('__yshot_post');if(e)e.remove();")
                                except Exception:
                                    pass
                            else: driver.save_screenshot(fp)
                            rel_dir = os.path.basename(tc_outdir) if tc_outdir != outdir else ""
                            log_cb(f"  S{si} スクショ: {rel_dir + '/' if rel_dir else ''}{fn}")
                            if save_source:
                                try:
                                    src_dir = source_dirs.get(tc_pid, source_root)
                                    src_base = fn.replace(".png", "")
                                    raw = driver.page_source or ""
                                    with open(os.path.join(src_dir, f"{src_base}_raw.html"), "w", encoding="utf-8") as sf:
                                        sf.write(_normalize_source(raw))
                                    dom = driver.execute_script("return document.documentElement.outerHTML;") or ""
                                    with open(os.path.join(src_dir, f"{src_base}_dom.html"), "w", encoding="utf-8") as sf:
                                        sf.write(_normalize_source(dom))
                                except Exception as sx:
                                    _flog.debug(f"Source save failed: {sx}")
                        except Exception as x: log_cb(f"  S{si} [WARN] スクショ失敗: {x}")
                    elif st == "検証":
                        vtype = step.get("verify_type", "attrs")
                        sel = step.get("selector", "").replace("{パターン}",value).replace("{pattern}",value)
                        sc += 1
                        safe_tc = _safe_filename(tc_name, 30)
                        safe_number = _safe_filename(tc_number, 10) if tc_number else ""
                        num_prefix = f"{safe_number}_" if safe_number else ""
                        if len(pats) > 1:
                            safe_label = _safe_filename(label, 30)
                            fn = f"{gss:03d}_{num_prefix}{safe_tc}_{safe_label}_ss{sc}.png"
                        else:
                            fn = f"{gss:03d}_{num_prefix}{safe_tc}_ss{sc}.png"
                        fp = os.path.join(tc_outdir, fn)
                        try:
                            verify_html = ""
                            if vtype == "post":
                                # POST値: パフォーマンスログから収集
                                import json as _json_mod
                                post_entries = []
                                try:
                                    perf_logs = driver.get_log("performance")
                                    for entry in perf_logs:
                                        try:
                                            msg = _json_mod.loads(entry["message"])["message"]
                                            if msg.get("method") != "Network.requestWillBeSent": continue
                                            req = msg.get("params", {}).get("request", {})
                                            if req.get("method") != "POST": continue
                                            pd = req.get("postData", "")
                                            if not pd: continue
                                            post_entries.append({"url": req.get("url",""), "postData": pd, "contentType": req.get("headers",{}).get("Content-Type","")})
                                        except Exception: continue
                                except Exception as plog_err:
                                    log_cb(f"  S{si} [WARN] パフォーマンスログ取得失敗: {plog_err}")
                                if post_entries:
                                    from urllib.parse import unquote_plus
                                    last_post = post_entries[-1]
                                    raw_pd = last_post["postData"]
                                    pairs = []
                                    for chunk in raw_pd.split("&"):
                                        if "=" in chunk: k,v = chunk.split("=",1); pairs.append((unquote_plus(k),unquote_plus(v)))
                                        else: pairs.append((unquote_plus(chunk),""))
                                    def _esc(s): return str(s).replace("&","&amp;").replace("<","&lt;").replace(">","&gt;").replace('"',"&quot;")
                                    tbl_rows = "".join(f'<tr><td style="border:1px solid #666;padding:6px 10px;font-weight:bold;background:#f5f5f5;white-space:nowrap">{_esc(k)}</td><td style="border:1px solid #666;padding:6px 10px;word-break:break-all">{_esc(v) if v else "<span style=color:#999>(空)</span>"}</td></tr>' for k,v in pairs)
                                    post_url = _esc(last_post["url"])
                                    verify_html = (f'<div style="margin:15px;padding:15px;border:3px solid #1565C0;border-radius:8px;background:#E3F2FD;font-family:sans-serif">'
                                        f'<div style="font-size:16px;font-weight:bold;color:#1565C0;margin-bottom:8px">📡 POST データ (実際の送信値)</div>'
                                        f'<div style="font-size:11px;color:#555;margin-bottom:10px;word-break:break-all">送信先: {post_url}</div>'
                                        f'<table style="border-collapse:collapse;width:100%;font-size:13px"><tr><th style="border:1px solid #666;padding:6px 10px;background:#1565C0;color:#fff;text-align:left">パラメータ名</th><th style="border:1px solid #666;padding:6px 10px;background:#1565C0;color:#fff;text-align:left">POST値</th></tr>{tbl_rows}</table></div>')
                                    log_cb(f"  S{si} 検証:POST {len(pairs)}件")
                                else: log_cb(f"  S{si} [INFO] POSTリクエストなし")
                            elif vtype in ("state", "attrs") and sel:
                                # 要素の状態/属性: 元ページから情報を収集
                                tgt_el = driver.find_element(*_sel_by(sel))
                                state_info = driver.execute_script("""
var el=arguments[0];var r={tagName:el.tagName.toLowerCase(),attrs:[],styles:[],dataUrl:''};
for(var i=0;i<el.attributes.length;i++){var a=el.attributes[i];r.attrs.push([a.name,a.value]);}
if(typeof el.value!=='undefined'&&el.value!=='')r.attrs.push(['[現在値]',el.value]);
if(!('value' in el)||el.tagName==='BUTTON'){var tc=(el.textContent||'').trim();if(tc&&tc.length<=200)r.attrs.push(['[textContent]',tc]);else if(tc)r.attrs.push(['[textContent]',tc.substring(0,200)+'…']);}
if(el.tagName==='IMG'&&el.naturalWidth>0){try{var cv=document.createElement('canvas');cv.width=el.naturalWidth;cv.height=el.naturalHeight;cv.getContext('2d').drawImage(el,0,0);r.dataUrl=cv.toDataURL('image/png');}catch(e){}};
if(arguments[1]){var cs=window.getComputedStyle(el);var ps=['display','visibility','opacity','width','height','background-image','background-color','color','font-size','border-width','border-style','border-color','position','overflow'];for(var j=0;j<ps.length;j++){var p=ps[j],v=cs.getPropertyValue(p);if(!v||v==='normal'||v==='auto'||v==='0px')continue;if(v==='none'&&({'background-image':1,'background-color':1,'border-style':1,'border-color':1})[p])continue;r.styles.push([p,v]);}}
return r;""", tgt_el, vtype == "state")
                                def _esc(s): return str(s).replace("&","&amp;").replace("<","&lt;").replace(">","&gt;").replace('"',"&quot;")
                                attr_rows = "".join(f'<tr><td style="border:1px solid #666;padding:6px 10px;font-weight:bold;background:#f5f5f5;white-space:nowrap">{_esc(k)}</td><td style="border:1px solid #666;padding:6px 10px;word-break:break-all">{_esc(v) if v else "<span style=color:#999>(空)</span>"}</td></tr>' for k,v in state_info.get("attrs",[]))
                                tag_name = _esc(state_info.get("tagName","?")); safe_sel = _esc(sel)
                                color = "#2E7D32" if vtype == "state" else "#6A1B9A"
                                bg = "#E8F5E9" if vtype == "state" else "#F3E5F5"
                                title = "🔍 要素の状態" if vtype == "state" else "🏷 要素の属性"
                                # img の場合はプレビューを追加（base64埋め込みで認証・CORS問題を回避）
                                _img_preview = ""
                                _data_url = state_info.get("dataUrl", "")
                                if tag_name == "img" and _data_url:
                                    _img_preview = (f'<div style="margin-bottom:12px;padding:10px;background:#fff;border:1px solid #ccc;border-radius:4px;text-align:center">'
                                        f'<div style="font-size:11px;color:#888;margin-bottom:6px">画像プレビュー</div>'
                                        f'<img src="{_data_url}" style="max-width:100%;max-height:400px;border:1px solid #ddd">'
                                        f'</div>')
                                verify_html = (f'<div style="margin:15px;padding:15px;border:3px solid {color};border-radius:8px;background:{bg};font-family:sans-serif">'
                                    f'<div style="font-size:16px;font-weight:bold;color:{color};margin-bottom:8px">{title}</div>'
                                    f'<div style="font-size:11px;color:#555;margin-bottom:10px;word-break:break-all">&lt;{tag_name}&gt; — {safe_sel}</div>'
                                    f'{_img_preview}'
                                    f'<table style="border-collapse:collapse;width:100%;font-size:13px"><tr><th style="border:1px solid #666;padding:6px 10px;background:{color};color:#fff;text-align:left">属性名</th><th style="border:1px solid #666;padding:6px 10px;background:{color};color:#fff;text-align:left">値</th></tr>{attr_rows}</table>')
                                if vtype == "state" and state_info.get("styles"):
                                    style_rows = "".join(f'<tr><td style="border:1px solid #666;padding:6px 10px;font-weight:bold;background:#f5f5f5;white-space:nowrap">{_esc(k)}</td><td style="border:1px solid #666;padding:6px 10px;word-break:break-all">{_esc(v)}</td></tr>' for k,v in state_info["styles"])
                                    verify_html += (f'<div style="font-size:13px;font-weight:bold;color:{color};margin:8px 0 4px">計算済みスタイル</div>'
                                        f'<table style="border-collapse:collapse;width:100%;font-size:13px"><tr><th style="border:1px solid #666;padding:6px 10px;background:#558B2F;color:#fff;text-align:left">プロパティ</th><th style="border:1px solid #666;padding:6px 10px;background:#558B2F;color:#fff;text-align:left">値</th></tr>{style_rows}</table>')
                                verify_html += '</div>'
                                n_info = len(state_info.get("attrs",[])) + len(state_info.get("styles",[]))
                                _vtype_label = {"state":"要素状態","attrs":"要素属性"}.get(vtype, vtype)
                                log_cb(f"  S{si} 検証:{_vtype_label} {n_info}件")
                            # 新タブで検証HTMLを表示 → スクショ → タブを閉じる（元ページに影響なし）
                            if verify_html:
                                _orig_handle = driver.current_window_handle
                                _new_tab_opened = False
                                try:
                                    driver.execute_script("window.open('about:blank','_blank');")
                                    driver.switch_to.window(driver.window_handles[-1])
                                    _new_tab_opened = True
                                    _page_html = f'<html><head><meta charset="utf-8"><title>y-shot 検証</title></head><body style="margin:0;background:#fff">{verify_html}</body></html>'
                                    driver.execute_script("document.write(arguments[0]);document.close();", _page_html)
                                    time.sleep(0.3)
                                    try:
                                        metrics = driver.execute_cdp_cmd('Page.getLayoutMetrics', {})
                                        cs_m = metrics.get('cssContentSize') or metrics.get('contentSize', {})
                                        cw, ch = cs_m.get('width', 1280), cs_m.get('height', 900)
                                        if ch > 16384: ch = 16384
                                        result = driver.execute_cdp_cmd('Page.captureScreenshot', {'format':'png','captureBeyondViewport':True,'clip':{'x':0,'y':0,'width':cw,'height':ch,'scale':1}})
                                        with open(fp, 'wb') as _f: _f.write(_b64.b64decode(result['data']))
                                    except Exception: driver.save_screenshot(fp)
                                    rel_dir = os.path.basename(tc_outdir) if tc_outdir != outdir else ""
                                    log_cb(f"  S{si} スクショ: {rel_dir + '/' if rel_dir else ''}{fn}")
                                finally:
                                    if _new_tab_opened:
                                        try: driver.close()
                                        except Exception: pass
                                        try: driver.switch_to.window(_orig_handle)
                                        except Exception: pass
                        except Exception as x: log_cb(f"  S{si} [WARN] 検証失敗: {x}")
                done_pats += 1
                if progress_cb and total_pats > 0:
                    progress_cb(done_pats, total_pats, f"{tc_number} {tc_name}")

        if stop_event and stop_event.is_set():
            log_cb(f"\n[中断完了] -> {outdir}")
        else:
            log_cb(f"\n[全完了] {len(test_cases)} テスト -> {outdir}")
        _used_pages = [pg for pg in (pages or []) if pg["_id"] in page_dirs]
        _generate_report(outdir, log_cb, pages=_used_pages)
        _generate_excel_report(outdir, log_cb, pages=_used_pages, test_cases=test_cases, run_label=run_label)
    except Exception as x:
        log_cb(f"[ERROR] {x}"); outdir = None
    finally:
        if driver:
            kill_driver(driver)
            if driver_ref is not None:
                try: driver_ref.remove(driver)
                except ValueError: pass
        done_cb(outdir)

# ===================================================================
# Path helpers
# ===================================================================

def get_app_dir():
    if getattr(sys, 'frozen', False): return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

def get_bundle_dir():
    if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'): return sys._MEIPASS
    return os.path.dirname(os.path.abspath(__file__))

_active_project_dir = [None]  # activate_project() でセット

def _data_path(filename):
    if _active_project_dir[0]:
        return os.path.join(_active_project_dir[0], filename)
    return os.path.join(get_app_dir(), filename)

# ===================================================================
# Persistence
# ===================================================================
CSV_HEADER = ["label", "value"]
CONFIG_FILE = "y_shot_config.ini"
TESTS_FILE = "y_shot_tests.json"
PATTERNS_FILE = "y_shot_patterns.json"
SELECTOR_BANK_FILE = "y_shot_selectors.json"
PAGES_FILE = "y_shot_pages.json"

def load_csv(path):
    if not os.path.isfile(path): return []
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        return [r for r in csv.DictReader(f) if "label" in r]
def save_csv(path, patterns):
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=CSV_HEADER); w.writeheader(); w.writerows(patterns)
def load_config():
    import configparser; c = configparser.ConfigParser(); c.read(_data_path(CONFIG_FILE), encoding="utf-8")
    return dict(c["settings"]) if "settings" in c else {}
def save_config(data):
    import configparser; c = configparser.ConfigParser()
    c["settings"] = {k: str(v) for k,v in data.items()}
    with open(_data_path(CONFIG_FILE), "w", encoding="utf-8") as f: c.write(f)
def _safe_json_save(filepath, data):
    """Atomic JSON save: write to .tmp, backup old file, then rename."""
    tmp = filepath + ".tmp"
    bak = filepath + ".backup"
    try:
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        if os.path.isfile(filepath):
            try: os.replace(filepath, bak)
            except Exception: pass
        os.replace(tmp, filepath)
    except Exception:
        # tmpファイルが残らないようにクリーンアップ
        try: os.remove(tmp)
        except OSError: pass
        raise

def _safe_json_load(filepath, default):
    """Load JSON with backup recovery on corruption."""
    for p in [filepath, filepath + ".backup"]:
        if os.path.isfile(p):
            try:
                with open(p, "r", encoding="utf-8") as f:
                    data = json.load(f)
                if p.endswith(".backup"):
                    _flog.warning(f"[WARN] メインファイル破損のためバックアップから復元: {filepath}")
                return data
            except (json.JSONDecodeError, Exception):
                continue
    return default

def load_tests():
    tests = _safe_json_load(_data_path(TESTS_FILE), [])
    for t in tests:
        if "name" in t and isinstance(t["name"], str):
            t["name"] = t["name"].replace("\r","").replace("\n","").strip()
    return tests
def save_tests(tests):
    _safe_json_save(_data_path(TESTS_FILE), tests)
def load_pattern_sets():
    return _safe_json_load(_data_path(PATTERNS_FILE), {})
def save_pattern_sets(ps):
    _safe_json_save(_data_path(PATTERNS_FILE), ps)
def load_selector_bank():
    return _safe_json_load(_data_path(SELECTOR_BANK_FILE), {})
def save_selector_bank(bank):
    _safe_json_save(_data_path(SELECTOR_BANK_FILE), bank)
def load_pages():
    return _safe_json_load(_data_path(PAGES_FILE), [])
def save_pages(pages):
    _safe_json_save(_data_path(PAGES_FILE), pages)

# ===================================================================
# Project management
# ===================================================================
PROJECTS_DIR = "projects"
PROJECTS_FILE = "projects.json"

def get_projects_dir():
    return os.path.join(get_app_dir(), PROJECTS_DIR)

def load_projects_registry():
    p = os.path.join(get_projects_dir(), PROJECTS_FILE)
    return _safe_json_load(p, {"last_active": "default", "projects": []})

def save_projects_registry(registry):
    os.makedirs(get_projects_dir(), exist_ok=True)
    _safe_json_save(os.path.join(get_projects_dir(), PROJECTS_FILE), registry)

def activate_project(project_id, registry):
    for proj in registry["projects"]:
        if proj["id"] == project_id:
            d = os.path.join(get_projects_dir(), proj["dir"])
            os.makedirs(d, exist_ok=True)
            _active_project_dir[0] = d
            registry["last_active"] = project_id
            return True
    return False

def _new_project_id(registry):
    max_id = 0
    for proj in registry["projects"]:
        try: max_id = max(max_id, int(proj["id"].split("_", 1)[1]))
        except (ValueError, IndexError): pass
    return f"proj_{max_id + 1}"

def _safe_dir_name(name):
    """Convert project name to a safe directory name."""
    safe = re.sub(r'[\\/:*?"<>|]', '_', name).strip()
    return safe or "project"

def migrate_to_projects():
    reg_path = os.path.join(get_projects_dir(), PROJECTS_FILE)
    if os.path.isfile(reg_path):
        return load_projects_registry()
    default_dir = os.path.join(get_projects_dir(), "default")
    os.makedirs(default_dir, exist_ok=True)
    for fname in [TESTS_FILE, PAGES_FILE, PATTERNS_FILE, SELECTOR_BANK_FILE, CONFIG_FILE]:
        src = os.path.join(get_app_dir(), fname)
        dst = os.path.join(default_dir, fname)
        if os.path.isfile(src) and not os.path.isfile(dst):
            shutil.copy2(src, dst)
    registry = {
        "last_active": "default",
        "projects": [{"id": "default", "name": "デフォルト", "dir": "default",
                       "created": datetime.now().isoformat()}]
    }
    save_projects_registry(registry)
    return registry

def get_templates_dir():
    user_dir = os.path.join(get_app_dir(), "templates")
    bundle_dir = os.path.join(get_bundle_dir(), "templates")
    if not os.path.isdir(user_dir):
        os.makedirs(user_dir, exist_ok=True)
        if os.path.isdir(bundle_dir) and bundle_dir != user_dir:
            import shutil
            for f in os.listdir(bundle_dir):
                if f.lower().endswith(".csv"):
                    src = os.path.join(bundle_dir, f)
                    dst = os.path.join(user_dir, f)
                    if not os.path.exists(dst): shutil.copy2(src, dst)
    return user_dir if os.path.isdir(user_dir) else None

# ===================================================================
# Flet UI
# ===================================================================

def main(page: ft.Page):
    # Log Flet version for diagnostics
    _flog.info(f"Flet version: {ft.__version__}")
    try:
        _main_inner(page)
    except Exception as _fatal:
        tb = traceback.format_exc()
        _flog.error(f"FATAL in main: {_fatal}\n{tb}")
        # Show error in page if possible
        try:
            page.clean()
            page.add(ft.Column([
                ft.Text(f"{APP_NAME} 起動エラー", size=20, weight=ft.FontWeight.BOLD, color=ft.Colors.RED_600),
                ft.Text(str(_fatal), size=14, selectable=True),
                ft.Text("詳細はlog/フォルダを確認してください", size=12, color=ft.Colors.GREY_600),
                ft.Text(tb, size=10, selectable=True, font_family="Consolas"),
            ], scroll=ft.ScrollMode.AUTO, expand=True))
            page.update()
        except Exception:
            pass
        raise

def _main_inner(page: ft.Page):
    page.title = f"{APP_NAME} - Web Screenshot Tool"
    page.window.width = 1500; page.window.height = 900
    # Set window icon to ebi
    # Flet window icon: .ico が必須 (Windows only, 公式仕様)
    for _icon_ext in ("shot_icon.ico", "shot_icon.png"):
        _icon_path = os.path.join(get_bundle_dir(), "assets", _icon_ext)
        if not os.path.isfile(_icon_path): _icon_path = os.path.join(get_app_dir(), "assets", _icon_ext)
        if os.path.isfile(_icon_path): break
    else:
        _icon_path = None
    if _icon_path and os.path.isfile(_icon_path):
        try: page.window.icon = _icon_path
        except Exception: pass
    page.theme_mode = ft.ThemeMode.LIGHT
    page.theme = ft.Theme(color_scheme_seed=ft.Colors.BLUE)

    # ── Project initialization ──
    _projects_registry = migrate_to_projects()
    _last_proj = _projects_registry.get("last_active", "default")
    if not activate_project(_last_proj, _projects_registry):
        # fallback to first project
        if _projects_registry["projects"]:
            activate_project(_projects_registry["projects"][0]["id"], _projects_registry)

    def _safe_load(loader, default, name):
        try:
            return loader()
        except Exception as x:
            _flog.error(f"Failed to load {name}: {x}")
            return default

    state = {
        "tests": _safe_load(load_tests, [], "tests"),
        "pattern_sets": _safe_load(load_pattern_sets, {}, "pattern_sets"),
        "config": _safe_load(load_config, {}, "config"),
        "selector_bank": _safe_load(load_selector_bank, {}, "selector_bank"),
        "pages": _safe_load(load_pages, [], "pages"),
        "browser_driver": None, "browser_elements": [],
        "selected_test": -1, "selected_pat_set": None, "selected_el": -1, "_prev_el_row": -1,
        "selected_page": None,
        "collapsed": set(),
        "_copied_pat": None,
        "_copied_step": None,
        "stop_event": None, "test_drivers": [], "running": False,
        "selected_test_per_page": {},
        "_tc_id_counter": 0, "_page_id_counter": 0,
    }
    def _reinit_id_counters():
        _max_id = 0
        for tc in state["tests"]:
            if "_id" in tc:
                try: _max_id = max(_max_id, int(tc["_id"].split("_", 1)[1]))
                except (ValueError, IndexError): pass
            else:
                state["_tc_id_counter"] += 1
                tc["_id"] = f"tc_{state['_tc_id_counter']}"
        state["_tc_id_counter"] = max(state["_tc_id_counter"], _max_id)
        _max_page_id = 0
        for pg in state["pages"]:
            try: _max_page_id = max(_max_page_id, int(pg["_id"].split("_", 1)[1]))
            except (ValueError, IndexError): pass
        state["_page_id_counter"] = _max_page_id
    _reinit_id_counters()

    def _new_tc_id():
        state["_tc_id_counter"] += 1; return f"tc_{state['_tc_id_counter']}"
    def _new_page_id():
        state["_page_id_counter"] += 1; return f"p_{state['_page_id_counter']}"

    def cur_page():
        pid = state["selected_page"]
        for p in state["pages"]:
            if p["_id"] == pid: return p
        return None

    # Index cache for O(1) lookups (invalidated on data change)
    _idx_cache = {"valid": False, "by_page": {}, "by_id": {}}
    def _rebuild_idx():
        c = _idx_cache
        c["by_page"] = {}; c["by_id"] = {}
        for i, tc in enumerate(state["tests"]):
            tid = tc.get("_id", "")
            pid = tc.get("page_id", "")
            c["by_id"][tid] = i
            c["by_page"].setdefault(pid, []).append(tc)
        c["valid"] = True
    def _invalidate_idx():
        _idx_cache["valid"] = False
    def tests_for_page(page_id):
        if not _idx_cache["valid"]: _rebuild_idx()
        return list(_idx_cache["by_page"].get(page_id, []))

    def auto_number_tests():
        """Re-number tests: page_number-sub_number.
        If a test has _sub_number set, use it and continue from there."""
        _invalidate_idx()  # data may have changed before this call
        for pg in state["pages"]:
            pnum = pg["number"]
            next_sub = int(_safe_float(pg.get("start_number", 1), 1))
            page_tests = tests_for_page(pg["_id"])
            for tc in page_tests:
                forced = tc.get("_sub_number")
                if forced is not None:
                    try: next_sub = int(forced)
                    except (ValueError, TypeError): pass
                tc["number"] = f"{pnum}-{next_sub}"
                next_sub += 1

    def _ensure_default_page():
        if not state["pages"]:
            dp = {"_id": _new_page_id(), "name": "ページ1", "number": "1", "start_number": 1, "url": ""}
            state["pages"].append(dp)
            for tc in state["tests"]:
                tc.setdefault("page_id", dp["_id"])
    _ensure_default_page()
    auto_number_tests()
    state["selected_page"] = state["pages"][0]["_id"]

    cfg = state["config"]

    # ── Helpers ──
    _init_done = [False]
    _save_timer = [None]
    _save_lock = threading.Lock()
    def schedule_save():
        if not _init_done[0]: return
        if _save_timer[0]: _save_timer[0].cancel()
        def do_save():
            with _save_lock:
                try: save_all()
                except Exception as x: _flog.error(f"auto-save failed: {x}")
        _save_timer[0] = threading.Timer(SAVE_DELAY_SEC, do_save)
        _save_timer[0].daemon = True; _save_timer[0].start()

    def log(msg):
        _flog.info(msg)
        ts = datetime.now().strftime("%H:%M:%S")
        if "[ERROR]" in msg: color = ft.Colors.RED_600
        elif "[WARN]" in msg: color = ft.Colors.ORANGE_700
        else: color = ft.Colors.GREY_700
        log_list.controls.append(ft.Text(f"[{ts}] {msg}", size=11, selectable=True, font_family="Consolas", color=color))
        if len(log_list.controls) > LOG_MAX_LINES: log_list.controls.pop(0)
        try:
            page.update()
            time.sleep(0.02)  # UIスレッドへの反映を待つ
        except Exception: pass
    def _log_error(context, exc):
        _flog.error(f"{context}: {exc}\n{traceback.format_exc()}")
        log(f"[ERROR] {context}: {exc}")
    def _guard_running():
        """Return True and show snack if tests are running (blocks editing)."""
        if state["running"]:
            snack("テスト実行中は編集できません", ft.Colors.ORANGE_700)
            return True
        return False
    def snack(msg, color=ft.Colors.GREEN_700):
        try:
            sb = ft.SnackBar(ft.Text(msg, color=ft.Colors.WHITE), bgcolor=color)
            page.open(sb)
        except Exception: pass
    def open_dlg(d, modal=True):
        d.modal = modal
        try: page.show_dialog(d)
        except RuntimeError:
            try: d.open = False; d.update()
            except Exception: pass
            try: page.show_dialog(d)
            except Exception: pass
    def close_dlg(d):
        try: d.open = False; d.update()
        except Exception: pass
    def save_all():
        save_config(state["config"]); save_tests(state["tests"])
        save_pattern_sets(state["pattern_sets"]); save_selector_bank(state["selector_bank"])
        save_pages(state["pages"])
    def close_browser():
        if state["browser_driver"]:
            kill_driver(state["browser_driver"]); state["browser_driver"] = None
    def get_all_selectors():
        sels = set()
        for elems in state["selector_bank"].values():
            for el in elems: sels.add(el["selector"])
        for el in state["browser_elements"]: sels.add(el["selector"])
        return sorted(sels)
    def cur_test():
        idx = state["selected_test"]
        if 0 <= idx < len(state["tests"]): return state["tests"][idx]
        return None
    def pat_set_names():
        return list(state["pattern_sets"].keys())

    # ================================================================
    # Page selector (Dropdown)
    # ================================================================

    def _page_dd_options():
        return [ft.dropdown.Option(key=pg["_id"], text=f"{pg['number']}. {pg['name']}")
                for pg in state["pages"]]

    def refresh_page_dd(update=True):
        page_dd.options = _page_dd_options()
        page_dd.value = state["selected_page"]
        pg = cur_page()
        if pg:
            n_tests = len(tests_for_page(pg["_id"]))
            pg_url = pg.get("url","")
            url_hint = pg_url[:40] + "..." if len(pg_url) > 40 else pg_url
            page_info_label.value = f"{n_tests}件 | {url_hint}" if url_hint else f"{n_tests}件"
        else:
            page_info_label.value = ""
        if update: page.update()

    def on_page_dd_change(e):
        if not page_dd.value: return
        state["selected_page"] = page_dd.value
        # Restore remembered test selection for this page
        remembered_id = state["selected_test_per_page"].get(page_dd.value, "")
        state["selected_test"] = -1
        if remembered_id:
            for i, tc in enumerate(state["tests"]):
                if tc.get("_id") == remembered_id:
                    state["selected_test"] = i; break
        refresh_page_dd(False); refresh_test_list(False); refresh_steps(False)
        # Auto-sync browser URL: project URL > page URL
        _proj_url = state["config"].get("project_url", "").strip()
        if _proj_url:
            browser_url.value = _proj_url
        else:
            pg = cur_page()
            if pg and pg.get("url", ""):
                browser_url.value = pg["url"]
        page.update()

    def add_page(e):
        if _guard_running(): return
        used_nums = {p["number"] for p in state["pages"]}
        next_num = str(len(state["pages"]) + 1)
        while next_num in used_nums:
            next_num = str(int(next_num) + 1)
        nf = ft.TextField(label="ページ名", width=350, value=f"ページ{next_num}")
        url_f = ft.TextField(label="起点URL", width=450, hint_text="このページの起点URL")
        numf = ft.TextField(label="ページ番号", width=100, value=next_num)
        startf = ft.TextField(label="テスト開始番号", width=100, value="1", hint_text="この番号から連番")
        def on_ok(e):
            try:
                name = nf.value.strip(); num = numf.value.strip()
                if not name: snack("ページ名を入力してください", ft.Colors.RED_700); return
                if not num: snack("ページ番号を入力してください", ft.Colors.RED_700); return
                if any(p["number"] == num for p in state["pages"]):
                    snack(f"番号 {num} は既に使用されています", ft.Colors.RED_700); return
                try: start = int(startf.value.strip() or "1")
                except ValueError: snack("開始番号は整数で入力してください", ft.Colors.RED_700); return
                new_pg = {"_id": _new_page_id(), "name": name, "number": num, "start_number": start,
                          "url": url_f.value.strip()}
                state["pages"].append(new_pg)
                state["selected_page"] = new_pg["_id"]
                auto_number_tests()
                refresh_page_dd(False); refresh_test_list(False); refresh_steps(False)
                page.update(); close_dlg(dlg)
            except Exception as x: _log_error("add_page", x); close_dlg(dlg)
        dlg = ft.AlertDialog(title=ft.Text("ページ追加"),
            content=ft.Column([nf, url_f, ft.Row([numf, startf], spacing=10)], tight=True, spacing=10, width=500),
            actions=[ft.TextButton("OK", on_click=on_ok), ft.TextButton("キャンセル", on_click=lambda e: close_dlg(dlg))])
        open_dlg(dlg)

    def edit_page(e):
        if _guard_running(): return
        pg = cur_page()
        if not pg: snack("ページを選択してください", ft.Colors.ORANGE_700); return
        nf = ft.TextField(label="ページ名", width=350, value=pg["name"])
        url_f = ft.TextField(label="起点URL", width=450, value=pg.get("url",""), hint_text="このページの起点URL")
        numf = ft.TextField(label="ページ番号", width=100, value=pg["number"])
        startf = ft.TextField(label="テスト開始番号", width=100, value=str(pg.get("start_number", 1)))
        def on_ok(e):
            try:
                name = nf.value.strip(); num = numf.value.strip()
                if not name: snack("ページ名を入力してください", ft.Colors.RED_700); return
                if not num: snack("ページ番号を入力してください", ft.Colors.RED_700); return
                if num != pg["number"] and any(p["number"] == num for p in state["pages"]):
                    snack(f"番号 {num} は既に使用されています", ft.Colors.RED_700); return
                try: start = int(startf.value.strip() or "1")
                except ValueError: snack("開始番号は整数で入力してください", ft.Colors.RED_700); return
                pg["name"] = name; pg["number"] = num; pg["start_number"] = start
                pg["url"] = url_f.value.strip()
                auto_number_tests()
                refresh_page_dd(False); refresh_test_list(False); refresh_steps(False)
                page.update(); close_dlg(dlg)
            except Exception as x: _log_error("edit_page", x); close_dlg(dlg)
        dlg = ft.AlertDialog(title=ft.Text("ページ編集"),
            content=ft.Column([nf, url_f, ft.Row([numf, startf], spacing=10)], tight=True, spacing=10, width=500),
            actions=[ft.TextButton("OK", on_click=on_ok), ft.TextButton("キャンセル", on_click=lambda e: close_dlg(dlg))])
        open_dlg(dlg)

    def move_page_up(e):
        if _guard_running(): return
        pg = cur_page()
        if not pg: return
        idx = next((i for i, p in enumerate(state["pages"]) if p["_id"] == pg["_id"]), -1)
        if idx <= 0: snack("先頭です", ft.Colors.ORANGE_700); return
        state["pages"][idx], state["pages"][idx-1] = state["pages"][idx-1], state["pages"][idx]
        auto_number_tests(); refresh_page_dd(False); refresh_test_list(False); refresh_steps(False); page.update()

    def move_page_down(e):
        if _guard_running(): return
        pg = cur_page()
        if not pg: return
        idx = next((i for i, p in enumerate(state["pages"]) if p["_id"] == pg["_id"]), -1)
        if idx < 0 or idx >= len(state["pages"]) - 1: snack("末尾です", ft.Colors.ORANGE_700); return
        state["pages"][idx], state["pages"][idx+1] = state["pages"][idx+1], state["pages"][idx]
        auto_number_tests(); refresh_page_dd(False); refresh_test_list(False); refresh_steps(False); page.update()

    def dup_page(e):
        if _guard_running(): return
        pg = cur_page()
        if not pg: snack("ページを選択してください", ft.Colors.ORANGE_700); return
        # ページを複製
        new_pg = copy.deepcopy(pg)
        new_pg["_id"] = _new_page_id()
        new_pg["name"] += " (コピー)"
        # 番号を自動で空き番号にする
        used = {p["number"] for p in state["pages"]}
        n = int(pg["number"]) + 1 if pg["number"].isdigit() else len(state["pages"]) + 1
        while str(n) in used: n += 1
        new_pg["number"] = str(n)
        # ページ挿入
        idx = next((i for i, p in enumerate(state["pages"]) if p["_id"] == pg["_id"]), len(state["pages"]))
        state["pages"].insert(idx + 1, new_pg)
        # テストケースも複製
        src_tests = tests_for_page(pg["_id"])
        for tc in src_tests:
            new_tc = copy.deepcopy(tc)
            new_tc["_id"] = _new_tc_id()
            new_tc["page_id"] = new_pg["_id"]
            new_tc.pop("_sub_number", None)
            state["tests"].append(new_tc)
        state["selected_page"] = new_pg["_id"]
        auto_number_tests()
        refresh_page_dd(False); refresh_test_list(False); refresh_steps(False)
        page.update()
        snack(f"ページ複製: {new_pg['name']} ({len(src_tests)}テスト)")

    def del_page(e):
        if _guard_running(): return
        if len(state["pages"]) <= 1:
            snack("最後のページは削除できません", ft.Colors.RED_700); return
        pg = cur_page()
        if not pg: return
        pid = pg["_id"]; n_tests = len(tests_for_page(pid))
        def on_yes(e):
            try:
                state["tests"] = [t for t in state["tests"] if t.get("page_id") != pid]
                state["pages"] = [p for p in state["pages"] if p["_id"] != pid]
                if state["selected_page"] == pid:
                    state["selected_page"] = state["pages"][0]["_id"] if state["pages"] else None
                state["selected_test"] = -1; auto_number_tests()
                refresh_page_dd(False); refresh_test_list(False); refresh_steps(False)
                page.update(); close_dlg(dlg)
            except Exception as x: _log_error("del_page", x); close_dlg(dlg)
        msg = f"「{pg['name']}」を削除しますか？"
        if n_tests > 0: msg += f"\n（{n_tests}件のテストケースも削除されます）"
        dlg = ft.AlertDialog(title=ft.Text("ページ削除確認"), content=ft.Text(msg),
            actions=[ft.TextButton("削除", on_click=on_yes, style=ft.ButtonStyle(color=ft.Colors.RED_600)),
                     ft.TextButton("キャンセル", on_click=lambda e: close_dlg(dlg))])
        open_dlg(dlg)

    def run_page(pid=None):
        if pid is None: pid = state["selected_page"]
        pg = None
        for p in state["pages"]:
            if p["_id"] == pid: pg = p; break
        tcs = tests_for_page(pid)
        if not tcs: snack("テストケース0件", ft.Colors.RED_700); return
        label = f"【{pg['number']}_{pg['name']}】" if pg else ""
        n_pats = sum(len(state["pattern_sets"].get(tc.get("pattern","") or "", [])) or 1 for tc in tcs)
        def on_yes(e):
            close_dlg(dlg); _do_run(tcs, label)
        dlg = ft.AlertDialog(title=ft.Text("ページテスト実行"),
            content=ft.Text(f"「{pg['name'] if pg else ''}」の {len(tcs)} テスト（{n_pats}パターン）を実行しますか？"),
            actions=[ft.TextButton("実行", on_click=on_yes), ft.TextButton("キャンセル", on_click=lambda e: close_dlg(dlg))])
        open_dlg(dlg)

    # ================================================================
    # Tab 1: Test Cases (1-column ReorderableListView)
    # ================================================================

    def _find_test_idx(tc_id):
        if not _idx_cache["valid"]: _rebuild_idx()
        return _idx_cache["by_id"].get(tc_id, -1)

    def refresh_test_list(update=True):
        _invalidate_idx()
        test_list.controls.clear()
        cur_pid = state["selected_page"]
        page_tests = tests_for_page(cur_pid) if cur_pid else []
        if not page_tests:
            test_list.controls.append(ft.Container(
                ft.Column([ft.Icon(ft.Icons.ADD_CIRCLE_OUTLINE, size=32, color=ft.Colors.GREY_400),
                    ft.Text("＋ボタンでテストケースを追加", size=12, color=ft.Colors.GREY_500, text_align=ft.TextAlign.CENTER)],
                    horizontal_alignment=ft.CrossAxisAlignment.CENTER, spacing=8),
                padding=ft.Padding(20, 40, 20, 20), key="empty"))
        for page_idx, tc in enumerate(page_tests):
            global_idx = _find_test_idx(tc.get("_id", ""))
            selected = (global_idx == state["selected_test"])
            pat = tc.get("pattern","")
            n_steps = len([s for s in tc.get("steps",[]) if s["type"] not in ("見出し","コメント")])
            n_pats = len(state["pattern_sets"].get(pat,[])) if pat else 0
            subtitle = f"{n_steps}ステップ"
            if pat: subtitle += f" | {pat} ({n_pats}件)"
            tc_id = tc.get("_id", f"tc_fallback_{global_idx}")
            tc_number = tc.get("number", "")
            is_forced = tc.get("_sub_number") is not None
            num_color = ft.Colors.ORANGE_700 if is_forced else ft.Colors.BLUE_600
            card = ft.Container(
                ft.Row([
                    ft.ReorderableDragHandle(content=ft.Icon(ft.Icons.DRAG_INDICATOR, size=16, color=ft.Colors.GREY_400),
                                             mouse_cursor=ft.MouseCursor.GRAB),
                    ft.Text(tc_number, size=11, color=num_color, weight=ft.FontWeight.BOLD, width=50),
                    ft.Column([
                        ft.Text(tc.get("name",""), weight=ft.FontWeight.BOLD, size=12,
                                color=ft.Colors.BLUE_800 if selected else ft.Colors.BLACK,
                                max_lines=1, overflow=ft.TextOverflow.ELLIPSIS),
                        ft.Text(subtitle, size=10, color=ft.Colors.GREY_500),
                    ], spacing=1, expand=True),
                    ft.PopupMenuButton(icon=ft.Icons.MORE_VERT, icon_size=14, icon_color=ft.Colors.GREY_400,
                        tooltip="操作", items=[
                            ft.PopupMenuItem(icon=ft.Icons.PLAY_ARROW, content="実行",
                                on_click=lambda e, tid=tc_id: run_single(_find_test_idx(tid))),
                            ft.PopupMenuItem(icon=ft.Icons.COPY, content="コピー",
                                on_click=lambda e, tid=tc_id: copy_test(_find_test_idx(tid))),
                            ft.PopupMenuItem(),
                            ft.PopupMenuItem(icon=ft.Icons.DELETE, content="削除",
                                on_click=lambda e, tid=tc_id: del_test(_find_test_idx(tid))),
                        ]),
                ], vertical_alignment=ft.CrossAxisAlignment.CENTER, spacing=4),
                bgcolor=ft.Colors.BLUE_50 if selected else None,
                ink=True, ink_color=ft.Colors.BLUE_100,
                padding=ft.Padding(8, 6, 4, 6), border_radius=6,
                border=ft.Border.all(2, ft.Colors.BLUE_300) if selected else ft.Border.all(1, ft.Colors.GREY_200),
                on_click=lambda e, tid=tc_id: select_test(_find_test_idx(tid)),
                key=tc_id)
            test_list.controls.append(card)
        has_any_url = bool(state["config"].get("project_url","").strip()) or any(p.get("url","").strip() for p in state["pages"]) or any(t.get("url","").strip() for t in state["tests"])
        has_tests = len(state["tests"]) > 0
        run_btn.disabled = not (has_any_url and has_tests)
        run_single_btn.disabled = not (has_any_url and has_tests)
        run_page_btn.disabled = not (has_any_url and page_tests)
        schedule_save()
        if update: page.update()

    _reorder_dedup = {}
    def _is_dup_reorder(handler, old, new):
        """Flet fires on_reorder twice per drag. Block same (handler, old, new) within 0.1s."""
        now = time.time()
        key = (handler, old, new)
        prev = _reorder_dedup.get(key)
        if prev and now - prev < 0.1: return True
        _reorder_dedup[key] = now
        if len(_reorder_dedup) > 50:
            stale = [k for k, v in _reorder_dedup.items() if now - v > 1.0]
            for k in stale: del _reorder_dedup[k]
        return False

    def on_test_reorder(e):
        try:
            old, new = e.old_index, e.new_index
            if old is None or new is None: return
            if _is_dup_reorder("test", old, new): return
            cur_pid = state["selected_page"]
            page_tests = tests_for_page(cur_pid) if cur_pid else []
            if not (0 <= old < len(page_tests) and 0 <= new <= len(page_tests)): return
            if old == new: return
            _flog.debug(f"on_test_reorder: old={old} new={new} len={len(page_tests)}")
            # 並び替え前に選択テストのIDを記録
            sel_idx = state["selected_test"]
            sel_id = state["tests"][sel_idx].get("_id") if 0 <= sel_idx < len(state["tests"]) else None
            # Build new page order: pop from old, insert at new
            old_tc = page_tests[old]
            new_order = list(page_tests)
            new_order.pop(old)
            new_order.insert(new, old_tc)
            # Rebuild global list preserving other pages' positions
            result = []
            page_inserted = False
            for t in state["tests"]:
                if t.get("page_id") == cur_pid:
                    if not page_inserted:
                        result.extend(new_order)
                        page_inserted = True
                    # skip original page tests (already added via new_order)
                else:
                    result.append(t)
            if not page_inserted:
                result.extend(new_order)
            state["tests"] = result
            # 選択テストのインデックスをIDで復元
            if sel_id:
                for i, t in enumerate(state["tests"]):
                    if t.get("_id") == sel_id:
                        state["selected_test"] = i
                        break
            auto_number_tests()
            refresh_test_list(False); refresh_page_dd(False); schedule_save(); page.update()
        except Exception as x: _log_error("on_test_reorder", x)

    def _update_test_highlight():
        """Update visual selection state on existing cards without rebuilding.
        Container structure: Container(key=tc_id) > Row > [DragHandle, Text(num), Column > [Text(name), Text(sub)], PopupMenu]"""
        sel_tc = state["tests"][state["selected_test"]] if 0 <= state["selected_test"] < len(state["tests"]) else None
        sel_id = sel_tc.get("_id") if sel_tc else None
        for ctrl in test_list.controls:
            if not isinstance(ctrl, ft.Container) or ctrl.key is None: continue
            is_sel = (ctrl.key == sel_id)
            ctrl.bgcolor = ft.Colors.BLUE_50 if is_sel else None
            ctrl.border = ft.Border.all(2, ft.Colors.BLUE_300) if is_sel else ft.Border.all(1, ft.Colors.GREY_200)
            try:
                row = ctrl.content  # Row
                name_txt = row.controls[2].controls[0]  # [0]=DragHandle, [1]=Text(num), [2]=Column > [0]=Text(name)
                name_txt.color = ft.Colors.BLUE_800 if is_sel else ft.Colors.BLACK
            except (AttributeError, IndexError):
                pass

    def select_test(idx):
        if idx < 0 or idx >= len(state["tests"]): return
        if idx == state["selected_test"]: return  # already selected
        state["selected_test"] = idx; state["collapsed"] = set()
        # Remember selection per page
        tc = state["tests"][idx]
        pid = tc.get("page_id", "")
        if pid: state["selected_test_per_page"][pid] = tc.get("_id", "")
        _update_test_highlight()
        refresh_steps(False); page.update()

    def add_test(e):
        if _guard_running(): return
        cur_pid = state["selected_page"]
        if not cur_pid: snack("ページを選択してください", ft.Colors.ORANGE_700); return
        new_tc = {"name": f"テスト{len(tests_for_page(cur_pid))+1}", "pattern": None, "steps": [],
                  "_id": _new_tc_id(), "page_id": cur_pid, "number": ""}
        state["tests"].append(new_tc); auto_number_tests()
        state["selected_test"] = len(state["tests"]) - 1
        refresh_page_dd(False); refresh_test_list(False); refresh_steps(False)
        try: page.update()
        except Exception: pass

    def copy_test(idx):
        if _guard_running(): return
        if 0 <= idx < len(state["tests"]):
            tc = copy.deepcopy(state["tests"][idx])
            tc["name"] += " (コピー)"; tc["_id"] = _new_tc_id(); tc.pop("_sub_number", None)
            state["tests"].insert(idx + 1, tc); auto_number_tests()
            state["selected_test"] = idx + 1
            refresh_page_dd(False); refresh_test_list(False); refresh_steps(False)
            try: page.update()
            except Exception: pass

    def del_test(idx):
        if _guard_running(): return
        if not (0 <= idx < len(state["tests"])): return
        tc = state["tests"][idx]; name = tc.get("name", f"テスト{idx+1}")
        pat = tc.get("pattern"); pat_orphan = False
        if pat and pat in state["pattern_sets"]:
            pat_orphan = not any(t.get("pattern") == pat for i, t in enumerate(state["tests"]) if i != idx)
        del_pat_cb = ft.Checkbox(label=f"パターンセット「{pat}」も削除", value=True, visible=pat_orphan)
        def on_yes(e):
            try:
                if pat_orphan and del_pat_cb.value and pat in state["pattern_sets"]:
                    del state["pattern_sets"][pat]
                    if state["selected_pat_set"] == pat: state["selected_pat_set"] = None
                state["tests"].pop(idx); auto_number_tests()
                if state["selected_test"] >= len(state["tests"]):
                    state["selected_test"] = len(state["tests"]) - 1  # -1 when empty
                refresh_page_dd(False); refresh_test_list(False); refresh_steps(False)
                refresh_pat_set_list(False); refresh_pats(False); close_dlg(dlg)
                try: page.update()
                except Exception: pass
            except Exception as x: _log_error("del_test", x); close_dlg(dlg)
        content = ft.Column([ft.Text(f"「{name}」を削除しますか？"), del_pat_cb],
                            tight=True, spacing=8) if pat_orphan else ft.Text(f"「{name}」を削除しますか？")
        dlg = ft.AlertDialog(title=ft.Text("削除確認"), content=content,
            actions=[ft.TextButton("削除", on_click=on_yes, style=ft.ButtonStyle(color=ft.Colors.RED_600)),
                     ft.TextButton("キャンセル", on_click=lambda e: close_dlg(dlg))])
        open_dlg(dlg)

    def edit_test_name(e):
        if _guard_running(): return
        tc = cur_test()
        if not tc: return
        nf = ft.TextField(label="テスト名", value=tc["name"], width=350)
        # 枝番: ページ番号は自動、枝番のみ編集可
        pg = None
        for p in state["pages"]:
            if p["_id"] == tc.get("page_id"): pg = p; break
        pg_prefix = pg["number"] if pg else "?"
        cur_sub = tc.get("_sub_number")
        sub_f = ft.TextField(label="枝番", width=100,
                             value=str(cur_sub) if cur_sub is not None else "",
                             hint_text="自動", keyboard_type=ft.KeyboardType.NUMBER)
        num_preview = ft.Text(f"現在: {tc.get('number','')}", size=11, color=ft.Colors.GREY_600)
        pat_opts = [ft.dropdown.Option("なし")] + [ft.dropdown.Option(n) for n in pat_set_names()]
        pf = ft.Dropdown(label="パターンセット", width=350, value=tc.get("pattern") or "なし", options=pat_opts)
        page_opts = [ft.dropdown.Option(key=pg2["_id"], text=f"{pg2['number']} {pg2['name']}") for pg2 in state["pages"]]
        page_dd_edit = ft.Dropdown(label="ページ", width=350, value=tc.get("page_id", ""), options=page_opts)
        tc_url_f = ft.TextField(label="開始URL（空欄でページURLを使用）", width=450, value=tc.get("url",""),
                                hint_text="個別URLが必要な場合のみ")
        def on_ok(e):
            try:
                tc["name"] = nf.value
                tc["pattern"] = None if pf.value == "なし" else pf.value
                tc["url"] = tc_url_f.value.strip()
                new_pid = page_dd_edit.value
                if new_pid and new_pid != tc.get("page_id"):
                    tc["page_id"] = new_pid; state["selected_page"] = new_pid
                sub_val = sub_f.value.strip()
                if sub_val:
                    try:
                        tc["_sub_number"] = int(sub_val)
                    except ValueError:
                        snack("枝番は整数で入力してください", ft.Colors.RED_700); return
                else:
                    tc.pop("_sub_number", None)
                auto_number_tests()
                refresh_page_dd(False); refresh_test_list(False); refresh_steps(); close_dlg(dlg)
            except Exception as x: _log_error("edit_test_name", x); close_dlg(dlg)
        dlg = ft.AlertDialog(title=ft.Text("テストケース設定"),
            content=ft.Column([nf,
                ft.Row([ft.Text(f"{pg_prefix} -", size=14, weight=ft.FontWeight.BOLD), sub_f, num_preview],
                       spacing=8, vertical_alignment=ft.CrossAxisAlignment.CENTER),
                ft.Text("枝番を指定すると、以降のテストも自動で連番になります。空欄で自動。",
                         size=11, color=ft.Colors.GREY_500),
                page_dd_edit, tc_url_f, pf], tight=True, spacing=10, width=500),
            actions=[ft.TextButton("OK", on_click=on_ok), ft.TextButton("キャンセル", on_click=lambda e: close_dlg(dlg))])
        open_dlg(dlg)

    # ── Steps ──
    def refresh_steps(update=True):
        step_reorder.controls.clear()
        tc = cur_test()
        if not tc:
            tc_header.value = "テストケースを選択してください"; tc_pattern_label.value = ""
            if update: page.update()
            return
        tc_header.value = f"{tc.get('number','')} {tc.get('name','')}"
        pat = tc.get("pattern")
        tc_pattern_label.value = f"パターン: {pat} ({len(state['pattern_sets'].get(pat,[]))}件)" if pat else "パターン: なし (1回実行)"
        if not tc.get("steps"):
            step_reorder.controls.append(ft.Container(
                ft.Column([ft.Icon(ft.Icons.TOUCH_APP, size=28, color=ft.Colors.GREY_400),
                    ft.Text("要素ブラウザからクイック追加\nまたは＋ボタンでステップ追加", size=11,
                            color=ft.Colors.GREY_500, text_align=ft.TextAlign.CENTER)],
                    horizontal_alignment=ft.CrossAxisAlignment.CENTER, spacing=6),
                padding=ft.Padding(20, 30, 20, 20), key="empty_steps"))
        collapsed = state["collapsed"]; hidden = False
        for i, s in enumerate(tc.get("steps",[])):
            t = s["type"]; key = str(i)
            if t == "見出し":
                sid = i; is_c = sid in collapsed; hidden = is_c
                ic = ft.Icons.EXPAND_LESS if not is_c else ft.Icons.EXPAND_MORE
                step_reorder.controls.append(ft.Container(ft.Row([
                    ft.Icon(ft.Icons.TITLE, color=ft.Colors.BLUE_800, size=16),
                    ft.Text(s.get("text",""), weight=ft.FontWeight.BOLD, size=13, color=ft.Colors.BLUE_800, expand=True),
                    ft.IconButton(ic, icon_size=16, on_click=lambda e, sid=sid: toggle_sec(sid)),
                    ft.IconButton(ft.Icons.COPY, icon_size=14, tooltip="コピー", on_click=lambda e, idx=i: copy_step(idx)),
                    ft.IconButton(ft.Icons.EDIT, icon_size=14, on_click=lambda e, idx=i: show_step_dlg(idx)),
                    ft.IconButton(ft.Icons.DELETE, icon_size=14, on_click=lambda e, idx=i: del_step(idx)),
                ], spacing=4, vertical_alignment=ft.CrossAxisAlignment.CENTER),
                    bgcolor=ft.Colors.BLUE_50, padding=ft.Padding(8,4,36,4), border_radius=4, key=key, height=36))
            elif t == "コメント":
                if hidden: continue
                step_reorder.controls.append(ft.Container(ft.Row([
                    ft.Icon(ft.Icons.COMMENT, color=ft.Colors.GREY_400, size=14),
                    ft.Text(s.get("text",""), size=11, italic=True, color=ft.Colors.GREY_500, expand=True),
                    ft.IconButton(ft.Icons.COPY, icon_size=14, tooltip="コピー", on_click=lambda e, idx=i: copy_step(idx)),
                    ft.IconButton(ft.Icons.EDIT, icon_size=14, on_click=lambda e, idx=i: show_step_dlg(idx)),
                    ft.IconButton(ft.Icons.DELETE, icon_size=14, on_click=lambda e, idx=i: del_step(idx)),
                ], spacing=4, vertical_alignment=ft.CrossAxisAlignment.CENTER),
                    padding=ft.Padding(8,2,36,2), key=key, height=28))
            else:
                if hidden: continue
                step_reorder.controls.append(ft.Container(ft.Row([
                    ft.Icon(STEP_ICONS.get(t, ft.Icons.HELP), color=ft.Colors.BLUE_600, size=16),
                    ft.Text({"アラートOK":"確認OK","アラートキャンセル":"確認×","ナビゲーション":"ナビ"}.get(t,t), size=10, color=ft.Colors.GREY_500, width=38),
                    ft.Text(step_short(s), size=11, expand=True, max_lines=1, overflow=ft.TextOverflow.ELLIPSIS, tooltip=step_short(s)),
                    ft.IconButton(ft.Icons.COPY, icon_size=14, tooltip="コピー", on_click=lambda e, idx=i: copy_step(idx)),
                    ft.IconButton(ft.Icons.EDIT, icon_size=14, on_click=lambda e, idx=i: show_step_dlg(idx)),
                    ft.IconButton(ft.Icons.DELETE, icon_size=14, on_click=lambda e, idx=i: del_step(idx)),
                ], spacing=4, vertical_alignment=ft.CrossAxisAlignment.CENTER),
                    padding=ft.Padding(8,2,36,2), key=key, height=28))
        schedule_save()
        if update: page.update()

    def on_reorder(e):
        try:
            tc = cur_test()
            if not tc: return
            old, new = e.old_index, e.new_index
            if old is None or new is None: return
            if _is_dup_reorder("step", old, new): return
            steps = tc["steps"]; vis = _get_vis(steps)
            if not (0 <= old < len(vis) and 0 <= new <= len(vis)): return
            if old == new: return
            _flog.debug(f"on_reorder(step): old={old} new={new} vis_len={len(vis)}")
            # Extract visible items, reorder, write back
            vis_items = [steps[i] for i in vis]
            item = vis_items.pop(old)
            vis_items.insert(new, item)
            for slot, reordered in zip(vis, vis_items):
                steps[slot] = reordered
            refresh_steps()
        except Exception as x: _log_error("on_reorder", x)

    def _get_vis(steps):
        vis, hidden = [], False
        for i, s in enumerate(steps):
            if s["type"] == "見出し": hidden = i in state["collapsed"]; vis.append(i)
            elif not hidden: vis.append(i)
        return vis

    def toggle_sec(sid):
        state["collapsed"].symmetric_difference_update({sid}); refresh_steps()

    def copy_step(idx):
        tc = cur_test()
        if not tc or not (0 <= idx < len(tc["steps"])): return
        state["_copied_step"] = copy.deepcopy(tc["steps"][idx])
        snack(f"ステップをコピー: {step_short(tc['steps'][idx])}")

    def paste_step(e):
        if _guard_running(): return
        tc = cur_test()
        if not tc: snack("テストケースを選択してください", ft.Colors.ORANGE_700); return
        s = state.get("_copied_step")
        if not s: snack("コピーされたステップなし", ft.Colors.ORANGE_700); return
        tc["steps"].append(copy.deepcopy(s))
        refresh_steps(False); refresh_test_list()
        try: step_reorder.scroll_to(offset=-1, duration=200)
        except Exception: pass
        snack(f"ステップを貼り付け: {step_short(s)}")

    def del_step(idx):
        if _guard_running(): return
        tc = cur_test()
        if not tc or not (0 <= idx < len(tc["steps"])): return
        def do_delete():
            tc["steps"].pop(idx)
            state["collapsed"] = {c if c < idx else c-1 for c in state["collapsed"] if c != idx}
            refresh_steps()
        if state["config"].get("confirm_step_delete", "1") != "1":
            do_delete(); return
        step = tc["steps"][idx]
        label = step_short(step)
        if len(label) > 30: label = label[:27] + "..."
        def on_yes(e):
            do_delete(); close_dlg(dlg)
        dlg = ft.AlertDialog(title=ft.Text("ステップ削除"),
            content=ft.Text(f"「{step['type']}: {label}」を削除しますか？"),
            actions=[ft.TextButton("削除", on_click=on_yes, style=ft.ButtonStyle(color=ft.Colors.RED_600)),
                     ft.TextButton("キャンセル", on_click=lambda e: close_dlg(dlg))])
        open_dlg(dlg)

    def show_step_dlg(idx):
        if _guard_running(): return
        tc = cur_test()
        if not tc: return
        if idx is not None and idx >= len(tc.get("steps", [])): return
        init = tc["steps"][idx] if idx is not None else {}
        t0 = init.get("type","入力")
        type_dd = ft.Dropdown(label="種類", width=160, value=t0, options=[ft.dropdown.Option(t) for t in STEP_TYPES])
        all_sels = get_all_selectors()
        sel_field = ft.TextField(label="セレクタ", expand=True, value=init.get("selector",""))
        def _pick_selector(e):
            if not all_sels: snack("要素を読み込んでください", ft.Colors.ORANGE_700); return
            def on_sel(s):
                sel_field.value = s; page.update(); close_dlg(pick_dlg)
            pick_dlg = ft.AlertDialog(title=ft.Text("セレクタ選択"),
                content=ft.Column([
                    ft.TextField(label="絞り込み", dense=True, on_change=lambda e: _filter_pick(e, pick_list, all_sels)),
                    pick_list := ft.ListView(controls=[
                        ft.TextButton(s, on_click=lambda e, s=s: on_sel(s), tooltip=s)
                        for s in all_sels], height=300, spacing=0)
                ], tight=True, spacing=8, width=500),
                actions=[ft.TextButton("閉じる", on_click=lambda e: close_dlg(pick_dlg))])
            open_dlg(pick_dlg, modal=False)
        def _filter_pick(e, pick_list, all_sels):
            q = (e.control.value or "").strip().lower()
            pick_list.controls = [ft.TextButton(s, on_click=lambda e, s=s: (
                sel_field.__setattr__("value", s), page.update()), tooltip=s)
                for s in all_sels if not q or q in s.lower()]
            page.update()
        sel_pick_btn = ft.IconButton(ft.Icons.LIST, tooltip="一覧から選択", icon_size=18, on_click=_pick_selector)
        init_val = init.get("value","")
        pat_names = pat_set_names()
        init_val_mode = "パターン" if init_val == "{パターン}" else "手入力"
        val_mode = ft.Dropdown(label="値の指定方法", width=200, value=init_val_mode,
            options=[ft.dropdown.Option("手入力"), ft.dropdown.Option("パターン")])
        init_pat_sel = tc.get("pattern", "") if init_val == "{パターン}" else (pat_names[0] if pat_names else "")
        pat_select = ft.Dropdown(label="パターンセット", width=240, value=init_pat_sel,
            options=[ft.dropdown.Option(n) for n in pat_names])
        val_field = ft.TextField(label="値 (固定値を直接入力)", width=450,
            value="" if init_val == "{パターン}" else init_val,
            multiline=True, min_lines=2, max_lines=4)
        # 入力モード
        input_mode_dd = ft.Dropdown(label="入力モード", width=200, value=init.get("input_mode","overwrite"),
            options=[ft.dropdown.Option(key=k, text=t) for k, t in INPUT_MODES])
        # ナビゲーションURL
        nav_url_f = ft.TextField(label="遷移先URL", width=450, value=init.get("url",""),
            hint_text="https://... ({パターン}可)")
        sec_field = ft.TextField(label="秒数", width=120, value=init.get("seconds","1.0"))
        # スクショモード: key=内部値, text=表示名 (区切り線でグループ化)
        _ss_options = [
            ft.dropdown.Option(key="fullpage", text="表示範囲のみ"),
            ft.dropdown.Option(key="fullshot", text="ページ全体 (縦長)"),
            ft.Divider(height=1),
            ft.dropdown.Option(key="element", text="要素のみ"),
            ft.dropdown.Option(key="margin", text="要素 + 余白"),
        ]
        mode_dd = ft.Dropdown(label="スクショ範囲", width=220, value=init.get("mode","fullpage"),
            options=_ss_options)
        margin_f = ft.TextField(label="マージン(px)", width=120, value=init.get("margin_px","500"))
        # 検証タイプ
        _VERIFY_TYPES = [
            ("post", "POST値"),
            ("state", "要素の状態"),
            ("attrs", "要素の属性"),
        ]
        verify_dd = ft.Dropdown(label="検証タイプ", width=220, value=init.get("verify_type","attrs"),
            options=[ft.dropdown.Option(key=k, text=t) for k, t in _VERIFY_TYPES])
        text_f = ft.TextField(label="テキスト", width=450, value=init.get("text",""), multiline=True, min_lines=1, max_lines=3)
        # Scroll controls
        scroll_mode_dd = ft.Dropdown(label="スクロール方法", width=220, value=init.get("scroll_mode","element"),
            options=[ft.dropdown.Option(key=k, text=t) for k, t in SCROLL_MODES])
        scroll_px_f = ft.TextField(label="位置(px)", width=120, value=init.get("scroll_px","0"), hint_text="上端からのpx")
        # Groups must be defined BEFORE upd() references them
        input_group = ft.Column([ft.Row([sel_field, sel_pick_btn], spacing=4), input_mode_dd, ft.Row([val_mode, pat_select], spacing=8), val_field], spacing=8, tight=True)
        nav_group = ft.Column([nav_url_f], spacing=8, tight=True)
        time_group = ft.Column([sec_field], spacing=8, tight=True)
        ss_group = ft.Column([ft.Row([mode_dd, margin_f], spacing=8)], spacing=8, tight=True)
        verify_group = ft.Column([verify_dd], spacing=8, tight=True)
        scroll_group = ft.Column([ft.Row([scroll_mode_dd, scroll_px_f], spacing=8)], spacing=8, tight=True)
        text_group = ft.Column([text_f], spacing=8, tight=True)
        def upd(e=None):
            try:
                t = type_dd.value
                is_input = (t == "入力")
                needs_sel = t in ("入力","クリック","ホバー","選択","要素待機") or (t=="スクショ" and mode_dd.value in ("element","margin")) or (t=="検証" and verify_dd.value in ("state","attrs")) or (t=="スクロール" and scroll_mode_dd.value=="element")
                sel_field.visible = needs_sel
                input_mode_dd.visible = is_input
                _needs_val = is_input or t == "選択"
                val_mode.visible = _needs_val
                pat_select.visible = _needs_val and val_mode.value == "パターン"
                val_field.visible = _needs_val and val_mode.value == "手入力"
                if is_input and input_mode_dd.value == "clear":
                    val_mode.visible = False; pat_select.visible = False; val_field.visible = False
                nav_url_f.visible = (t == "ナビゲーション")
                sec_field.visible = t in ("待機", "戻る", "更新", "要素待機"); mode_dd.visible = (t=="スクショ"); verify_dd.visible = (t=="検証")
                scroll_mode_dd.visible = (t == "スクロール")
                scroll_px_f.visible = (t == "スクロール" and scroll_mode_dd.value == "pixel")
                if t == "要素待機": sec_field.label = "タイムアウト(秒)"
                else: sec_field.label = "秒数"
                margin_f.visible = (t=="スクショ" and mode_dd.value=="margin")
                text_f.visible = t in ("見出し","コメント")
                if t == "選択": val_mode.label = "選択肢の指定方法"
                else: val_mode.label = "値の指定方法"
                # Hide empty groups entirely
                input_group.visible = needs_sel or is_input or t == "選択"
                nav_group.visible = (t == "ナビゲーション")
                time_group.visible = t in ("待機", "戻る", "更新", "要素待機")
                ss_group.visible = (t == "スクショ")
                verify_group.visible = (t == "検証")
                scroll_group.visible = (t == "スクロール")
                text_group.visible = t in ("見出し","コメント")
                try: page.update()
                except Exception: pass
            except Exception as x: _log_error("show_step_dlg.upd", x)
        type_dd.on_select = upd; mode_dd.on_select = upd; verify_dd.on_select = upd; val_mode.on_select = upd; input_mode_dd.on_select = upd; scroll_mode_dd.on_select = upd
        # 初期表示を正しく設定
        upd()
        def on_ok(e):
            try:
                t = type_dd.value; step = {"type": t}
                if t in ("見出し","コメント"): step["text"] = text_f.value
                elif t in ("入力","クリック","ホバー","選択"):
                    s = sel_field.value if hasattr(sel_field,'value') else ""
                    if not s: snack("セレクタを入力してください", ft.Colors.RED_700); return
                    step["selector"] = s
                    if t == "入力":
                        step["input_mode"] = input_mode_dd.value
                        if input_mode_dd.value != "clear":
                            if val_mode.value == "手入力": step["value"] = val_field.value
                            else:
                                pn = pat_select.value
                                if not pn or pn not in state["pattern_sets"]:
                                    snack("パターンセットを選択してください", ft.Colors.RED_700); return
                                existing = tc.get("pattern")
                                if existing and existing != pn:
                                    snack(f"このテストには既に「{existing}」が設定されています（1テスト1パターン）", ft.Colors.RED_700); return
                                step["value"] = "{パターン}"; tc["pattern"] = pn
                    elif t == "選択":
                        if val_mode.value == "手入力": step["value"] = val_field.value
                        else:
                            pn = pat_select.value
                            if not pn or pn not in state["pattern_sets"]:
                                snack("パターンセットを選択してください", ft.Colors.RED_700); return
                            existing = tc.get("pattern")
                            if existing and existing != pn:
                                snack(f"このテストには既に「{existing}」が設定されています（1テスト1パターン）", ft.Colors.RED_700); return
                            step["value"] = "{パターン}"; tc["pattern"] = pn
                elif t in ("戻る", "更新"):
                    try: step["seconds"] = str(float(sec_field.value))
                    except Exception: snack("秒数を正しく入力してください", ft.Colors.RED_700); return
                elif t == "ナビゲーション":
                    url = nav_url_f.value.strip()
                    if not url: snack("URLを入力してください", ft.Colors.RED_700); return
                    step["url"] = url
                elif t == "待機":
                    try: step["seconds"] = str(float(sec_field.value))
                    except Exception: snack("秒数を正しく入力してください", ft.Colors.RED_700); return
                elif t == "要素待機":
                    s = sel_field.value if hasattr(sel_field,'value') else ""
                    if not s: snack("セレクタを入力してください", ft.Colors.RED_700); return
                    step["selector"] = s
                    try: step["seconds"] = str(float(sec_field.value or "10"))
                    except Exception: snack("秒数を正しく入力してください", ft.Colors.RED_700); return
                elif t == "スクロール":
                    step["scroll_mode"] = scroll_mode_dd.value
                    if scroll_mode_dd.value == "element":
                        s = sel_field.value if hasattr(sel_field,'value') else ""
                        if not s: snack("セレクタを入力してください", ft.Colors.RED_700); return
                        step["selector"] = s
                    elif scroll_mode_dd.value == "pixel":
                        try: step["scroll_px"] = str(int(scroll_px_f.value or "0"))
                        except Exception: snack("整数で入力してください", ft.Colors.RED_700); return
                elif t == "スクショ":
                    step["mode"] = mode_dd.value
                    s = sel_field.value if hasattr(sel_field,'value') else ""
                    if mode_dd.value in ("element","margin") and not s: snack("セレクタを入力してください", ft.Colors.RED_700); return
                    if s: step["selector"] = s
                    if mode_dd.value == "margin":
                        try: step["margin_px"] = str(int(margin_f.value))
                        except Exception: snack("整数で入力してください", ft.Colors.RED_700); return
                elif t == "検証":
                    step["verify_type"] = verify_dd.value
                    s = sel_field.value if hasattr(sel_field,'value') else ""
                    if verify_dd.value in ("state","attrs") and not s: snack("セレクタを入力してください", ft.Colors.RED_700); return
                    if s: step["selector"] = s
                if idx is not None: tc["steps"][idx] = step
                else:
                    tc["steps"].append(step)
                    try: step_reorder.scroll_to(offset=-1, duration=200)
                    except Exception: pass
                refresh_steps(False); refresh_test_list(); close_dlg(dlg)
            except Exception as x: _log_error("show_step_dlg.on_ok", x); close_dlg(dlg)
        dlg = ft.AlertDialog(title=ft.Text("ステップ編集" if idx is not None else "ステップ追加"),
            content=ft.Column([type_dd, text_group, input_group, nav_group, time_group, scroll_group, ss_group, verify_group],
                tight=True, spacing=12, scroll=ft.ScrollMode.AUTO, width=500),
            actions=[ft.TextButton("OK", on_click=on_ok), ft.TextButton("キャンセル", on_click=lambda e: close_dlg(dlg))])
        open_dlg(dlg)

    # ── Element browser ──
    def _el_loading_start(msg="読込中..."):
        el_loading.visible = True; load_btn.disabled = True
        el_status.value = msg; el_status.color = ft.Colors.BLUE_600
        page.update()
    def _el_loading_end(msg="", color=ft.Colors.GREY_500):
        el_loading.visible = False; load_btn.disabled = False
        if msg: el_status.value = msg; el_status.color = color
        page.update()
    def load_page_click(e):
        url = browser_url.value
        if not url: snack("URLを入力してください", ft.Colors.RED_700); return
        _el_loading_start("ブラウザ起動中...")
        page.run_thread(do_load_page, url)
    def do_load_page(url):
        try:
            from selenium import webdriver
            if state["browser_driver"] is None:
                _br_opts = webdriver.ChromeOptions()
                _br_opts.add_argument("--ignore-certificate-errors")
                _br_opts.add_argument("--allow-insecure-localhost")
                _br_opts.add_argument("--disable-features=HttpsUpgrades")
                _br_opts.add_argument("--disable-blink-features=AutomationControlled")
                _br_opts.add_experimental_option("excludeSwitches", ["enable-automation"])
                _br_opts.add_argument("--disable-notifications")
                state["browser_driver"] = webdriver.Chrome(options=_br_opts); state["browser_driver"].set_window_size(1280,900)
                try: state["browser_driver"].execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {"source": "Object.defineProperty(navigator,'webdriver',{get:()=>false});"})
                except Exception: pass
            ba = state["config"].get("basic_auth_user","").strip()
            if ba:
                setup_basic_auth(state["browser_driver"], state["config"])
            el_status.value = "ページ読込中..."; page.update()
            lu = build_auth_url(url, ba, state["config"].get("basic_auth_pass","")) if ba else url
            state["browser_driver"].get(lu)
            try: w = float(browser_wait.value)
            except Exception: w = 3.0
            time.sleep(w); log(f"[DEBUG] {state['browser_driver'].title}")
            el_status.value = "要素を収集中..."; page.update()
            _do_collect_elements(url)
        except Exception as x:
            if state["browser_driver"]:
                try: state["browser_driver"].title
                except Exception:
                    kill_driver(state["browser_driver"]); state["browser_driver"] = None
            log(f"[ERROR] {x}")
        finally: _el_loading_end()
    def _do_collect_elements(url=None):
        """Collect elements from current DOM state (no page navigation)."""
        drv = state["browser_driver"]
        elems = collect_elements_js(drv, include_hidden=True)
        _flog.info(f"[要素] メインフレーム: {len(elems)} 要素")
        # Detect iframes and collect their elements too
        from selenium.webdriver.common.by import By
        try:
            iframes = drv.find_elements(By.CSS_SELECTOR, "iframe, frame")
            _flog.info(f"[要素] iframe検出: {len(iframes)} 件")
            for fi, iframe in enumerate(iframes):
                frame_id = iframe.get_attribute("id") or iframe.get_attribute("name") or f"frame_{fi}"
                try:
                    drv.switch_to.frame(iframe)
                    frame_elems = collect_elements_js(drv, include_hidden=True)
                    _flog.info(f"[要素] iframe[{frame_id}]: {len(frame_elems)} 要素")
                    for fe in frame_elems:
                        fe["hint"] = f"[iframe:{frame_id}] " + fe.get("hint", "")
                        fe["_frame"] = frame_id
                        fe["_frame_index"] = fi
                    elems.extend(frame_elems)
                    drv.switch_to.default_content()
                except Exception as _ifr_err:
                    _flog.warning(f"[要素] iframe[{frame_id}] 収集失敗: {_ifr_err}")
                    try: drv.switch_to.default_content()
                    except Exception: pass
        except Exception as _ifr_outer_err:
            _flog.warning(f"[要素] iframe検出失敗: {_ifr_outer_err}")
        state["browser_elements"] = list(elems)
        if url:
            bank = state["selector_bank"]
            bank[url.split("?")[0]] = [el for el in elems if el.get("visible", True)]
            # A4: LRU limit — keep only newest 50 URLs
            if len(bank) > BANK_MAX_URLS:
                keys = list(bank.keys())
                for old_key in keys[:len(keys) - BANK_MAX_URLS]:
                    del bank[old_key]
        filter_el_table(); schedule_save()
        # Debug: report checkbox/radio hint coverage
        cb_radio = [e for e in elems if e.get("type", "").lower() in ("checkbox", "radio")]
        cb_with_hint = [e for e in cb_radio if e.get("hint", "") and e["hint"] != e.get("type", "")]
        log(f"[要素] DOM再取得 {len(elems)} 要素" + (f" (checkbox/radio: {len(cb_radio)}件, ヒントあり: {len(cb_with_hint)}件)" if cb_radio else ""))
    def reload_dom_click(e):
        """Re-collect elements from current DOM without navigating."""
        if not state["browser_driver"]: snack("先にページを読み込んでください", ft.Colors.ORANGE_700); return
        _el_loading_start("DOM再取得中...")
        try:
            _do_collect_elements()
        except Exception as x:
            log(f"[ERROR] DOM再取得失敗: {x}")
        finally: _el_loading_end()
    def on_el_sort_change(key):
        try:
            el_sort_dd.data = key
            filter_el_table()
        except Exception as x: _log_error("on_el_sort_change", x)

    _el_idx_to_row = {}  # 要素index → 行index のマッピング（filter_el_tableで構築）

    def filter_el_table(update=True):
        """Filter and display elements based on search text, hidden visibility, and sort."""
        _el_idx_to_row.clear()
        el_table.rows.clear()
        query = (el_search.value or "").strip().lower()
        show_hidden = el_show_hidden.value
        visible_count = 0
        total_count = len(state["browser_elements"])
        hidden_count = sum(1 for el in state["browser_elements"] if not el.get("visible", True))
        # Build sorted index
        sort_key = el_sort_dd.data or "dom"
        indexed_els = list(enumerate(state["browser_elements"]))
        if sort_key == "tag":
            indexed_els.sort(key=lambda x: (x[1].get("tag",""), x[1].get("type","")))
        elif sort_key == "type":
            indexed_els.sort(key=lambda x: (x[1].get("type",""), x[1].get("tag","")))
        elif sort_key == "id":
            indexed_els.sort(key=lambda x: (x[1].get("id","") or x[1].get("name","") or "zzz"))
        for i, el in indexed_els:
            is_visible = el.get("visible", True)
            # Filter: hidden visibility
            if not is_visible and not show_hidden:
                continue
            # Filter: search query (match against tag, type, id, name, hint, selector)
            if query:
                searchable = " ".join([
                    el.get("tag", ""), el.get("type", ""), el.get("id", ""),
                    el.get("name", ""), el.get("hint", ""), el.get("selector", "")
                ]).lower()
                if query not in searchable:
                    continue
            _el_idx_to_row[i] = visible_count
            visible_count += 1
            # Row color: dim for hidden elements
            row_color = ft.Colors.ORANGE_50 if not is_visible else None
            reason = el.get("hidden_reason", "")
            vis_indicator = "" if is_visible else f" [{reason}]" if reason else " [hidden]"
            is_selected = (i == state["selected_el"])
            hint_text = el.get("hint", "")
            hint_display = hint_text[:25] if hint_text else ""
            if not is_visible: hint_display += " *"
            hint_tooltip = hint_text + (f"\n{vis_indicator.strip()}" if vis_indicator else "")
            id_or_name = el.get("id") or el.get("name", "")
            sel_text = el["selector"]
            # Build label/validation info column
            meta = el.get("meta", {})
            label_text = meta.get("label", "")
            badges = []
            if meta.get("required"): badges.append("*必須")
            if meta.get("maxlength"): badges.append(f"max{meta['maxlength']}")
            if meta.get("pattern"): badges.append("正規表現")
            if meta.get("disabled"): badges.append("無効")
            if meta.get("readonly"): badges.append("読取専用")
            if meta.get("option_count"): badges.append(f"{meta['option_count']}件")
            info_parts = []
            if label_text: info_parts.append(label_text)
            if badges: info_parts.append(" ".join(badges))
            info_text = " | ".join(info_parts) if info_parts else ""
            info_color = ft.Colors.RED_700 if meta.get("required") else ft.Colors.GREY_600
            el_table.rows.append(ft.DataRow(
                cells=[ft.DataCell(ft.Text(el["tag"],size=11)),
                       ft.DataCell(ft.Text(el.get("type",""),size=11)),
                       ft.DataCell(ft.Text(id_or_name,size=11,tooltip=id_or_name)),
                       ft.DataCell(ft.Text(hint_display,size=11,tooltip=hint_tooltip,
                                           color=ft.Colors.ORANGE_700 if not is_visible else None)),
                       ft.DataCell(ft.Text(info_text,size=10,color=info_color,tooltip=info_text)),
                       ft.DataCell(ft.Text(sel_text,size=10,color=ft.Colors.GREY_600,tooltip=sel_text))],
                on_select_change=lambda e, idx=i: on_el_click(idx),
                selected=is_selected,
                color=row_color))
        status_parts = [f"{visible_count}/{total_count} 要素"]
        if hidden_count > 0:
            status_parts.append(f"(非表示: {hidden_count})")
        if query:
            status_parts.append(f"検索: \"{el_search.value}\"")
        state["_prev_el_row"] = _el_idx_to_row.get(state["selected_el"], -1)
        el_status.value = " ".join(status_parts); el_status.color = ft.Colors.GREY_500
        if update:
            try: page.update()
            except Exception: pass
    def show_bank_dlg(e):
        """Show saved selector bank history in a dialog."""
        bank = state["selector_bank"]
        if not bank: snack("保存済みURLなし", ft.Colors.ORANGE_600); return
        def on_select(url):
            clean = url.split("?")[0]
            if clean in bank:
                state["browser_elements"] = list(bank[clean])
                browser_url.value = url
                filter_el_table(); snack(f"バンク {len(bank[clean])} 要素")
            close_dlg(dlg)
        items = ft.Column([
            ft.TextButton(u, on_click=lambda e, u=u: on_select(u), tooltip=f"{len(bank[u])} 要素")
            for u in bank.keys()
        ], spacing=2, scroll=ft.ScrollMode.AUTO, height=300, width=450)
        dlg = ft.AlertDialog(title=ft.Text("保存済みURL"),
            content=items,
            actions=[ft.TextButton("閉じる", on_click=lambda e: close_dlg(dlg))])
        open_dlg(dlg, modal=False)
    def on_el_click(idx):
        state["selected_el"] = idx
        el = state["browser_elements"][idx] if idx < len(state["browser_elements"]) else None
        # 前回選択行の解除 + 新しい行の選択（2行だけ更新）
        prev = state["_prev_el_row"]
        if 0 <= prev < len(el_table.rows):
            el_table.rows[prev].selected = False
        vis_row = _el_idx_to_row.get(idx, -1)
        if 0 <= vis_row < len(el_table.rows):
            el_table.rows[vis_row].selected = True
        state["_prev_el_row"] = vis_row
        # テーブルのみ即時更新（page全体ではなく）
        try: el_table.update()
        except Exception: pass
        # ハイライトはバックグラウンドで実行
        if el and state["browser_driver"]:
            drv = state["browser_driver"]
            sel_str = el["selector"]
            def _highlight_async():
                try:
                    result_json = drv.execute_script(HIGHLIGHT_JS, sel_str)
                    if result_json:
                        import json as _json
                        info = _json.loads(result_json)
                        found = info.get("found", 0)
                        if found == 0:
                            el_status.value = f"セレクタ不一致: {sel_str}"
                        elif found > 1:
                            el_status.value = f"セレクタ {found}件一致（曖昧）: {sel_str}"
                        else:
                            el_status.value = f"一致: {sel_str} ({info.get('tag','')})"
                    try: el_status.update()
                    except Exception: pass
                except Exception:
                    pass
            threading.Thread(target=_highlight_async, daemon=True).start()
    def show_el_detail(e):
        """Show all attributes of the selected element in a dialog."""
        idx = state["selected_el"]
        if idx < 0 or idx >= len(state["browser_elements"]):
            snack("要素を選択してください", ft.Colors.ORANGE_700); return
        el = state["browser_elements"][idx]
        # Fetch all attributes from live DOM via JS
        attrs = {}
        if state["browser_driver"]:
            try:
                attrs = state["browser_driver"].execute_script(
                    "var e=document.querySelector(arguments[0]);"
                    "if(!e)return {};"
                    "var r={};"
                    "for(var i=0;i<e.attributes.length;i++){r[e.attributes[i].name]=e.attributes[i].value;}"
                    "r['__tagName']=e.tagName.toLowerCase();"
                    "r['__textContent']=(e.textContent||'').trim().substring(0,200);"
                    "r['__innerText']=(e.innerText||'').trim().substring(0,200);"
                    "r['__visible']=e.offsetParent!==null||e.getClientRects().length>0?'true':'false';"
                    "var cs=window.getComputedStyle(e);"
                    "r['__display']=cs.display;r['__visibility']=cs.visibility;r['__opacity']=cs.opacity;"
                    "return r;", el["selector"]) or {}
            except Exception:
                pass
        rows = []
        # Show stored info first
        for key in ("selector", "tag", "type", "id", "name", "hint", "visible", "hidden_reason"):
            val = str(el.get(key, ""))
            if val: rows.append(ft.DataRow(cells=[
                ft.DataCell(ft.Text(key, size=11, weight=ft.FontWeight.BOLD, selectable=True)),
                ft.DataCell(ft.Text(val, size=11, selectable=True))]))
        # Show collected metadata
        meta = el.get("meta", {})
        if meta:
            rows.append(ft.DataRow(cells=[ft.DataCell(ft.Text("── テスト情報 ──", size=10, color=ft.Colors.TEAL_600)),
                                          ft.DataCell(ft.Text(""))]))
            meta_labels = {"label": "ラベル", "required": "必須", "maxlength": "最大文字数",
                          "pattern": "入力パターン", "disabled": "無効", "readonly": "読取専用",
                          "option_count": "選択肢数", "href": "リンク先"}
            for mk, mv in meta.items():
                label = meta_labels.get(mk, mk)
                rows.append(ft.DataRow(cells=[
                    ft.DataCell(ft.Text(label, size=11, weight=ft.FontWeight.BOLD, selectable=True,
                                       color=ft.Colors.RED_700 if mk == "required" else None)),
                    ft.DataCell(ft.Text(str(mv), size=11, selectable=True))]))
        # Show live DOM attributes
        if attrs:
            rows.append(ft.DataRow(cells=[ft.DataCell(ft.Text("── DOM属性 ──", size=10, color=ft.Colors.BLUE_600)),
                                          ft.DataCell(ft.Text(""))]))
            for k, v in sorted(attrs.items()):
                if k.startswith("__"): continue
                rows.append(ft.DataRow(cells=[
                    ft.DataCell(ft.Text(k, size=11, weight=ft.FontWeight.BOLD, selectable=True)),
                    ft.DataCell(ft.Text(v[:100], size=11, selectable=True))]))
            rows.append(ft.DataRow(cells=[ft.DataCell(ft.Text("── 計算値 ──", size=10, color=ft.Colors.BLUE_600)),
                                          ft.DataCell(ft.Text(""))]))
            for k in ("__tagName", "__visible", "__display", "__visibility", "__opacity", "__textContent", "__innerText"):
                v = attrs.get(k, "")
                if v: rows.append(ft.DataRow(cells=[
                    ft.DataCell(ft.Text(k.lstrip("_"), size=11, weight=ft.FontWeight.BOLD, selectable=True)),
                    ft.DataCell(ft.Text(v[:100], size=11, selectable=True))]))
        tbl = ft.DataTable(columns=[ft.DataColumn(ft.Text("属性", size=11)), ft.DataColumn(ft.Text("値", size=11))],
            rows=rows, column_spacing=12, data_row_min_height=26, heading_row_height=28)
        dlg = ft.AlertDialog(title=ft.Text(f"要素詳細: {el['selector'][:40]}"),
            content=ft.Container(ft.Column([tbl], scroll=ft.ScrollMode.AUTO), width=500, height=400),
            actions=[ft.TextButton("閉じる", on_click=lambda e: close_dlg(dlg))])
        open_dlg(dlg, modal=False)
    def on_el_search_change(e):
        try: filter_el_table()
        except Exception as x: _log_error("on_el_search_change", x)
    def on_show_hidden_change(e):
        try: filter_el_table()
        except Exception as x: _log_error("on_show_hidden_change", x)
    def _resolve_el_selector(el_info):
        """セレクタモードに応じてCSS or XPathを返す"""
        use_xpath = state.get("_sel_mode") == "xpath"
        if not use_xpath:
            return el_info["selector"]
        # XPathモード: ブラウザからXPath生成
        driver = state["browser_driver"]
        css_sel = el_info["selector"]
        if css_sel.startswith("//") or css_sel.startswith("(/"): return css_sel  # 既にXPath
        if not driver: return css_sel  # ブラウザなしならCSS fallback
        try:
            target = driver.find_element(*_sel_by(css_sel))
            xpath = driver.execute_script(XPATH_JS, target)
            return xpath if xpath else css_sel
        except Exception:
            return css_sel

    def quick_add(stype):
        tc = cur_test()
        if not tc: snack("テストケースを選択してください", ft.Colors.ORANGE_700); return
        idx = state["selected_el"]
        if idx < 0 or idx >= len(state["browser_elements"]): snack("要素を選択してください", ft.Colors.ORANGE_700); return
        el_info = state["browser_elements"][idx]; sel = _resolve_el_selector(el_info)
        tag = el_info.get("tag", ""); etype = el_info.get("type", "").lower()
        actual_type = stype
        if stype != "ホバー":
            if tag == "select": actual_type = "選択"
            elif etype in ("radio", "checkbox"): actual_type = "クリック"
            elif tag in ("button", "a") or etype in ("submit", "button", "reset", "image"): actual_type = "クリック"
        converted = actual_type != stype
        step = {"type": actual_type, "selector": sel}
        if "_frame" in el_info: step["_frame"] = el_info["_frame"]; step["_frame_index"] = el_info.get("_frame_index", 0)
        if actual_type in ("入力", "選択"): step["value"] = ""
        tc["steps"].append(step); refresh_steps(False); refresh_test_list()
        try: step_reorder.scroll_to(offset=-1, duration=200)
        except Exception: pass
        if converted:
            snack(f"要素に合わせて「{actual_type}」に変更: {sel}", ft.Colors.BLUE_600)
        else:
            snack(f"{actual_type}: {sel}")
    def quick_add_all_options(e):
        if not state["browser_driver"]: snack("先にページを読み込んでください", ft.Colors.ORANGE_700); return
        idx = state["selected_el"]
        if idx < 0 or idx >= len(state["browser_elements"]): snack("要素を選択してください", ft.Colors.ORANGE_700); return
        el_info = state["browser_elements"][idx]; tag = el_info.get("tag", ""); etype = el_info.get("type", "").lower()
        if tag != "select" and etype != "radio": snack("セレクトボックスまたはラジオボタンを選択", ft.Colors.ORANGE_700); return
        step_type, options = collect_element_options(state["browser_driver"], el_info)
        if not options: snack("選択肢が取得できませんでした", ft.Colors.RED_700); return
        el_type_name = "セレクトボックス" if tag == "select" else "ラジオボタン"
        sel = el_info.get("selector", ""); el_label = el_info.get("name", "") or el_info.get("hint", "") or sel
        if len(el_label) > 25: el_label = el_label[:22] + "..."
        base_name = f"{el_label} ({el_type_name})"; pat_name = base_name; n = 1
        while pat_name in state["pattern_sets"]: n += 1; pat_name = f"{base_name} {n}"
        tc_name = f"{el_label} 全パターン"
        add_ss = ft.Checkbox(label="選択ごとにスクショを追加", value=True)
        info_text = f"{el_type_name} {sel} から {len(options)} 件検出。\nTC「{tc_name}」とPS「{pat_name}」を作成。"
        def on_ok(ev):
            try:
                state["pattern_sets"][pat_name] = list(options); state["selected_pat_set"] = pat_name
                new_steps = []
                if step_type == "選択": new_steps.append({"type": "選択", "selector": sel, "value": "{パターン}"})
                else: new_steps.append({"type": "クリック", "selector": "{パターン}"})
                if add_ss.value: new_steps.append({"type": "スクショ", "mode": "fullpage"})
                cur_pid = state["selected_page"] or (state["pages"][0]["_id"] if state["pages"] else None)
                new_tc = {"name": tc_name, "pattern": pat_name, "steps": new_steps, "_id": _new_tc_id(),
                          "page_id": cur_pid, "number": ""}
                state["tests"].append(new_tc); auto_number_tests()
                state["selected_test"] = len(state["tests"]) - 1
                refresh_steps(False); refresh_test_list(False); refresh_page_dd(False)
                refresh_pat_set_list(False); refresh_pats()
                snack(f"{el_type_name}の全パターン {len(options)} 件を追加"); close_dlg(dlg)
            except Exception as x: _log_error("quick_add_all_options", x); close_dlg(dlg)
        dlg = ft.AlertDialog(title=ft.Text(f"{el_type_name}の全パターン追加"),
            content=ft.Column([ft.Text(info_text, size=13), ft.Divider(),
                ft.Text("プレビュー:", size=11, weight=ft.FontWeight.BOLD),
                ft.Column([ft.Text(f"  {o['label']}" + (f" = {o['value'][:30]}" if o['value'] != o['label'] else ""),
                    size=11, color=ft.Colors.GREY_600) for o in options[:10]]
                    + ([ft.Text(f"  ... 他 {len(options)-10} 件", size=11, color=ft.Colors.GREY_400)]
                       if len(options) > 10 else []), spacing=2),
                ft.Divider(), add_ss], tight=True, spacing=8, width=450, scroll=ft.ScrollMode.AUTO, height=350),
            actions=[ft.TextButton("作成", on_click=on_ok), ft.TextButton("キャンセル", on_click=lambda ev: close_dlg(dlg))])
        open_dlg(dlg)
    def capture_form(e):
        tc = cur_test()
        if not tc: snack("テストケースを選択してください", ft.Colors.ORANGE_700); return
        if not state["browser_driver"]: snack("先にページを読み込んでください", ft.Colors.ORANGE_700); return
        try:
            fs = capture_form_values(state["browser_driver"])
            if not fs: snack("フォーム値なし", ft.Colors.ORANGE_700); return
            tc["steps"].extend(fs); refresh_steps(False); refresh_test_list(); snack(f"フォーム値 {len(fs)} 件")
        except Exception as x: log(f"[ERROR] {x}")
    def _clipboard_copy(text):
        """クリップボードにテキストをコピー（Windows対応）"""
        try:
            import subprocess
            p = subprocess.Popen(['clip'], stdin=subprocess.PIPE)
            p.communicate(text.encode('utf-16-le'))
        except Exception:
            try: page.set_clipboard(text)
            except Exception: pass

    def copy_el_selector(e):
        """選択中の要素のセレクタをクリップボードにコピー（モードに連動）"""
        idx = state["selected_el"]
        if idx < 0 or idx >= len(state["browser_elements"]): snack("要素を選択してください", ft.Colors.ORANGE_700); return
        el_info = state["browser_elements"][idx]
        sel = _resolve_el_selector(el_info)
        if sel:
            _clipboard_copy(sel)
            snack(f"コピー: {sel[:50]}")

    def copy_el_xpath(e):
        """選択中の要素のXPathをクリップボードにコピー"""
        idx = state["selected_el"]
        if idx < 0 or idx >= len(state["browser_elements"]): snack("要素を選択してください", ft.Colors.ORANGE_700); return
        el = state["browser_elements"][idx]
        sel = el.get("selector", "")
        # XPathセレクタならそのままコピー
        if sel.startswith("//") or sel.startswith("(/"):
            _clipboard_copy(sel)
            snack(f"XPathコピー: {sel[:50]}"); return
        # CSSセレクタからXPathを生成（ブラウザ使用）
        driver = state["browser_driver"]
        if not driver: snack("先にページを読み込んでください", ft.Colors.ORANGE_700); return
        try:
            target = driver.find_element(*_sel_by(sel))
            if target:
                xpath = driver.execute_script(XPATH_JS, target)
                _clipboard_copy(xpath)
                snack(f"XPathコピー: {xpath[:50]}")
            else:
                snack("要素が見つかりません", ft.Colors.RED_700)
        except Exception as x:
            snack(f"XPath取得失敗: {x}", ft.Colors.RED_700)

    def copy_el_id(e):
        """選択中の要素のID/nameをクリップボードにコピー"""
        idx = state["selected_el"]
        if idx < 0 or idx >= len(state["browser_elements"]): snack("要素を選択してください", ft.Colors.ORANGE_700); return
        el = state["browser_elements"][idx]
        val = el.get("id") or el.get("name") or ""
        if val:
            _clipboard_copy(val)
            snack(f"コピー: {val}")
        else:
            snack("この要素にはID/name属性がありません", ft.Colors.ORANGE_700)

    def test_selector_dlg(e):
        """Open a dialog to test CSS/XPath selectors by highlighting matches in the browser."""
        if not state["browser_driver"]: snack("先にページを読み込んでください", ft.Colors.ORANGE_700); return
        # Pre-fill with selected element's selector
        init_sel = ""
        idx = state["selected_el"]
        if 0 <= idx < len(state["browser_elements"]):
            init_sel = state["browser_elements"][idx].get("selector", "")
        tf = ft.TextField(label="セレクタ (CSS / XPath)", width=420, value=init_sel, autofocus=True)
        result_text = ft.Text("", size=12)
        def do_test(e):
            sel_val = (tf.value or "").strip()
            if not sel_val: result_text.value = "セレクタを入力"; result_text.color = ft.Colors.RED_700; page.update(); return
            try:
                matches = state["browser_driver"].find_elements(*_sel_by(sel_val))
                if not matches:
                    result_text.value = f"該当なし"; result_text.color = ft.Colors.RED_700
                else:
                    state["browser_driver"].execute_script(HIGHLIGHT_JS, sel_val)
                    result_text.value = f"一致: {len(matches)} 要素"
                    result_text.color = ft.Colors.GREEN_700 if len(matches) == 1 else ft.Colors.ORANGE_700
            except Exception as x:
                result_text.value = f"エラー: {x}"; result_text.color = ft.Colors.RED_700
            page.update()
        dlg = ft.AlertDialog(title=ft.Text("セレクタテスト"),
            content=ft.Column([tf, ft.OutlinedButton("テスト", icon=ft.Icons.PLAY_ARROW, on_click=do_test), result_text],
                tight=True, spacing=10, width=450),
            actions=[ft.TextButton("閉じる", on_click=lambda e: close_dlg(dlg))])
        open_dlg(dlg, modal=False)
    def close_br(e):
        close_browser(); el_table.rows.clear(); state["browser_elements"].clear()
        el_status.value = "未読込"; el_status.color = ft.Colors.GREY_500; page.update()

    # ================================================================
    # Tab 2: Pattern Sets
    # ================================================================
    def on_ps_search_change(e):
        refresh_pat_set_list()

    def refresh_pat_set_list(update=True):
        pat_set_list.controls.clear()
        ps_query = ""
        try: ps_query = (ps_search.value or "").strip().lower()
        except NameError: pass
        if not state["pattern_sets"]:
            pat_set_list.controls.append(ft.Container(
                ft.Column([ft.Icon(ft.Icons.ADD_CIRCLE_OUTLINE, size=28, color=ft.Colors.GREY_400),
                    ft.Text("＋ボタンでパターンセットを追加", size=11, color=ft.Colors.GREY_500, text_align=ft.TextAlign.CENTER)],
                    horizontal_alignment=ft.CrossAxisAlignment.CENTER, spacing=6),
                padding=ft.Padding(16, 30, 16, 16), key="empty_ps"))
        for name in state["pattern_sets"].keys():
            if ps_query and ps_query not in name.lower(): continue
            pats = state["pattern_sets"][name]; selected = (state["selected_pat_set"] == name)
            pat_set_list.controls.append(ft.Container(
                ft.Row([ft.Column([
                    ft.Text(name, weight=ft.FontWeight.BOLD, size=13,
                            color=ft.Colors.BLUE_800 if selected else ft.Colors.BLACK),
                    ft.Text(f"{len(pats)} パターン", size=10, color=ft.Colors.GREY_500)], spacing=2, expand=True),
                    ft.PopupMenuButton(icon=ft.Icons.MORE_VERT, icon_size=18, icon_color=ft.Colors.GREY_500,
                        tooltip="操作", items=[
                            ft.PopupMenuItem(icon=ft.Icons.EDIT, content="リネーム", on_click=lambda e, n=name: rename_pat_set(n)),
                            ft.PopupMenuItem(),
                            ft.PopupMenuItem(icon=ft.Icons.DELETE, content="削除", on_click=lambda e, n=name: del_pat_set(n))])],
                    vertical_alignment=ft.CrossAxisAlignment.CENTER),
                bgcolor=ft.Colors.BLUE_50 if selected else None,
                padding=ft.Padding(10, 6, 36, 6), border_radius=6,
                border=ft.Border.all(2, ft.Colors.BLUE_300) if selected else ft.Border.all(1, ft.Colors.GREY_200),
                on_click=lambda e, n=name: select_pat_set(n), key=f"ps_{name}"))
        schedule_save()
        if update: page.update()

    def on_pat_set_reorder(e):
        try:
            names = list(state["pattern_sets"].keys())
            old, new = e.old_index, e.new_index
            if old is None or new is None: return
            if _is_dup_reorder("patset", old, new): return
            _flog.debug(f"on_pat_set_reorder: old={old} new={new} len={len(names)}")
            if 0 <= old < len(names) and 0 <= new <= len(names):
                if old == new: return
                item = names.pop(old)
                names.insert(new, item)
                state["pattern_sets"] = {n: state["pattern_sets"][n] for n in names}
                refresh_pat_set_list()
        except Exception as x: _log_error("on_pat_set_reorder", x)

    def select_pat_set(name):
        state["selected_pat_set"] = name; refresh_pat_set_list(False); refresh_pats()

    def add_pat_set(e):
        if _guard_running(): return
        nf = ft.TextField(label="パターンセット名", width=300)
        def on_ok(e):
            try:
                n = nf.value.strip()
                if not n: snack("パターンセット名を入力してください", ft.Colors.RED_700); return
                if n in state["pattern_sets"]: snack("同名のパターンセットが既に存在します", ft.Colors.RED_700); return
                state["pattern_sets"][n] = []; state["selected_pat_set"] = n
                refresh_pat_set_list(False); refresh_pats(); close_dlg(dlg)
            except Exception as x: _log_error("add_pat_set", x); close_dlg(dlg)
        dlg = ft.AlertDialog(title=ft.Text("パターンセット追加"), content=nf,
            actions=[ft.TextButton("OK", on_click=on_ok), ft.TextButton("キャンセル", on_click=lambda e: close_dlg(dlg))])
        open_dlg(dlg)

    def rename_pat_set(old_name):
        if old_name not in state["pattern_sets"]: return
        nf = ft.TextField(label="新しい名前", width=300, value=old_name)
        def on_ok(e):
            try:
                new_name = nf.value.strip()
                if not new_name: snack("パターンセット名を入力してください", ft.Colors.RED_700); return
                if new_name != old_name and new_name in state["pattern_sets"]: snack("同名のパターンセットが既に存在します", ft.Colors.RED_700); return
                if new_name != old_name:
                    new_ps = {}
                    for k, v in state["pattern_sets"].items(): new_ps[new_name if k == old_name else k] = v
                    state["pattern_sets"] = new_ps
                    for tc in state["tests"]:
                        if tc.get("pattern") == old_name: tc["pattern"] = new_name
                    if state["selected_pat_set"] == old_name: state["selected_pat_set"] = new_name
                refresh_pat_set_list(False); refresh_pats(False); refresh_test_list(); close_dlg(dlg)
            except Exception as x: _log_error("rename_pat_set", x); close_dlg(dlg)
        dlg = ft.AlertDialog(title=ft.Text("リネーム"), content=nf,
            actions=[ft.TextButton("OK", on_click=on_ok), ft.TextButton("キャンセル", on_click=lambda e: close_dlg(dlg))])
        open_dlg(dlg)

    def del_pat_set(name):
        if _guard_running(): return
        if name not in state["pattern_sets"]: return
        cnt = len(state["pattern_sets"][name])
        def on_yes(e):
            try:
                del state["pattern_sets"][name]
                if state["selected_pat_set"] == name: state["selected_pat_set"] = None
                refresh_pat_set_list(False); refresh_pats(False); refresh_test_list(); close_dlg(dlg)
            except Exception as x: _log_error("del_pat_set", x); close_dlg(dlg)
        dlg = ft.AlertDialog(title=ft.Text("削除確認"), content=ft.Text(f"「{name}」({cnt}件) を削除しますか？"),
            actions=[ft.TextButton("削除", on_click=on_yes, style=ft.ButtonStyle(color=ft.Colors.RED_600)),
                     ft.TextButton("キャンセル", on_click=lambda e: close_dlg(dlg))])
        open_dlg(dlg)

    def on_pat_reorder(e):
        try:
            name = state["selected_pat_set"]
            if not name or name not in state["pattern_sets"]: return
            pats = state["pattern_sets"][name]
            old, new = e.old_index, e.new_index
            if old is None or new is None: return
            if _is_dup_reorder("pat", old, new): return
            _flog.debug(f"on_pat_reorder: old={old} new={new} len={len(pats)}")
            if 0 <= old < len(pats) and 0 <= new <= len(pats):
                if old == new: return
                item = pats.pop(old)
                pats.insert(new, item)
                refresh_pats()
        except Exception as x: _log_error("on_pat_reorder", x)

    def refresh_pats(update=True):
        pat_items.controls.clear()
        name = state["selected_pat_set"]
        if not name or name not in state["pattern_sets"]:
            pat_header.value = "パターンセットを選択"
            if update: page.update()
            return
        pats = state["pattern_sets"][name]
        pat_header.value = f"{name} ({len(pats)} 件)"
        for i, p in enumerate(pats):
            v = p["value"]; d = v if len(v) <= 55 else v[:52]+"..."
            pat_items.controls.append(ft.Container(ft.Row([
                ft.Text(f"{i+1}", size=10, color=ft.Colors.BLUE_600, weight=ft.FontWeight.BOLD, width=28),
                ft.Column([
                    ft.Text(p["label"], weight=ft.FontWeight.BOLD, size=12, color=ft.Colors.BLUE_800),
                    ft.Text(d, size=11, color=ft.Colors.GREY_600, max_lines=1, overflow=ft.TextOverflow.ELLIPSIS),
                ], spacing=1, expand=True),
                ft.Text(f"{len(v)}字", size=10, color=ft.Colors.GREY_400, width=40),
                ft.IconButton(ft.Icons.COPY, icon_size=14, tooltip="コピー", on_click=lambda e, idx=i: copy_pat(idx)),
                ft.IconButton(ft.Icons.EDIT, icon_size=14, on_click=lambda e, idx=i: edit_pat(idx)),
                ft.IconButton(ft.Icons.DELETE, icon_size=14, icon_color=ft.Colors.RED_400, on_click=lambda e, idx=i: del_pat(idx)),
            ], vertical_alignment=ft.CrossAxisAlignment.CENTER, spacing=4),
                padding=ft.Padding(10, 6, 36, 6), border=ft.Border.all(1, ft.Colors.GREY_200), border_radius=4, key=f"pat_{i}"))
        schedule_save()
        if update: page.update()

    def add_pat(e):
        if _guard_running(): return
        name = state["selected_pat_set"]
        if not name: return
        edit_pat(None)

    def edit_pat(idx):
        if _guard_running(): return
        name = state["selected_pat_set"]
        if not name: return
        pats = state["pattern_sets"][name]
        init = pats[idx] if idx is not None else {}
        lf = ft.TextField(label="ラベル", width=400, value=init.get("label",""))
        vf = ft.TextField(label="入力値", width=400, value=init.get("value",""), multiline=True, min_lines=3, max_lines=6)
        def on_ok(e):
            try:
                if not lf.value: snack("ラベルを入力してください", ft.Colors.RED_700); return
                p = {"label": lf.value, "value": vf.value}
                if idx is not None: pats[idx] = p
                else: pats.append(p)
                refresh_pats(False); refresh_pat_set_list(False); refresh_test_list(); close_dlg(dlg)
            except Exception as x: _log_error("edit_pat", x); close_dlg(dlg)
        dlg = ft.AlertDialog(title=ft.Text("パターン"),
            content=ft.Column([lf, vf], tight=True, spacing=10, width=450),
            actions=[ft.TextButton("OK", on_click=on_ok), ft.TextButton("キャンセル", on_click=lambda e: close_dlg(dlg))])
        open_dlg(dlg)

    def del_pat(idx):
        if _guard_running(): return
        name = state["selected_pat_set"]
        if not name or name not in state["pattern_sets"]: return
        pats = state["pattern_sets"][name]
        if not (0 <= idx < len(pats)): return
        pat = pats[idx]
        label = pat.get("label", f"パターン{idx+1}")
        def on_yes(e):
            pats.pop(idx); refresh_pats(False); refresh_pat_set_list(False); refresh_test_list(); close_dlg(dlg)
        dlg = ft.AlertDialog(title=ft.Text("パターン削除"),
            content=ft.Text(f"「{label}」を削除しますか？"),
            actions=[ft.TextButton("削除", on_click=on_yes, style=ft.ButtonStyle(color=ft.Colors.RED_600)),
                     ft.TextButton("キャンセル", on_click=lambda e: close_dlg(dlg))])
        open_dlg(dlg)

    def copy_pat(idx):
        name = state["selected_pat_set"]
        if not name or name not in state["pattern_sets"]: return
        pats = state["pattern_sets"][name]
        if 0 <= idx < len(pats):
            state["_copied_pat"] = copy.deepcopy(pats[idx])
            snack(f"パターンをコピー: {pats[idx].get('label','')}")

    def paste_pat(e):
        if _guard_running(): return
        name = state["selected_pat_set"]
        if not name or name not in state["pattern_sets"]: snack("パターンセットを選択してください", ft.Colors.ORANGE_700); return
        p = state.get("_copied_pat")
        if not p: snack("コピーされたパターンなし", ft.Colors.ORANGE_700); return
        state["pattern_sets"][name].append(copy.deepcopy(p))
        refresh_pats(False); refresh_pat_set_list(False); refresh_test_list()
        snack(f"パターンを貼り付け: {p.get('label','')}")

    async def export_csv(e):
        name = state["selected_pat_set"]
        if not name or name not in state["pattern_sets"]: snack("パターンセットを選択してください", ft.Colors.ORANGE_700); return
        pats = state["pattern_sets"][name]
        if not pats: snack("パターンなし", ft.Colors.ORANGE_700); return
        default_name = f"{_safe_filename(name, 50)}.csv"
        initial_dir = state["config"].get("output_dir", os.path.join(get_app_dir(), "screenshots"))
        try:
            fp = await _export_file_picker.save_file(
                dialog_title="CSVエクスポート先を選択",
                file_name=default_name,
                initial_directory=initial_dir,
                allowed_extensions=["csv"],
            )
            if not fp: return  # cancelled
            if not fp.lower().endswith(".csv"):
                fp += ".csv"
            _dir = os.path.dirname(fp)
            if _dir: os.makedirs(_dir, exist_ok=True)
            save_csv(fp, pats); snack(f"エクスポート: {fp}")
        except Exception as x:
            _log_error("export_csv", x); snack(f"CSVエクスポート失敗: {x}", ft.Colors.RED_600)

    def load_template(e):
        name = state["selected_pat_set"]
        if not name: snack("パターンセットを選択してください", ft.Colors.ORANGE_700); return
        td = get_templates_dir()
        csvs = sorted([f for f in os.listdir(td) if f.lower().endswith(".csv")]) if td else []
        csv_cache = {f: load_csv(os.path.join(td, f)) for f in csvs} if td else {}
        def on_sel(fn):
            try:
                state["pattern_sets"][name].extend(csv_cache[fn])
                refresh_pats(False); refresh_pat_set_list(False); refresh_test_list()
                snack(f"{len(csv_cache[fn])} 件追加"); close_dlg(dlg)
            except Exception as x: _log_error("load_template", x); close_dlg(dlg)
        def on_import_csv(e):
            try:
                fp = csv_path_f.value.strip()
                if not fp or not os.path.isfile(fp):
                    snack("ファイルが見つかりません", ft.Colors.RED_700); return
                rows = load_csv(fp)
                if not rows: snack("データなし（label,value ヘッダが必要）", ft.Colors.RED_700); return
                state["pattern_sets"][name].extend(rows)
                refresh_pats(False); refresh_pat_set_list(False); refresh_test_list()
                snack(f"{len(rows)} 件インポート"); close_dlg(dlg)
            except Exception as x: _log_error("import_csv", x); snack(f"インポート失敗: {x}", ft.Colors.RED_600)
        csv_path_f = ft.TextField(label="CSVファイルパス", width=380, hint_text="label,value 形式のCSV", dense=True)
        cards = [ft.Card(ft.Container(ft.Column([
            ft.Text(os.path.splitext(f)[0], weight=ft.FontWeight.BOLD, size=13),
            ft.Text(f"{len(csv_cache[f])} 件", size=11, color=ft.Colors.GREY_600)], spacing=2),
            padding=12, on_click=lambda e, fn=f: on_sel(fn)), elevation=2) for f in csvs]
        content_col = ft.Column(spacing=8, scroll=ft.ScrollMode.AUTO, width=420, tight=True)
        content_col.controls.append(ft.Text("CSVファイルから直接インポート:", size=12, weight=ft.FontWeight.BOLD))
        content_col.controls.append(ft.Row([csv_path_f, ft.TextButton("読込", on_click=on_import_csv)], spacing=4))
        if cards:
            content_col.controls.append(ft.Divider())
            content_col.controls.append(ft.Text("テンプレート:", size=12, weight=ft.FontWeight.BOLD))
            content_col.controls.extend(cards)
        dlg = ft.AlertDialog(title=ft.Text("テンプレート / CSVインポート"),
            content=content_col,
            actions=[ft.TextButton("閉じる", on_click=lambda e: close_dlg(dlg))])
        open_dlg(dlg)

    def _ensure_pat_set(ml_hint=""):
        """パターンセット未選択なら自動作成して返す。選択中ならその名前を返す。"""
        name = state["selected_pat_set"]
        if name and name in state["pattern_sets"]:
            return name
        # 自動作成
        auto_name = f"パターンセット（{ml_hint}字）" if ml_hint else f"パターンセット{len(state['pattern_sets'])+1}"
        # 同名があればサフィックス
        base = auto_name; cnt = 2
        while auto_name in state["pattern_sets"]:
            auto_name = f"{base}_{cnt}"; cnt += 1
        state["pattern_sets"][auto_name] = []
        state["selected_pat_set"] = auto_name
        refresh_pat_set_list(False)
        return auto_name

    def gen_input_check(e):
        cf = ft.TextField(label="繰り返す文字", width=80, value="あ")
        mf = ft.TextField(label="max_length", width=140, hint_text="例: 50")
        def on_ok(e):
            try:
                ch = cf.value or "あ"; ml = mf.value.strip()
                if not ml or not ml.isdigit() or int(ml) < 1: snack("max_lengthを正の整数で", ft.Colors.RED_700); return
                n = int(ml)
                name = _ensure_pat_set(ml)
                ps = []
                if n > 1: ps.append({"label": f"max-1({n-1}文字)", "value": ch * (n - 1)})
                ps.append({"label": f"max({n}文字)", "value": ch * n})
                ps.append({"label": f"max+1({n+1}文字)", "value": ch * (n + 1)})
                state["pattern_sets"][name].extend(ps)
                select_pat_set(name)
                refresh_pats(False); refresh_pat_set_list(False); refresh_test_list()
                snack(f"{name} に {len(ps)} 件追加"); close_dlg(dlg)
            except Exception as x: _log_error("gen_input_check", x); close_dlg(dlg)
        dlg = ft.AlertDialog(title=ft.Text("max_length用 境界値生成（文字）"),
            content=ft.Column([cf, mf,
                ft.Text("指定文字を max-1, max, max+1 文字ずつ生成", size=11, color=ft.Colors.GREY_600),
                ft.Text("パターンセット未選択なら自動作成します", size=10, color=ft.Colors.BLUE_400)],
                tight=True, spacing=10, width=350),
            actions=[ft.TextButton("追加", on_click=on_ok), ft.TextButton("閉じる", on_click=lambda e: close_dlg(dlg))])
        open_dlg(dlg)

    def gen_numeric_check(e):
        mf = ft.TextField(label="max_length", width=140, hint_text="例: 10")
        DIGITS = "1234567890"
        def _make_num(length):
            """半角数値文字列を指定長で生成 (1234567890を繰り返し)"""
            if length <= 0: return ""
            return (DIGITS * (length // 10 + 1))[:length]
        def on_ok(e):
            try:
                ml = mf.value.strip()
                if not ml or not ml.isdigit() or int(ml) < 1: snack("max_lengthを正の整数で", ft.Colors.RED_700); return
                n = int(ml)
                name = _ensure_pat_set(ml)
                ps = []
                if n > 1: ps.append({"label": f"数値max-1({n-1}桁)", "value": _make_num(n - 1)})
                ps.append({"label": f"数値max({n}桁)", "value": _make_num(n)})
                ps.append({"label": f"数値max+1({n+1}桁)", "value": _make_num(n + 1)})
                state["pattern_sets"][name].extend(ps)
                select_pat_set(name)
                refresh_pats(False); refresh_pat_set_list(False); refresh_test_list()
                snack(f"{name} に {len(ps)} 件追加"); close_dlg(dlg)
            except Exception as x: _log_error("gen_numeric_check", x); close_dlg(dlg)
        dlg = ft.AlertDialog(title=ft.Text("max_length用 境界値生成（半角数値）"),
            content=ft.Column([mf,
                ft.Text("1234567890 を繰り返して max-1, max, max+1 桁生成", size=11, color=ft.Colors.GREY_600),
                ft.Text("パターンセット未選択なら自動作成します", size=10, color=ft.Colors.BLUE_400)],
                tight=True, spacing=10, width=350),
            actions=[ft.TextButton("追加", on_click=on_ok), ft.TextButton("閉じる", on_click=lambda e: close_dlg(dlg))])
        open_dlg(dlg)

    # ── Settings / Info / Run ──
    def show_settings(e):
        c = state["config"]
        auf = ft.TextField(label="Basic認証ID", value=c.get("basic_auth_user",""), width=210)
        apf = ft.TextField(label="パスワード", value=c.get("basic_auth_pass",""), password=True, width=210)
        of = ft.TextField(label="出力フォルダ", value=c.get("output_dir", os.path.join(get_app_dir(), "screenshots")), width=450)
        hl = ft.Checkbox(label="ヘッドレスモード (ブラウザ非表示)", value=c.get("headless")=="1")
        ss = ft.Checkbox(label="HTMLソース保存 (diff比較用)", value=c.get("save_source")=="1")
        csd = ft.Checkbox(label="ステップ削除時に確認する", value=c.get("confirm_step_delete","1")=="1")
        def on_ok(e):
            try:
                state["config"].update({"basic_auth_user":auf.value,"basic_auth_pass":apf.value,
                    "output_dir":of.value,"headless":"1" if hl.value else "0",
                    "save_source":"1" if ss.value else "0",
                    "confirm_step_delete":"1" if csd.value else "0"})
                state["config"].pop("url", None)  # 旧グローバルURL設定を除去
                save_config(state["config"]); snack("設定保存")
                refresh_test_list(False); page.update(); close_dlg(dlg)
            except Exception as x: _log_error("show_settings", x); close_dlg(dlg)
        dlg = ft.AlertDialog(title=ft.Text("設定"),
            content=ft.Column([ft.Row([auf, apf], spacing=10), of, hl, ss, csd], tight=True, spacing=12, width=500),
            actions=[ft.TextButton("OK", on_click=on_ok), ft.TextButton("キャンセル", on_click=lambda e: close_dlg(dlg))])
        open_dlg(dlg)

    # ── Project Export / Import ──
    async def export_project(e):
        """Export pages, tests, pattern sets, config, and selector bank as a single JSON project file."""
        try:
            save_all()
            proj_name = _current_project_name() or "project"
            ts = datetime.now().strftime("%Y%m%d%H%M%S")
            default_name = f"{_safe_filename(proj_name, 30)}_{ts}.yshot.json"
            initial_dir = state["config"].get("output_dir", os.path.join(get_app_dir(), "screenshots"))
            fp = await _export_file_picker.save_file(
                dialog_title="プロジェクトエクスポート先を選択",
                file_name=default_name,
                initial_directory=initial_dir,
                allowed_extensions=["json"],
            )
            if not fp: return  # cancelled
            if not fp.endswith(".yshot.json"):
                fp += ".yshot.json"
            project_data = {
                "app": APP_NAME, "version": APP_VERSION,
                "project_name": proj_name,
                "pages": state["pages"],
                "tests": state["tests"],
                "pattern_sets": state["pattern_sets"],
                "config": state["config"],
                "selector_bank": state["selector_bank"],
            }
            _dir = os.path.dirname(fp)
            if _dir: os.makedirs(_dir, exist_ok=True)
            with open(fp, "w", encoding="utf-8") as f:
                json.dump(project_data, f, ensure_ascii=False, indent=2)
            snack(f"エクスポート: {fp}")
            log(f"[プロジェクト] エクスポート: {fp}")
        except Exception as x:
            _log_error("export_project", x); snack(f"エクスポート失敗: {x}", ft.Colors.RED_600)

    async def import_project(e):
        """Import a .yshot.json project file via OS file picker + confirmation dialog."""
        initial_dir = state["config"].get("output_dir", os.path.join(get_app_dir(), "screenshots"))
        _import_picker = ft.FilePicker()
        picked = await _import_picker.pick_files(
            dialog_title="インポートするプロジェクトファイルを選択",
            initial_directory=initial_dir,
            allowed_extensions=["json"],
            allow_multiple=False,
        )
        if not picked: return
        fp = picked[0].path

        # Read and validate
        try:
            with open(fp, "r", encoding="utf-8") as f:
                data = json.load(f)
        except json.JSONDecodeError:
            snack("JSONパースエラー", ft.Colors.RED_600); return
        except Exception as x:
            snack(f"ファイル読込失敗: {x}", ft.Colors.RED_600); return
        if "pages" not in data or "tests" not in data:
            snack("無効なプロジェクトファイルです", ft.Colors.RED_700); return

        # Build preview
        n_pages = len(data.get("pages", []))
        n_tests = len(data.get("tests", []))
        n_pats = len(data.get("pattern_sets", {}))
        ver = data.get("version", "?")
        pname = data.get("project_name", "")
        has_config = "設定あり" if data.get("config") else "設定なし"
        proj_url = data.get("config", {}).get("project_url", "")
        urls = [pg.get("url","") for pg in data.get("pages",[]) if pg.get("url","")]
        url_hint = proj_url[:40] if proj_url else (urls[0][:40] if urls else "URL未設定")
        if len(url_hint) >= 40: url_hint += "..."
        name_hint = f"「{pname}」 " if pname else ""
        preview_str = f"{name_hint}v{ver} | {n_pages}ページ, {n_tests}テスト, {n_pats}パターン, {has_config} | {url_hint}"

        mode_dd = ft.Dropdown(label="インポート方法", width=350, value="new_project",
            options=[ft.dropdown.Option(key="replace", text="置換（現在のデータを上書き）"),
                     ft.dropdown.Option(key="merge", text="マージ（現在のデータに追加）"),
                     ft.dropdown.Option(key="new_project", text="新規プロジェクトとしてインポート")])

        def on_ok(ev):
            try:
                imp_pages = data.get("pages", [])
                imp_tests = data.get("tests", [])
                imp_pats = data.get("pattern_sets", {})
                imp_config = {k: v for k, v in data.get("config", {}).items() if k != "output_dir"}
                imp_selbank = data.get("selector_bank", {})
                # Sanitize: strip \r\n from names (common with Windows CRLF data)
                for _t in imp_tests:
                    if "name" in _t: _t["name"] = _t["name"].replace("\r","").replace("\n","").strip()
                for _p in imp_pages:
                    if "name" in _p: _p["name"] = _p["name"].replace("\r","").replace("\n","").strip()

                if mode_dd.value == "new_project":
                    # Create a new project and import data into it
                    save_all()
                    proj_name = data.get("project_name") or os.path.basename(fp).replace(".yshot.json", "") or "Imported"
                    new_id = _new_project_id(_projects_registry)
                    dir_name = _safe_dir_name(proj_name)
                    base_dir = dir_name; counter = 1
                    while os.path.isdir(os.path.join(get_projects_dir(), dir_name)):
                        counter += 1; dir_name = f"{base_dir}_{counter}"
                    _projects_registry["projects"].append({
                        "id": new_id, "name": proj_name, "dir": dir_name,
                        "created": datetime.now().isoformat()
                    })
                    activate_project(new_id, _projects_registry)
                    save_projects_registry(_projects_registry)
                    state["pages"] = imp_pages
                    state["tests"] = imp_tests
                    state["pattern_sets"] = imp_pats
                    state["selector_bank"] = imp_selbank
                    # 新規プロジェクトでは旧configをクリアし、output_dirだけ現マシンの値を保持
                    kept_output_dir = state["config"].get("output_dir", "")
                    state["config"] = dict(imp_config)
                    if kept_output_dir:
                        state["config"]["output_dir"] = kept_output_dir
                    _reinit_id_counters()
                    _ensure_default_page()
                    auto_number_tests()
                    state["selected_page"] = state["pages"][0]["_id"]
                    state["selected_test"] = -1
                    state["selected_pat_set"] = None
                    save_all()
                    project_dd.options = _project_dd_options()
                    project_dd.value = new_id
                    refresh_page_dd(False); refresh_test_list(False); refresh_steps(False)
                    refresh_pat_set_list(False); refresh_pats(False)
                    page.title = f"{APP_NAME} - {_current_project_name()}"
                    page.update()
                    snack(f"新規プロジェクト「{proj_name}」にインポート完了 ({len(imp_pages)}ページ, {len(imp_tests)}テスト)")
                    log(f"[プロジェクト] 新規プロジェクトインポート: {fp} -> {proj_name}")
                    close_dlg(dlg); return

                if mode_dd.value == "replace":
                    state["pages"] = imp_pages
                    state["tests"] = imp_tests
                    state["pattern_sets"] = imp_pats
                    if "config" in data:
                        state["config"].update(imp_config)
                    if "selector_bank" in data:
                        state["selector_bank"] = imp_selbank
                else:
                    # Merge: remap IDs and page numbers to avoid collision
                    _reinit_id_counters()  # カウンタを現状に合わせてからID生成する
                    # 既存ページの最大番号を取得して続き番号を振る
                    max_page_num = 0
                    for pg in state["pages"]:
                        try: max_page_num = max(max_page_num, int(pg.get("number", 0)))
                        except (ValueError, TypeError): pass
                    page_id_map = {}
                    for i, pg in enumerate(imp_pages):
                        old_id = pg["_id"]
                        new_id = _new_page_id()
                        pg["_id"] = new_id
                        pg["number"] = str(max_page_num + 1 + i)
                        page_id_map[old_id] = new_id
                    for tc in imp_tests:
                        tc["_id"] = _new_tc_id()
                        old_pid = tc.get("page_id", "")
                        if old_pid in page_id_map:
                            tc["page_id"] = page_id_map[old_pid]
                    state["pages"].extend(imp_pages)
                    state["tests"].extend(imp_tests)
                    pat_name_map = {}  # old_name -> new_name (only for renamed sets)
                    for k, v in imp_pats.items():
                        if k in state["pattern_sets"]:
                            new_k = f"{k}_imported"
                            n = 2
                            while new_k in state["pattern_sets"]:
                                new_k = f"{k}_imported_{n}"; n += 1
                            state["pattern_sets"][new_k] = v
                            pat_name_map[k] = new_k
                        else:
                            state["pattern_sets"][k] = v
                    # Remap pattern references in imported tests
                    if pat_name_map:
                        for tc in imp_tests:
                            old_pat = tc.get("pattern")
                            if old_pat in pat_name_map:
                                tc["pattern"] = pat_name_map[old_pat]
                    # Merge selector bank (import側で上書きしない、既存を優先)
                    if imp_selbank:
                        for url, sels in imp_selbank.items():
                            if url not in state["selector_bank"]:
                                state["selector_bank"][url] = sels

                _reinit_id_counters()
                _ensure_default_page()
                state["selected_page"] = state["pages"][0]["_id"]
                state["selected_test"] = -1
                state["selected_pat_set"] = None
                auto_number_tests()
                save_all()
                refresh_page_dd(False); refresh_test_list(False); refresh_steps(False)
                refresh_pat_set_list(False); refresh_pats(False); page.update()
                src_ver = data.get("version", "?")
                snack(f"インポート完了 ({len(imp_pages)}ページ, {len(imp_tests)}テスト)")
                log(f"[プロジェクト] インポート: {fp} (v{src_ver}, {mode_dd.value})")
                close_dlg(dlg)
            except Exception as x:
                _log_error("import_project", x); snack(f"インポート失敗: {x}", ft.Colors.RED_600)

        dlg = ft.AlertDialog(title=ft.Text("プロジェクトインポート"),
            content=ft.Column([
                ft.Text(os.path.basename(fp), size=12, weight=ft.FontWeight.BOLD),
                ft.Text(preview_str, size=11, color=ft.Colors.GREY_600),
                mode_dd,
            ], tight=True, spacing=10, width=400),
            actions=[ft.TextButton("インポート", on_click=on_ok),
                     ft.TextButton("キャンセル", on_click=lambda e: close_dlg(dlg))])
        open_dlg(dlg)

    def show_info(e):
        dlg = ft.AlertDialog(title=ft.Text("情報"),
            content=ft.Column([ft.Text(f"{APP_NAME}  v{APP_VERSION}", size=20, weight=ft.FontWeight.BOLD),
                ft.Text("Web Screenshot Automation Tool"), ft.Divider(),
                ft.Text(f"Developed by {APP_AUTHOR}")], tight=True, spacing=8, width=300),
            actions=[ft.TextButton("閉じる", on_click=lambda e: close_dlg(dlg))])
        open_dlg(dlg, modal=False)

    def _do_run(test_cases_to_run, run_label=""):
        c = state["config"]
        if not test_cases_to_run: snack("テストケース0件", ft.Colors.RED_700); return
        # URL pre-check: project URL が設定されていればスキップ
        _proj_url = c.get("project_url", "").strip()
        no_url_tests = []
        if not _proj_url:
            _page_url_map = {pg["_id"]: pg.get("url","").strip() for pg in state["pages"]}
            for tc in test_cases_to_run:
                tc_url = tc.get("url","").strip()
                if not tc_url and not _page_url_map.get(tc.get("page_id",""), ""):
                    no_url_tests.append(tc)
            if len(no_url_tests) == len(test_cases_to_run):
                snack("起点URLが未設定です（プロジェクト設定またはページ編集でURLを入力してください）", ft.Colors.RED_700); return
        if no_url_tests:
            names = "\n".join(f"  - {tc.get('number','')} {tc.get('name','')}" for tc in no_url_tests[:10])
            if len(no_url_tests) > 10: names += f"\n  ... 他 {len(no_url_tests)-10} 件"
            def on_continue(e):
                close_dlg(warn_dlg)
                _do_run_execute(test_cases_to_run, run_label)
            warn_dlg = ft.AlertDialog(title=ft.Text("URL未設定のテストケースがあります"),
                content=ft.Column([
                    ft.Text(f"{len(no_url_tests)} 件のテストはURL未設定のためスキップされます:", size=12),
                    ft.Text(names, size=11, font_family="Consolas"),
                ], tight=True, spacing=8, width=450),
                actions=[ft.TextButton("続行", on_click=on_continue),
                         ft.TextButton("キャンセル", on_click=lambda e: close_dlg(warn_dlg))])
            open_dlg(warn_dlg); return
        _do_run_execute(test_cases_to_run, run_label)

    def _do_run_execute(test_cases_to_run, run_label=""):
        c = state["config"]
        close_browser()
        state["running"] = True
        stop_ev = threading.Event(); state["stop_event"] = stop_ev
        run_btn.disabled = True; run_single_btn.disabled = True; run_page_btn.disabled = True
        run_btn.icon = ft.Icons.HOURGLASS_TOP; run_single_btn.icon = ft.Icons.HOURGLASS_TOP; run_page_btn.icon = ft.Icons.HOURGLASS_TOP
        stop_btn.visible = True; stop_btn.disabled = False; open_folder_btn.visible = False
        run_spinner.visible = True; progress.visible = True; progress.value = 0; progress_label.visible = True; progress_label.value = ""
        nav_bar.selected_index = 0; switch_tab(0); page.update(); save_all()
        def on_progress(current, total, tc_label=""):
            progress.value = current / total if total > 0 else None
            label = f"{current}/{total} パターン"
            if tc_label: label += f" — {tc_label}"
            progress_label.value = label
            try: progress.update(); progress_label.update()
            except Exception:
                try: page.update()
                except Exception: pass
        def on_done(outdir=None):
            state["running"] = False
            run_btn.disabled = False; run_single_btn.disabled = False; run_page_btn.disabled = False
            run_btn.icon = ft.Icons.PLAY_ARROW; run_single_btn.icon = ft.Icons.PLAY_ARROW; run_page_btn.icon = ft.Icons.PLAY_ARROW
            stop_btn.visible = False; run_spinner.visible = False; progress.visible = False; progress_label.visible = False
            state["stop_event"] = None
            if outdir and os.path.isdir(outdir):
                open_folder_btn.data = outdir; open_folder_btn.visible = True
            page.update()
        page.run_thread(run_all_tests, dict(c), list(test_cases_to_run),
                        dict(state["pattern_sets"]), lambda m: log(m), on_done, stop_ev, on_progress,
                        state["test_drivers"], list(state["pages"]), run_label, _current_project_name())

    def run_click(e):
        tcs = state["tests"]
        if not tcs: snack("テストケース0件", ft.Colors.RED_700); return
        n_pats = sum(len(state["pattern_sets"].get(tc.get("pattern","") or "", [])) or 1 for tc in tcs)
        def on_yes(e):
            close_dlg(dlg); _do_run(tcs, "【全テスト】")
        dlg = ft.AlertDialog(title=ft.Text("全テスト実行"),
            content=ft.Text(f"全 {len(tcs)} テスト（{n_pats}パターン）を実行しますか？"),
            actions=[ft.TextButton("実行", on_click=on_yes), ft.TextButton("キャンセル", on_click=lambda e: close_dlg(dlg))])
        open_dlg(dlg)
    def run_single(idx):
        if not (0 <= idx < len(state["tests"])):
            snack("テストケースを選択してください", ft.Colors.ORANGE_700); return
        tc = state["tests"][idx]
        label = f"【{tc.get('number','')}_{tc.get('name','')}】"
        _do_run([tc], label)
    def stop_click(e):
        ev = state.get("stop_event")
        if ev: ev.set(); stop_btn.disabled = True; log("[中断] 中断リクエスト送信..."); page.update()

    def switch_tab(idx):
        tc_content.visible = (idx == 0); ps_content.visible = (idx == 1)
        if idx == 0: refresh_page_dd(False); refresh_test_list(False); refresh_steps(False)
        else: refresh_pat_set_list(False); refresh_pats(False)
        page.update()
    def on_nav(e): switch_tab(e.control.selected_index)

    # ── Build controls ──
    _init_browser_url = cfg.get("browser_url","")
    if not _init_browser_url:
        _init_browser_url = state["config"].get("project_url", "").strip()
    if not _init_browser_url:
        _init_pg = cur_page()
        if _init_pg: _init_browser_url = _init_pg.get("url", "")
    browser_url = ft.TextField(label="URL", expand=True, dense=True, value=_init_browser_url)
    browser_wait = ft.TextField(label="秒", width=55, dense=True, value=cfg.get("browser_wait","3.0"))
    load_btn = ft.Button("読込", icon=ft.Icons.DOWNLOAD, on_click=load_page_click)
    el_loading = ft.ProgressRing(width=14, height=14, stroke_width=2, visible=False)
    el_status = ft.Text("未読込", size=11, color=ft.Colors.GREY_500)
    def _on_sel_mode_change(e):
        state["_sel_mode"] = e.control.value
    sel_mode_radio = ft.RadioGroup(content=ft.Row([
        ft.Radio(value="css", label="CSS", label_style=ft.TextStyle(size=10)),
        ft.Radio(value="xpath", label="XPath", label_style=ft.TextStyle(size=10)),
    ], spacing=0), value="css", on_change=_on_sel_mode_change)
    el_search = ft.TextField(label="検索", expand=True, dense=True, hint_text="セレクタ/id/name/ヒント",
                             on_change=on_el_search_change, prefix_icon=ft.Icons.SEARCH)
    el_show_hidden = ft.Checkbox(label="非表示", value=False, on_change=on_show_hidden_change)
    el_sort_dd = ft.PopupMenuButton(
        icon=ft.Icons.SORT, icon_size=18, tooltip="並び替え",
        items=[
            ft.PopupMenuItem(content="DOM順", on_click=lambda e: on_el_sort_change("dom")),
            ft.PopupMenuItem(content="タグ別", on_click=lambda e: on_el_sort_change("tag")),
            ft.PopupMenuItem(content="type別", on_click=lambda e: on_el_sort_change("type")),
            ft.PopupMenuItem(content="id/name別", on_click=lambda e: on_el_sort_change("id")),
        ])
    el_sort_dd.data = "dom"
    el_table = ft.DataTable(
        columns=[ft.DataColumn(ft.Text("タグ",size=11)), ft.DataColumn(ft.Text("type",size=11)),
                 ft.DataColumn(ft.Text("id/name",size=11)), ft.DataColumn(ft.Text("ヒント",size=11)),
                 ft.DataColumn(ft.Text("ラベル/検証",size=11)), ft.DataColumn(ft.Text("セレクタ",size=11))],
        rows=[], column_spacing=8, data_row_min_height=28, heading_row_height=30,
        show_checkbox_column=True)

    # ── Project selector ──
    def _project_dd_options():
        return [ft.dropdown.Option(key=p["id"], text=p["name"])
                for p in _projects_registry["projects"]]

    def _current_project_name():
        for p in _projects_registry["projects"]:
            if p["id"] == _projects_registry.get("last_active"): return p["name"]
        return ""

    def _reload_project_data():
        """Reload all data from current project directory."""
        state["tests"] = _safe_load(load_tests, [], "tests")
        state["pages"] = _safe_load(load_pages, [], "pages")
        state["pattern_sets"] = _safe_load(load_pattern_sets, {}, "pattern_sets")
        state["selector_bank"] = _safe_load(load_selector_bank, {}, "selector_bank")
        state["config"] = _safe_load(load_config, {}, "config")
        _reinit_id_counters()
        _ensure_default_page()
        _invalidate_idx()
        auto_number_tests()
        state["selected_page"] = state["pages"][0]["_id"]
        state["selected_test"] = -1
        state["selected_pat_set"] = None
        state["collapsed"] = set()
        state["selected_test_per_page"] = {}

    def on_project_change(e):
        if _guard_running(): return
        if not project_dd.value: return
        if project_dd.value == _projects_registry.get("last_active"): return
        save_all()
        activate_project(project_dd.value, _projects_registry)
        save_projects_registry(_projects_registry)
        _reload_project_data()
        refresh_page_dd(False); refresh_test_list(False); refresh_steps(False)
        refresh_pat_set_list(False); refresh_pats(False)
        page.title = f"{APP_NAME} - {_current_project_name()}"
        page.update()

    def add_project(e):
        if _guard_running(): return
        nf = ft.TextField(label="プロジェクト名", width=350, autofocus=True)
        def on_ok(e):
            name = nf.value.strip()
            if not name: snack("名前を入力してください", ft.Colors.RED_700); return
            save_all()
            new_id = _new_project_id(_projects_registry)
            dir_name = _safe_dir_name(name)
            # Ensure unique directory name
            base_dir = dir_name
            counter = 1
            while os.path.isdir(os.path.join(get_projects_dir(), dir_name)):
                counter += 1; dir_name = f"{base_dir}_{counter}"
            _projects_registry["projects"].append({
                "id": new_id, "name": name, "dir": dir_name,
                "created": datetime.now().isoformat()
            })
            activate_project(new_id, _projects_registry)
            save_projects_registry(_projects_registry)
            _reload_project_data()
            project_dd.options = _project_dd_options()
            project_dd.value = new_id
            refresh_page_dd(False); refresh_test_list(False); refresh_steps(False)
            refresh_pat_set_list(False); refresh_pats(False)
            page.title = f"{APP_NAME} - {_current_project_name()}"
            page.update(); close_dlg(dlg)
        dlg = ft.AlertDialog(title=ft.Text("プロジェクト追加"),
            content=ft.Column([nf], tight=True, spacing=10, width=400),
            actions=[ft.TextButton("OK", on_click=on_ok),
                     ft.TextButton("キャンセル", on_click=lambda e: close_dlg(dlg))])
        open_dlg(dlg)

    def rename_project(e):
        cur_id = _projects_registry.get("last_active")
        cur_proj = None
        for p in _projects_registry["projects"]:
            if p["id"] == cur_id: cur_proj = p; break
        if not cur_proj: return
        nf = ft.TextField(label="プロジェクト名", width=400, value=cur_proj["name"], autofocus=True)
        uf = ft.TextField(label="プロジェクトURL (設定時は全ページで優先)", width=400,
                          value=state["config"].get("project_url", ""),
                          hint_text="空欄の場合はページごとのURLを使用")
        def on_ok(e):
            name = nf.value.strip()
            if not name: snack("名前を入力してください", ft.Colors.RED_700); return
            cur_proj["name"] = name
            save_projects_registry(_projects_registry)
            state["config"]["project_url"] = uf.value.strip()
            save_config(state["config"])
            project_dd.options = _project_dd_options()
            page.title = f"{APP_NAME} - {_current_project_name()}"
            refresh_test_list(False); page.update(); close_dlg(dlg)
        dlg = ft.AlertDialog(title=ft.Text("プロジェクト設定"),
            content=ft.Column([nf, uf], tight=True, spacing=10, width=450),
            actions=[ft.TextButton("OK", on_click=on_ok),
                     ft.TextButton("キャンセル", on_click=lambda e: close_dlg(dlg))])
        open_dlg(dlg)

    def del_project(e):
        if _guard_running(): return
        cur_id = _projects_registry.get("last_active")
        if cur_id == "default":
            snack("デフォルトプロジェクトは削除できません", ft.Colors.RED_700); return
        cur_proj = None
        for p in _projects_registry["projects"]:
            if p["id"] == cur_id: cur_proj = p; break
        if not cur_proj: return
        def on_ok(e):
            close_dlg(dlg)
            proj_dir = os.path.join(get_projects_dir(), cur_proj["dir"])
            _projects_registry["projects"].remove(cur_proj)
            activate_project("default", _projects_registry)
            save_projects_registry(_projects_registry)
            # Remove project directory
            try: shutil.rmtree(proj_dir)
            except Exception: pass
            _reload_project_data()
            project_dd.options = _project_dd_options()
            project_dd.value = "default"
            refresh_page_dd(False); refresh_test_list(False); refresh_steps(False)
            refresh_pat_set_list(False); refresh_pats(False)
            page.title = f"{APP_NAME} - {_current_project_name()}"
            page.update()
        dlg = ft.AlertDialog(title=ft.Text("プロジェクト削除"),
            content=ft.Text(f"「{cur_proj['name']}」を削除しますか？\nすべてのページ・テストが削除されます。"),
            actions=[ft.TextButton("削除", on_click=on_ok, style=ft.ButtonStyle(color=ft.Colors.RED_700)),
                     ft.TextButton("キャンセル", on_click=lambda e: close_dlg(dlg))])
        open_dlg(dlg)

    project_dd = ft.Dropdown(label="プロジェクト", expand=True, dense=True,
                             options=_project_dd_options(),
                             value=_projects_registry.get("last_active"),
                             on_select=on_project_change)

    # Page selector
    page_dd = ft.Dropdown(label="ページ", expand=True, dense=True,
                          options=_page_dd_options(), value=state["selected_page"],
                          on_select=on_page_dd_change)
    page_info_label = ft.Text("", size=10, color=ft.Colors.GREY_500)

    test_list = ft.ReorderableListView(controls=[], on_reorder=on_test_reorder, spacing=3, expand=True, show_default_drag_handles=False)
    step_reorder = ft.ReorderableListView(controls=[], on_reorder=on_reorder, spacing=1, expand=True)
    log_list = ft.ListView(spacing=1, auto_scroll=True, height=130)
    _log_expanded = [False]
    def toggle_log(e):
        _log_expanded[0] = not _log_expanded[0]
        log_list.height = 350 if _log_expanded[0] else 130
        log_toggle_btn.icon = ft.Icons.EXPAND_LESS if _log_expanded[0] else ft.Icons.EXPAND_MORE
        log_toggle_btn.tooltip = "ログを縮小" if _log_expanded[0] else "ログを拡大"; page.update()
    log_toggle_btn = ft.IconButton(ft.Icons.EXPAND_MORE, icon_size=16, tooltip="ログを拡大", on_click=toggle_log)
    tc_header = ft.Text("", weight=ft.FontWeight.BOLD, size=15)
    tc_pattern_label = ft.Text("", size=11, color=ft.Colors.GREY_600)
    ps_search = ft.TextField(label="検索", width=250, dense=True, hint_text="パターンセット名",
                             on_change=on_ps_search_change, prefix_icon=ft.Icons.SEARCH)
    pat_set_list = ft.ReorderableListView(controls=[], on_reorder=on_pat_set_reorder, spacing=4, expand=True)
    pat_items = ft.ReorderableListView(controls=[], on_reorder=on_pat_reorder, spacing=3, expand=True)
    pat_header = ft.Text("", weight=ft.FontWeight.BOLD, size=15)
    run_spinner = ft.ProgressRing(width=16, height=16, stroke_width=2, visible=False)
    progress = ft.ProgressBar(visible=False, value=0)
    progress_label = ft.Text("", size=11, color=ft.Colors.GREY_600, visible=False)
    run_single_btn = ft.Button("選択テスト実行", icon=ft.Icons.PLAY_ARROW, bgcolor=ft.Colors.GREEN_600,
                               color=ft.Colors.WHITE, on_click=lambda e: run_single(state["selected_test"]), height=42)
    run_page_btn = ft.Button("ページ実行", icon=ft.Icons.PLAY_ARROW, bgcolor=ft.Colors.TEAL_600,
                             color=ft.Colors.WHITE, on_click=lambda e: run_page(), height=42)
    run_btn = ft.Button("全テスト実行", icon=ft.Icons.PLAY_ARROW, bgcolor=ft.Colors.BLUE_600,
                        color=ft.Colors.WHITE, on_click=run_click, height=42)
    stop_btn = ft.Button("中断", icon=ft.Icons.STOP, bgcolor=ft.Colors.RED_600,
                         color=ft.Colors.WHITE, on_click=stop_click, height=42, visible=False)
    def open_folder_click(e):
        path = open_folder_btn.data
        if path and os.path.isdir(path):
            import subprocess; subprocess.Popen(["explorer", os.path.normpath(path)])
    open_folder_btn = ft.Button("出力フォルダを開く", icon=ft.Icons.FOLDER_OPEN,
                                on_click=open_folder_click, height=42, visible=False)

    # ── Layout: Tab 1 (collapsible test list) ──
    _tc_panel_collapsed = [False]
    tc_panel_full = ft.Column([
        ft.Row([project_dd,
                ft.IconButton(ft.Icons.ADD, tooltip="プロジェクト追加", icon_size=16, icon_color=ft.Colors.GREY_700, style=ft.ButtonStyle(padding=4), on_click=add_project),
                ft.PopupMenuButton(icon=ft.Icons.MORE_VERT, icon_size=16, icon_color=ft.Colors.GREY_700,
                    tooltip="プロジェクト操作", items=[
                        ft.PopupMenuItem(icon=ft.Icons.EDIT, content="プロジェクト設定", on_click=rename_project),
                        ft.PopupMenuItem(icon=ft.Icons.DELETE, content="削除", on_click=del_project),
                    ]),
               ], spacing=0, vertical_alignment=ft.CrossAxisAlignment.CENTER),
        ft.Divider(height=1),
        ft.Row([page_dd,
                ft.IconButton(ft.Icons.ADD, tooltip="ページ追加", icon_size=16, icon_color=ft.Colors.GREY_700, style=ft.ButtonStyle(padding=4), on_click=add_page),
                ft.PopupMenuButton(icon=ft.Icons.MORE_VERT, icon_size=16, icon_color=ft.Colors.GREY_700,
                    tooltip="ページ操作", items=[
                        ft.PopupMenuItem(icon=ft.Icons.EDIT, content="編集", on_click=edit_page),
                        ft.PopupMenuItem(icon=ft.Icons.COPY, content="複製", on_click=dup_page),
                        ft.PopupMenuItem(),
                        ft.PopupMenuItem(icon=ft.Icons.ARROW_UPWARD, content="一つ上へ", on_click=move_page_up),
                        ft.PopupMenuItem(icon=ft.Icons.ARROW_DOWNWARD, content="一つ下へ", on_click=move_page_down),
                        ft.PopupMenuItem(),
                        ft.PopupMenuItem(icon=ft.Icons.DELETE, content="削除", on_click=del_page),
                    ]),
               ], spacing=0, vertical_alignment=ft.CrossAxisAlignment.CENTER),
        page_info_label,
        ft.Divider(height=1),
        ft.Row([ft.Text("テストケース", weight=ft.FontWeight.BOLD, size=13),
                ft.IconButton(ft.Icons.ADD, tooltip="テスト追加", icon_size=16, on_click=add_test)],
               alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
        test_list,
    ], spacing=4, expand=True)
    tc_panel_mini = ft.Column([
        ft.Text("TC", size=10, color=ft.Colors.GREY_500, text_align=ft.TextAlign.CENTER),
    ], spacing=4, horizontal_alignment=ft.CrossAxisAlignment.CENTER, visible=False)
    def toggle_tc_panel(e):
        _tc_panel_collapsed[0] = not _tc_panel_collapsed[0]
        collapsed = _tc_panel_collapsed[0]
        tc_panel_full.visible = not collapsed
        tc_panel_mini.visible = collapsed
        tc_panel_container.width = 40 if collapsed else 320
        tc_collapse_btn.icon = ft.Icons.CHEVRON_RIGHT if collapsed else ft.Icons.CHEVRON_LEFT
        tc_collapse_btn.tooltip = "テスト一覧を展開" if collapsed else "テスト一覧を折りたたむ"
        page.update()
    tc_collapse_btn = ft.IconButton(ft.Icons.CHEVRON_LEFT, icon_size=16, tooltip="テスト一覧を折りたたむ",
                                     on_click=toggle_tc_panel, icon_color=ft.Colors.GREY_500, style=ft.ButtonStyle(padding=2))
    tc_panel_container = ft.Container(ft.Column([
        ft.Row([tc_collapse_btn], alignment=ft.MainAxisAlignment.END),
        tc_panel_full, tc_panel_mini,
    ], spacing=0, expand=True), width=320, padding=8, border=ft.Border.all(1, ft.Colors.GREY_300), border_radius=8)

    tc_content = ft.Row([
        tc_panel_container,
        ft.Column([
            ft.Container(ft.Column([
                ft.Row([tc_header, ft.IconButton(ft.Icons.EDIT, icon_size=16, tooltip="テスト設定", on_click=edit_test_name)],
                       alignment=ft.MainAxisAlignment.START),
                tc_pattern_label,
                ft.Row([ft.IconButton(ft.Icons.ADD, tooltip="ステップ追加", icon_size=18, on_click=lambda e: show_step_dlg(None)),
                        ft.IconButton(ft.Icons.PASTE, tooltip="ステップ貼り付け", icon_size=18, on_click=paste_step),
                        ft.IconButton(ft.Icons.TITLE, tooltip="見出し", icon_size=18,
                            on_click=lambda e: (cur_test() and cur_test()["steps"].append({"type":"見出し","text":"セクション"}), refresh_steps())),
                        ft.IconButton(ft.Icons.COMMENT, tooltip="コメント", icon_size=18,
                            on_click=lambda e: (cur_test() and cur_test()["steps"].append({"type":"コメント","text":""}), refresh_steps()))], spacing=0),
                step_reorder,
            ], spacing=4), padding=8, border=ft.Border.all(1, ft.Colors.GREY_300), border_radius=8, expand=True),
            ft.Container(ft.Column([
                ft.Row([ft.Text("ログ", size=12, weight=ft.FontWeight.BOLD), log_toggle_btn],
                       alignment=ft.MainAxisAlignment.SPACE_BETWEEN), log_list]),
                padding=8, border=ft.Border.all(1, ft.Colors.GREY_300), border_radius=8),
        ], expand=3, spacing=6),
        ft.Container(ft.Column([
            ft.Text("要素ブラウザ", weight=ft.FontWeight.BOLD, size=13),
            ft.Row([browser_url, browser_wait,
                    ft.IconButton(ft.Icons.HISTORY, tooltip="保存済みURL", icon_size=18, on_click=show_bank_dlg)], spacing=4),
            ft.Row([load_btn, ft.OutlinedButton("DOM再取得", icon=ft.Icons.REFRESH, on_click=reload_dom_click),
                    ft.OutlinedButton("閉じる", on_click=close_br)], spacing=4, wrap=True),
            ft.Row([el_search, el_sort_dd, el_show_hidden], spacing=4, vertical_alignment=ft.CrossAxisAlignment.CENTER),
            ft.Row([el_loading, el_status], spacing=6, vertical_alignment=ft.CrossAxisAlignment.CENTER),
            ft.Container(ft.Column([ft.Row([el_table], scroll=ft.ScrollMode.AUTO)], scroll=ft.ScrollMode.AUTO),
                expand=True, border=ft.Border.all(1, ft.Colors.GREY_200), border_radius=4),
            ft.Row([
                    sel_mode_radio,
                    ft.VerticalDivider(width=1),
                    ft.IconButton(ft.Icons.EDIT, tooltip="入力", icon_size=18, on_click=lambda e: quick_add("入力")),
                    ft.IconButton(ft.Icons.MOUSE, tooltip="クリック", icon_size=18, on_click=lambda e: quick_add("クリック")),
                    ft.IconButton(ft.Icons.NEAR_ME, tooltip="ホバー", icon_size=18, on_click=lambda e: quick_add("ホバー")),
                    ft.IconButton(ft.Icons.ARROW_DROP_DOWN_CIRCLE, tooltip="選択", icon_size=18, on_click=lambda e: quick_add("選択")),
                    ft.VerticalDivider(width=1),
                    ft.IconButton(ft.Icons.LIST, tooltip="全パターン", icon_size=18, on_click=quick_add_all_options),
                    ft.IconButton(ft.Icons.SAVE_ALT, tooltip="値取込", icon_size=18, on_click=capture_form),
                    ft.VerticalDivider(width=1),
                    ft.IconButton(ft.Icons.SEARCH, tooltip="セレクタテスト", icon_size=18, on_click=test_selector_dlg),
                    ft.IconButton(ft.Icons.INFO_OUTLINE, tooltip="要素詳細", icon_size=18, on_click=show_el_detail),
                    ft.VerticalDivider(width=1),
                    ft.IconButton(ft.Icons.CONTENT_COPY, tooltip="セレクタをコピー", icon_size=18, on_click=copy_el_selector),
                    ft.IconButton(ft.Icons.CODE, tooltip="XPathをコピー", icon_size=18, on_click=copy_el_xpath),
                   ], spacing=0, wrap=True, vertical_alignment=ft.CrossAxisAlignment.CENTER),
        ], spacing=4), expand=3, padding=8, border=ft.Border.all(1, ft.Colors.GREY_300), border_radius=8),
    ], spacing=8, expand=True, vertical_alignment=ft.CrossAxisAlignment.STRETCH)

    # ── Layout: Tab 2 ──
    ps_content = ft.Row([
        ft.Container(ft.Column([
            ft.Row([ft.Text("パターンセット", weight=ft.FontWeight.BOLD, size=14),
                    ft.IconButton(ft.Icons.ADD, tooltip="追加", icon_size=18, on_click=add_pat_set)],
                   alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
            ps_search,
            pat_set_list,
        ], spacing=4, expand=True), width=320, padding=8, border=ft.Border.all(1, ft.Colors.GREY_300), border_radius=8),
        ft.Column([
            ft.Row([pat_header,
                    ft.Row([ft.Button("追加", icon=ft.Icons.ADD, on_click=add_pat),
                            ft.Button("貼り付け", icon=ft.Icons.PASTE, on_click=paste_pat),
                            ft.Button("テンプレート", icon=ft.Icons.FOLDER_OPEN, on_click=load_template),
                            ft.Button("文字max", icon=ft.Icons.STRAIGHTEN, on_click=gen_input_check, tooltip="max_length境界値(文字)"),
                            ft.Button("数値max", icon=ft.Icons.PIN, on_click=gen_numeric_check, tooltip="max_length境界値(半角数値)"),
                            ft.Button("CSVエクスポート", icon=ft.Icons.FILE_UPLOAD, on_click=export_csv)], spacing=4)],
                   alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
            ft.Container(pat_items, expand=True, padding=ft.Padding(4,4,4,4),
                border=ft.Border.all(1, ft.Colors.GREY_200), border_radius=6),
        ], expand=True, spacing=6),
    ], spacing=8, expand=True, visible=False, vertical_alignment=ft.CrossAxisAlignment.STRETCH)

    nav_bar = ft.NavigationBar(
        destinations=[ft.NavigationBarDestination(icon=ft.Icons.LIST_ALT, label="テストケース"),
                      ft.NavigationBarDestination(icon=ft.Icons.DATASET, label="パターンセット")],
        selected_index=0, on_change=on_nav)

    # ── File picker for export dialogs ──
    _export_file_picker = ft.FilePicker()

    page.appbar = ft.AppBar(title=ft.Text(APP_NAME, weight=ft.FontWeight.BOLD), center_title=False,
        bgcolor=ft.Colors.BLUE_50,
        actions=[ft.IconButton(ft.Icons.FILE_UPLOAD, tooltip="プロジェクトエクスポート", on_click=export_project),
                 ft.IconButton(ft.Icons.FILE_DOWNLOAD, tooltip="プロジェクトインポート", on_click=import_project),
                 ft.IconButton(ft.Icons.SETTINGS, tooltip="設定", on_click=show_settings),
                 ft.IconButton(ft.Icons.INFO_OUTLINE, tooltip="情報", on_click=show_info)])

    page.add(ft.Column([ft.Stack([tc_content, ps_content], expand=True),
        ft.Row([run_spinner, progress, progress_label], spacing=8, vertical_alignment=ft.CrossAxisAlignment.CENTER),
        ft.Row([open_folder_btn, stop_btn, run_single_btn, run_page_btn, run_btn],
               alignment=ft.MainAxisAlignment.END, spacing=8)], expand=True, spacing=4))
    page.navigation_bar = nav_bar

    # ── Keyboard shortcuts ──
    def on_keyboard(e: ft.KeyboardEvent):
        if not _init_done[0]: return
        key = e.key; ctrl = e.ctrl or e.meta
        if ctrl and key.lower() == "s":
            save_all(); snack("保存しました")
        elif ctrl and key.lower() == "n":
            if nav_bar.selected_index == 0:
                add_test(None)
            else:
                add_pat_set(None)
        elif key == "Delete":
            if nav_bar.selected_index == 0:
                idx = state["selected_test"]
                if 0 <= idx < len(state["tests"]): del_test(idx)
            else:
                name = state["selected_pat_set"]
                if name: del_pat_set(name)
    page.on_keyboard_event = on_keyboard

    refresh_page_dd(False); refresh_test_list(False); refresh_steps(False)
    refresh_pat_set_list(False); refresh_pats(False)
    page.title = f"{APP_NAME} - {_current_project_name()}"
    page.update()
    _init_done[0] = True
    _flog.info(f"{APP_NAME} v{APP_VERSION} started (project={_current_project_name()}, {len(state['pages'])} pages, {len(state['tests'])} tests, {len(state['pattern_sets'])} pattern sets)")

    def _cleanup_all_drivers():
        """Best-effort quit of Selenium drivers (non-blocking)."""
        for drv in list(state.get("test_drivers", [])):
            try: drv.quit()
            except Exception: pass
        state["test_drivers"] = []
        if state["browser_driver"]:
            try: state["browser_driver"].quit()
            except Exception: pass
            state["browser_driver"] = None

    def _kill_children_and_exit():
        """Kill all child processes (Flet runtime, WebView) but NOT this process.
        Then sys.exit(0) so PyInstaller's atexit cleanup can delete _MEIxxxxxx."""
        my_pid = os.getpid()
        _flog.info(f"Killing child processes of PID={my_pid}")
        if sys.platform == 'win32':
            import subprocess as _sp
            child_pids = []
            # Method 1: wmic (Win10, some Win11)
            try:
                result = _sp.run(
                    ['wmic', 'process', 'where', f'ParentProcessId={my_pid}', 'get', 'ProcessId'],
                    capture_output=True, text=True, timeout=3, creationflags=WIN_CREATE_NO_WINDOW)
                for line in result.stdout.strip().split('\n'):
                    cpid = line.strip()
                    if cpid.isdigit() and cpid != str(my_pid):
                        child_pids.append(cpid)
            except Exception:
                pass
            # Method 2: PowerShell (Win11 where wmic is removed)
            if not child_pids:
                try:
                    ps_cmd = f"Get-CimInstance Win32_Process -Filter 'ParentProcessId={my_pid}' | Select-Object -ExpandProperty ProcessId"
                    result = _sp.run(['powershell', '-NoProfile', '-Command', ps_cmd],
                        capture_output=True, text=True, timeout=5, creationflags=WIN_CREATE_NO_WINDOW)
                    for line in result.stdout.strip().split('\n'):
                        cpid = line.strip()
                        if cpid.isdigit() and cpid != str(my_pid):
                            child_pids.append(cpid)
                except Exception:
                    pass
            # Kill found children
            for cpid in child_pids:
                try:
                    _sp.run(['taskkill', '/F', '/T', '/PID', cpid],
                            capture_output=True, timeout=3, creationflags=WIN_CREATE_NO_WINDOW)
                except Exception:
                    pass
            if child_pids:
                time.sleep(0.2)
            _flog.info(f"Killed {len(child_pids)} child processes")
        else:
            try:
                import signal as _sig
                os.killpg(os.getpgid(my_pid), _sig.SIGTERM)
            except Exception:
                pass
        # Normal exit → PyInstaller atexit cleanup runs → _MEI folder deleted
        _deadline = threading.Timer(1.0, lambda: os._exit(0))
        _deadline.daemon = True; _deadline.start()
        sys.exit(0)

    def on_window_event(e: ft.WindowEvent):
        if e.type == ft.WindowEventType.CLOSE:
            # 1. Save data (fast)
            if _save_timer[0]: _save_timer[0].cancel()
            with _save_lock:
                try: save_all()
                except Exception: pass
            # 2. Signal any running tests to stop
            ev = state.get("stop_event")
            if ev: ev.set()
            # 3. Best-effort Selenium cleanup
            _cleanup_all_drivers()
            # 4. Kill Flet children, then clean exit
            _kill_children_and_exit()

    import signal
    def _signal_cleanup(signum, frame):
        try: save_all()
        except Exception: pass
        _kill_children_and_exit()
    try:
        signal.signal(signal.SIGTERM, _signal_cleanup)
        signal.signal(signal.SIGINT, _signal_cleanup)
    except (OSError, ValueError): pass

    page.window.prevent_close = True
    page.window.on_event = on_window_event

if __name__ == "__main__":
    ft.run(main)
